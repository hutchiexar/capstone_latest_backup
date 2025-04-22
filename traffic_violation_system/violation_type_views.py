from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q
from django.urls import reverse
from .models import ViolationType, Violation
from .utils import log_activity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'SUPERVISOR'])
def violation_types(request):
    """
    View for managing violation types.
    Shows a list of all violation types with filtering options.
    """
    # Get all violation types
    violation_types_list = ViolationType.objects.all()
    
    # Get filter parameters
    category = request.GET.get('category', '')
    ncap = request.GET.get('ncap', '')
    search_query = request.GET.get('search', '')
    
    # Apply filters
    if category:
        violation_types_list = violation_types_list.filter(category=category)
    
    if ncap == 'ncap':
        violation_types_list = violation_types_list.filter(is_ncap=True)
    elif ncap == 'regular':
        violation_types_list = violation_types_list.filter(is_ncap=False)
        
    if search_query:
        violation_types_list = violation_types_list.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    
    # Order by category and name
    violation_types_list = violation_types_list.order_by('category', 'name')
    
    # Pagination
    paginator = Paginator(violation_types_list, 10)  # Show 10 types per page
    page = request.GET.get('page', 1)
    
    try:
        violation_types_page = paginator.page(page)
    except PageNotAnInteger:
        violation_types_page = paginator.page(1)
    except EmptyPage:
        violation_types_page = paginator.page(paginator.num_pages)
    
    # Prepare context
    context = {
        'violation_types': violation_types_page,
        'selected_category': category,
        'selected_ncap': ncap,
        'search_query': search_query,
        'categories': ViolationType.objects.values_list('category', flat=True).distinct(),
    }
    
    return render(request, 'violations/violation_types.html', context)

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'SUPERVISOR'])
def add_violation_type(request):
    """
    View for adding a new violation type.
    """
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            amount = request.POST.get('amount')
            category = request.POST.get('category')
            is_active = 'is_active' in request.POST
            is_ncap = 'is_ncap' in request.POST
            
            # Validate required fields
            if not name or not amount or not category:
                messages.error(request, "Please fill all required fields")
                return redirect('violation_types')
                
            # Check for duplicate name
            if ViolationType.objects.filter(name=name).exists():
                messages.error(request, f"Violation type '{name}' already exists")
                return redirect('violation_types')
                
            # Create the violation type
            ViolationType.objects.create(
                name=name,
                description=description,
                amount=amount,
                category=category,
                is_active=is_active,
                is_ncap=is_ncap
            )
            
            # Log the activity
            log_activity(
                request.user, 
                f"Added new violation type: {name}", 
                "VIOLATION_TYPE_ADDED"
            )
            
            return redirect(f"{reverse('violation_types')}?success=1&action=add")
            
        except Exception as e:
            messages.error(request, f"Error adding violation type: {str(e)}")
            return redirect('violation_types')
    else:
        # GET request, redirect to violation types list
        return redirect('violation_types')

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'SUPERVISOR'])
def edit_violation_type(request, type_id):
    """
    View for editing an existing violation type.
    """
    violation_type = get_object_or_404(ViolationType, id=type_id)
    
    if request.method == 'POST':
        try:
            violation_type.name = request.POST.get('name')
            violation_type.description = request.POST.get('description', '')
            violation_type.amount = request.POST.get('amount')
            violation_type.category = request.POST.get('category')
            violation_type.is_active = 'is_active' in request.POST
            violation_type.is_ncap = 'is_ncap' in request.POST
            
            # Save the changes
            violation_type.save()
            
            # Log the activity
            log_activity(
                request.user, 
                f"Updated violation type: {violation_type.name}", 
                "VIOLATION_TYPE_UPDATED"
            )
            
            return redirect(f"{reverse('violation_types')}?success=1&action=edit")
            
        except Exception as e:
            messages.error(request, f"Error updating violation type: {str(e)}")
            return redirect('violation_types')
    else:
        # GET request, redirect to violation types list
        return redirect('violation_types')

@login_required
@user_passes_test(lambda u: u.userprofile.role == 'ADMIN')
def delete_violation_type(request, type_id):
    """
    View for deleting a violation type.
    """
    violation_type = get_object_or_404(ViolationType, id=type_id)
    
    if request.method == 'POST':
        try:
            name = violation_type.name
            
            # Check if this violation type is used in any violations
            if Violation.objects.filter(violation_type_obj=violation_type).exists():
                messages.error(request, f"Cannot delete violation type '{name}' as it is used in existing violations")
                return redirect('violation_types')
            
            # Delete the violation type
            violation_type.delete()
            
            # Log the activity
            log_activity(
                request.user, 
                f"Deleted violation type: {name}", 
                "VIOLATION_TYPE_DELETED"
            )
            
            return redirect(f"{reverse('violation_types')}?success=1&action=delete")
            
        except Exception as e:
            messages.error(request, f"Error deleting violation type: {str(e)}")
            return redirect('violation_types')
    else:
        # GET request, redirect to violation types list
        return redirect('violation_types')

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'SUPERVISOR'])
def export_violation_types(request):
    """
    View for exporting violation types to Excel.
    """
    # Filter parameters
    category = request.GET.get('category', '')
    ncap = request.GET.get('ncap', '')
    
    # Query violation types with filters
    violation_types = ViolationType.objects.all()
    
    if category:
        violation_types = violation_types.filter(category=category)
    
    if ncap == 'ncap':
        violation_types = violation_types.filter(is_ncap=True)
    elif ncap == 'regular':
        violation_types = violation_types.filter(is_ncap=False)
    
    # Order by category and name
    violation_types = violation_types.order_by('category', 'name')
    
    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Violation Types"
    
    # Add header row
    headers = ["Name", "Category", "Amount (₱)", "Status", "NCAP", "Description"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    for row_num, violation_type in enumerate(violation_types, 2):
        worksheet.cell(row=row_num, column=1).value = violation_type.name
        worksheet.cell(row=row_num, column=2).value = violation_type.get_category_display()
        worksheet.cell(row=row_num, column=3).value = float(violation_type.amount)
        worksheet.cell(row=row_num, column=4).value = "Active" if violation_type.is_active else "Inactive"
        worksheet.cell(row=row_num, column=5).value = "Yes" if violation_type.is_ncap else "No"
        worksheet.cell(row=row_num, column=6).value = violation_type.description
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=violation_types.xlsx'
    
    # Log the activity
    log_activity(
        request.user, 
        "Exported violation types to Excel", 
        "VIOLATION_TYPES_EXPORTED"
    )
    
    # Save to response
    workbook.save(response)
    return response 