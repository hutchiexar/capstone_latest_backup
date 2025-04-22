from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now
from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now

# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                        
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path


@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


@login_required
def edit_profile(request):
    # Get or create UserProfile
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'enforcer_id': generate_enforcer_id(),
            'phone_number': '',
            'address': ''
        }
    )

    if request.method == 'POST':
        try:
            user = request.user

            # Update user information
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            # Update profile information
            profile.phone_number = request.POST.get('phone_number')
            profile.address = request.POST.get('address')

            if 'avatar' in request.FILES:
                # Add debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Uploading avatar for user {user.username}")
                logger.info(f"Avatar file name: {request.FILES['avatar'].name}")
                logger.info(f"Using MEDIA_ROOT: {settings.MEDIA_ROOT}")
                
                profile.avatar = request.FILES['avatar']
                logger.info(f"Avatar saved to: {profile.avatar.path if profile.avatar else 'None'}")

            profile.save()

            # Store current page in session
            request.session['current_page'] = 'profile'
            
            # Add success message
            messages.success(request, 'Profile updated successfully!')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Updated Profile',
                details='Profile information was updated',
                category='user'
            )

            return redirect('profile')
        except Exception as e:
            # Log detailed exception info
            import logging, traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Profile update error: {str(e)}")
            logger.error(traceback.format_exc())
            
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    return render(request, 'edit_profile.html', {'user': request.user})


@login_required
def users_list(request):
    user_profiles = UserProfile.objects.all().select_related('user').order_by('user__date_joined')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if search_query:
        user_profiles = user_profiles.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(enforcer_id__icontains=search_query)
        )

    if role_filter:
        user_profiles = user_profiles.filter(role=role_filter)

    if status_filter:
        is_active = status_filter == 'active'
        user_profiles = user_profiles.filter(user__is_active=is_active)

    # Pagination - 10 users per page
    paginator = Paginator(user_profiles, 10)
    page = request.GET.get('page', 1)
    try:
        user_profiles = paginator.page(page)
    except PageNotAnInteger:
        user_profiles = paginator.page(1)
    except EmptyPage:
        user_profiles = paginator.page(paginator.num_pages)
    
    context = {
        'user_profiles': user_profiles,
        'page_obj': user_profiles,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': paginator.count
    }
    return render(request, 'users/users_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        try:
            # Create user
            user = User.objects.create_user(
                username=request.POST['username'],
                password=request.POST['password'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Instead of creating a new profile, update the one created by the signal
            profile = user.userprofile  # This exists because of the signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Created User',
                details=f'Created user account for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User created successfully!')
            # Correctly add query parameter to redirect
            from django.urls import reverse
            return redirect(reverse('users_list') + '?success=created')
        except Exception as e:
            # If there's an error, delete the user to maintain consistency
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error creating user: {str(e)}')
    
    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/create_user.html', {'roles': roles})


@login_required
def user_detail(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    return render(request, 'users/user_detail.html', {'profile': profile})


@login_required
def edit_user(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    
    if request.method == 'POST':
        try:
            user = profile.user
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.save()

            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Updated User',
                details=f'Updated user profile for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User profile updated successfully!')
            return redirect('users_list')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')

    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/edit_user.html', {
        'profile': profile,
        'roles': roles
    })


@login_required
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            profile = get_object_or_404(UserProfile, id=user_id)
            user = profile.user
            
            # Toggle the user's active status
            user.is_active = not user.is_active
            user.save()
            
            # Add success message
            messages.success(request, f'User {user.username} has been {"activated" if user.is_active else "deactivated"} successfully.')
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'message': f'User {user.username} has been {"activated" if user.is_active else "deactivated"}'
            })
        except Exception as e:
            messages.error(request, f'Error changing user status: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def activity_history(request):
    activities = ActivityLog.objects.select_related('user', 'user__userprofile').all()

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        activities = activities.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(user__userprofile__enforcer_id__icontains=search_query)
        ).distinct()

    # Get filter parameters
    category = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_id = request.GET.get('user_id', '')

    # Apply filters
    if category:
        activities = activities.filter(category=category)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            activities = activities.filter(timestamp__gte=date_from)
        except ValueError:
            pass

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to = date_to + timedelta(days=1)
            activities = activities.filter(timestamp__lt=date_to)
        except ValueError:
            pass

    if user_id:
        activities = activities.filter(user_id=user_id)

    # Order by timestamp descending (most recent first)
    activities = activities.order_by('-timestamp')

    # Pagination - 25 activities per page
    paginator = Paginator(activities, 25)
    page = request.GET.get('page', 1)
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)

    # Get all categories and users for filters
    categories = ActivityLog.CATEGORY_CHOICES
    users = User.objects.filter(
        id__in=activities.paginator.object_list.values_list('user_id', flat=True).distinct()
    ).select_related('userprofile')

    context = {
        'activities': activities,
        'categories': categories,
        'users': users,
        'current_category': category,
        'current_user': user_id,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': activities,
        'search_query': search_query,
        'total_count': paginator.count
    }

    return render(request, 'activity_history.html', context)


@login_required
def get_statistics(request, time_range):
    """API endpoint for chart data"""
    today = timezone.now()
    
    if time_range == 'week':
        start_date = today - timedelta(days=7)
        date_format = '%Y-%m-%d'
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
        date_format = '%Y-%m-%d'
    else:  # year
        start_date = today - timedelta(days=365)
        date_format = '%Y-%m'

    # Get violations data
    violations = Violation.objects.filter(
        violation_date__gte=start_date
    ).annotate(
        date=TruncDate('violation_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')

    # Format data for charts
    labels = []
    data = []
    
    # Create a date range for all dates in the period
    date_range = []
    delta = today - start_date
    for i in range(delta.days + 1):
        date_range.append(start_date + timedelta(days=i))
    
    # Create a lookup of existing data
    data_lookup = {item['date']: item['count'] for item in violations}
    
    # Fill in data for all dates
    for date in date_range:
        labels.append(date.strftime(date_format))
        data.append(data_lookup.get(date.date(), 0))

    violations_chart = {
        'labels': labels,
        'datasets': [{
            'label': 'Total Violations',
            'data': data,
            'borderColor': '#2563eb',
            'backgroundColor': 'rgba(37, 99, 235, 0.1)',
            'tension': 0.4,
            'fill': True
        }]
    }

    return JsonResponse({
        'violations': violations_chart
    })


from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now
from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now

# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path


@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


@login_required
def edit_profile(request):
    # Get or create UserProfile
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'enforcer_id': generate_enforcer_id(),
            'phone_number': '',
            'address': ''
        }
    )

    if request.method == 'POST':
        try:
            user = request.user

            # Update user information
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            # Update profile information
            profile.phone_number = request.POST.get('phone_number')
            profile.address = request.POST.get('address')

            if 'avatar' in request.FILES:
                # Add debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Uploading avatar for user {user.username}")
                logger.info(f"Avatar file name: {request.FILES['avatar'].name}")
                logger.info(f"Using MEDIA_ROOT: {settings.MEDIA_ROOT}")
                
                profile.avatar = request.FILES['avatar']
                logger.info(f"Avatar saved to: {profile.avatar.path if profile.avatar else 'None'}")

            profile.save()

            # Store current page in session
            request.session['current_page'] = 'profile'
            
            # Add success message
            messages.success(request, 'Profile updated successfully!')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Updated Profile',
                details='Profile information was updated',
                category='user'
            )

            return redirect('profile')
        except Exception as e:
            # Log detailed exception info
            import logging, traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Profile update error: {str(e)}")
            logger.error(traceback.format_exc())
            
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    return render(request, 'edit_profile.html', {'user': request.user})


@login_required
def users_list(request):
    user_profiles = UserProfile.objects.all().select_related('user').order_by('user__date_joined')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if search_query:
        user_profiles = user_profiles.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(enforcer_id__icontains=search_query)
        )

    if role_filter:
        user_profiles = user_profiles.filter(role=role_filter)

    if status_filter:
        is_active = status_filter == 'active'
        user_profiles = user_profiles.filter(user__is_active=is_active)

    # Pagination - 10 users per page
    paginator = Paginator(user_profiles, 10)
    page = request.GET.get('page', 1)
    try:
        user_profiles = paginator.page(page)
    except PageNotAnInteger:
        user_profiles = paginator.page(1)
    except EmptyPage:
        user_profiles = paginator.page(paginator.num_pages)
    
    context = {
        'user_profiles': user_profiles,
        'page_obj': user_profiles,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': paginator.count
    }
    return render(request, 'users/users_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        try:
            # Create user
            user = User.objects.create_user(
                username=request.POST['username'],
                password=request.POST['password'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Instead of creating a new profile, update the one created by the signal
            profile = user.userprofile  # This exists because of the signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Created User',
                details=f'Created user account for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User created successfully!')
            # Correctly add query parameter to redirect
            from django.urls import reverse
            return redirect(reverse('users_list') + '?success=created')
        except Exception as e:
            # If there's an error, delete the user to maintain consistency
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error creating user: {str(e)}')
    
    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/create_user.html', {'roles': roles})


@login_required
def user_detail(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    return render(request, 'users/user_detail.html', {'profile': profile})


@login_required
def edit_user(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    
    if request.method == 'POST':
        try:
            user = profile.user
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.save()

            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Updated User',
                details=f'Updated user profile for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User profile updated successfully!')
            return redirect('users_list')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')

    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/edit_user.html', {
        'profile': profile,
        'roles': roles
    })


@login_required
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            profile = get_object_or_404(UserProfile, id=user_id)
            user = profile.user
            
            # Toggle the user's active status
            user.is_active = not user.is_active
            user.save()
            
            # Add success message
            messages.success(request, f'User {user.username} has been {"activated" if user.is_active else "deactivated"} successfully.')
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'message': f'User {user.username} has been {"activated" if user.is_active else "deactivated"}'
            })
        except Exception as e:
            messages.error(request, f'Error changing user status: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def activity_history(request):
    activities = ActivityLog.objects.select_related('user', 'user__userprofile').all()

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        activities = activities.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(user__userprofile__enforcer_id__icontains=search_query)
        ).distinct()

    # Get filter parameters
    category = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_id = request.GET.get('user_id', '')

    # Apply filters
    if category:
        activities = activities.filter(category=category)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            activities = activities.filter(timestamp__gte=date_from)
        except ValueError:
            pass

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to = date_to + timedelta(days=1)
            activities = activities.filter(timestamp__lt=date_to)
        except ValueError:
            pass

    if user_id:
        activities = activities.filter(user_id=user_id)

    # Order by timestamp descending (most recent first)
    activities = activities.order_by('-timestamp')

    # Pagination - 25 activities per page
    paginator = Paginator(activities, 25)
    page = request.GET.get('page', 1)
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)

    # Get all categories and users for filters
    categories = ActivityLog.CATEGORY_CHOICES
    users = User.objects.filter(
        id__in=activities.paginator.object_list.values_list('user_id', flat=True).distinct()
    ).select_related('userprofile')

    context = {
        'activities': activities,
        'categories': categories,
        'users': users,
        'current_category': category,
        'current_user': user_id,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': activities,
        'search_query': search_query,
        'total_count': paginator.count
    }

    return render(request, 'activity_history.html', context)


@login_required
def get_statistics(request, time_range):
    """API endpoint for chart data"""
    today = timezone.now()
    
    if time_range == 'week':
        start_date = today - timedelta(days=7)
        date_format = '%Y-%m-%d'
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
        date_format = '%Y-%m-%d'
    else:  # year
        start_date = today - timedelta(days=365)
        date_format = '%Y-%m'

    # Get violations data
    violations = Violation.objects.filter(
        violation_date__gte=start_date
    ).annotate(
        date=TruncDate('violation_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')

    # Format data for charts
    labels = []
    data = []
    
    # Create a date range for all dates in the period
    date_range = []
    delta = today - start_date
    for i in range(delta.days + 1):
        date_range.append(start_date + timedelta(days=i))
    
    # Create a lookup of existing data
    data_lookup = {item['date']: item['count'] for item in violations}
    
    # Fill in data for all dates
    for date in date_range:
        labels.append(date.strftime(date_format))
        data.append(data_lookup.get(date.date(), 0))

    violations_chart = {
        'labels': labels,
        'datasets': [{
            'label': 'Total Violations',
            'data': data,
            'borderColor': '#2563eb',
            'backgroundColor': 'rgba(37, 99, 235, 0.1)',
            'tension': 0.4,
            'fill': True
        }]
    }

    return JsonResponse({
        'violations': violations_chart
    })


@login_required
def violation_detail_modal(request, violation_id):
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Check if it's an NCAP violation by checking for any images
        is_ncap = bool(violation.image) or bool(violation.driver_photo) or bool(violation.vehicle_photo) or bool(violation.secondary_photo)
        
        # Get previous violations for the same violator
        previous_violations = Violation.objects.filter(
            violator=violation.violator
        ).exclude(
            id=violation.id
        ).order_by('-violation_date')
        
        # Find violations from same name violators
        same_name_violations = []
        if violation.violator:
            try:
                # Find violations with the same first and last name
                same_name_violations = Violation.objects.filter(
                    violator__first_name__iexact=violation.violator.first_name,
                    violator__last_name__iexact=violation.violator.last_name
                ).exclude(
                    violator=violation.violator  # Exclude the current violator as they're already in previous_violations
                ).order_by('-violation_date')
            except Exception as e:
                print(f"Error finding same name violations: {e}")
        
        # Find violations from the same submission
        submission_violations = []
        if is_ncap:
            try:
                # First, try to find violations with the same submission_id
                if violation.submission_id:
                    submission_violations = list(Violation.objects.filter(
                        submission_id=violation.submission_id
                    ).order_by('id'))
                    
                    print(f"Found {len(submission_violations)} violations with submission_id {violation.submission_id}")
                
                # If no submission_id or no other violations found with it, fall back to time window approach
                if not submission_violations:
                    # Find violations submitted within a 5-minute window of this violation
                    if violation.violation_date:
                        time_window_start = violation.violation_date - timezone.timedelta(minutes=5)
                        time_window_end = violation.violation_date + timezone.timedelta(minutes=5)
                        
                        # Find all violations from the same time window by the same violator
                        submission_violations = list(Violation.objects.filter(
                            violator=violation.violator,
                            violation_date__gte=time_window_start,
                            violation_date__lte=time_window_end
                        ).order_by('id'))
                        
                        print(f"Fallback: Found {len(submission_violations)} violations within 5 min window for violator {violation.violator.id}")
                
                # Debug each violation found
                for idx, v in enumerate(submission_violations):
                    print(f"  Violation #{idx+1}: ID={v.id}, Type={v.violation_type}, Same as current={v.id == int(violation_id)}")
            except Exception as e:
                print(f"Error finding related violations: {e}")
                # In case of error, just use the current violation
                submission_violations = [violation]
        
        # Use the appropriate template
        template_name = 'violations/ncap_violation_detail_modal.html' if is_ncap else 'violations/violation_detail_modal.html'
        
        context = {
            'violation': violation,
            'is_modal': True,
            'previous_violations': previous_violations,
            'same_name_violations': same_name_violations,
            'submission_violations': submission_violations
        }
        
        return render(request, template_name, context)
    except Exception as e:
        print(f"Error loading violation details: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading violation details: {str(e)}</div>', 
            status=500
        )

@login_required
def violations_list(request):
    # Get all violations EXCLUDING NCAP violations (those with any type of image)
    all_violations = Violation.objects.filter(
        Q(image='') | Q(image__isnull=True),
        Q(driver_photo='') | Q(driver_photo__isnull=True),
        Q(vehicle_photo='') | Q(vehicle_photo__isnull=True),
        Q(secondary_photo='') | Q(secondary_photo__isnull=True)
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    # Pagination - 15 items per page
    paginator = Paginator(unique_violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}
    
    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/violations_list.html', context)

@login_required
def update_violation_status(request, violation_id):
    if request.method == 'POST':
        try:
            violation = get_object_or_404(Violation, id=violation_id)
            new_status = request.POST.get('status')
            
            if new_status in dict(Violation.STATUS_CHOICES):
                old_status = violation.status
                violation.status = new_status
                violation.save()
                
                # Log the activity
                log_activity(
                    user=request.user,
                    action='Updated Violation Status',
                    details=f'Changed status of Violation #{violation.id} from {old_status} to {new_status}',
                    category='violation'
                )
                
                messages.success(request, f'Violation status updated successfully to {new_status}')
                return JsonResponse({
                    'status': 'success',
                    'message': f'Violation status updated to {new_status}',
                    'new_status': new_status
                })
            else:
                messages.error(request, 'Invalid status provided')
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid status provided'
                }, status=400)
                
        except Exception as e:
            messages.error(request, f'Error updating violation status: {str(e)}')
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
            
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)

@login_required
def user_detail_modal(request, user_id):
    try:
        # Get the User object first, then its profile
        user = get_object_or_404(User, id=user_id)
        profile = user.userprofile
        
        # Log the user ID for debugging
        logger.debug(f"Fetching user modal for User ID: {user_id}, Username: {user.username}")
        
        context = {
            'profile': profile,
            'user': user,
            'is_modal': True,
            'debug_user_id': user_id  # Add this for debugging
        }
        return render(request, 'users/user_detail.html', context)
    except Exception as e:
        logger.error(f"Error in user_detail_modal for user_id {user_id}: {str(e)}")
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading user details (ID: {user_id}). Error: {str(e)}</div>', 
            status=500
        )

@login_required
def ncap_violations_list(request):
    # Get violations with any images (NCAP violations)
    all_violations = Violation.objects.filter(
        Q(image__isnull=False) | 
        Q(driver_photo__isnull=False) | 
        Q(vehicle_photo__isnull=False) | 
        Q(secondary_photo__isnull=False)
    ).exclude(
        Q(image='') & 
        Q(driver_photo='') & 
        Q(vehicle_photo='') & 
        Q(secondary_photo='')
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)

    # Pagination - 12 items per page
    paginator = Paginator(unique_violations, 12)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/ncap_violations_list.html', context)

@login_required
def adjudication_list(request):
    # Get all violations without PAID status
    violations = Violation.objects.select_related('violator', 'enforcer').exclude(status='PAID').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        violations = violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Pagination - 15 items per page
    paginator = Paginator(violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    return render(request, 'violations/adjudication_list.html', context)

@login_required
def adjudication_form(request, violation_id):
    from django.db.models import Q
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    # Get all other violations from the same violator (excluding current one)
    # Now including all statuses, not just PENDING
    previous_violations = Violation.objects.filter(
        violator=violation.violator
    ).exclude(id=violation_id).order_by('-violation_date')
    
    # Also find violations with same license number but different violator records
    if violation.violator and violation.violator.license_number:
        license_violations = Violation.objects.filter(
            violator__license_number=violation.violator.license_number
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).order_by('-violation_date')
        
        # Combine both querysets
        previous_violations = list(previous_violations) + list(license_violations)
    
    # Also find violations with the same name, even if they're different violator records
    if violation.violator:
        name_violations = Violation.objects.filter(
            Q(violator__first_name__iexact=violation.violator.first_name) &
            Q(violator__last_name__iexact=violation.violator.last_name)
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).exclude(
            id__in=[v.id for v in previous_violations]
        ).order_by('-violation_date')
        
        # Add name-matched violations to the list
        previous_violations = list(previous_violations) + list(name_violations)
        
    # Sort by violation date (most recent first)
    if previous_violations:
        previous_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    context = {
        'violation': violation,
        'previous_violations': previous_violations
    }
    
    return render(request, 'violations/adjudication_form.html', context)

@login_required
def submit_adjudication(request, violation_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    try:
        # Process the main violation
        process_adjudication(request, violation)
        
        # Check if there are batch violations to process
        batch_violation_ids_json = request.POST.get('batch_violation_ids', '[]')
        is_batch = False
        if batch_violation_ids_json and batch_violation_ids_json != '[]':
            batch_violation_ids = json.loads(batch_violation_ids_json)
            is_batch = len(batch_violation_ids) > 0
            
            # Log batch adjudication in the activity log
            if batch_violation_ids:
                ActivityLog.objects.create(
                    user=request.user,
                    action="BATCH_ADJUDICATION",
                    details=f"Batch adjudicated {len(batch_violation_ids)} violations along with violation #{violation_id}",
                    category="violation"
                )
            
            # Process each batch violation
            for batch_id in batch_violation_ids:
                try:
                    batch_violation = Violation.objects.get(id=batch_id)
                    
                    # Get removed violation types for this batch item
                    removed_types_key = f'removed_violation_types_{batch_id}'
                    removed_types_json = request.POST.get(removed_types_key, '[]')
                    
                    # Process the batch violation with its specific removed violations
                    process_adjudication(request, batch_violation, removed_types_json)
                    
                except Violation.DoesNotExist:
                    logger.error(f"Batch violation ID {batch_id} not found during adjudication")
                except Exception as e:
                    logger.error(f"Error processing batch violation {batch_id}: {str(e)}")
        
        # Get settled status for success message
        is_settled = request.POST.get('settled_status', 'false') == 'true'
        
        # Redirect with success parameters
        redirect_url = f"{reverse('adjudication_list')}?success=true"
        if is_settled:
            redirect_url += "&settled=true"
        if is_batch:
            redirect_url += "&batch=true"
            
        return redirect(redirect_url)
    
    except Exception as e:
        logger.error(f"Error in submit_adjudication: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)})

def process_adjudication(request, violation, batch_removed_types_json=None):
    """Helper function to process a single violation adjudication"""
    adjudication_status = request.POST.get('adjudication_status', 'adjudicated')
    requires_approval = request.POST.get('requires_approval', 'true') == 'true'
    settled_status = request.POST.get('settled_status', 'false') == 'true'
    
    # Get form data
    total_penalty = request.POST.get('total_penalty', '0')
    remarks = request.POST.get('remarks', '')
    settlement_reason = request.POST.get('settlement_reason', '')
    
    # Parse removed violations - use batch specific ones if provided
    if batch_removed_types_json:
        removed_violations_json = batch_removed_types_json
    else:
        removed_violations_json = request.POST.get('removed_violations', '[]')
    
    removed_violations = []
    try:
        removed_violations = json.loads(removed_violations_json)
    except json.JSONDecodeError:
        logger.error(f"Error parsing removed violations JSON: {removed_violations_json}")
    
    # Handle payment due date
    payment_due_str = request.POST.get('payment_due', '')
    payment_due_date = None
    if payment_due_str:
        try:
            payment_due_date = datetime.strptime(payment_due_str, '%Y-%m-%d').date()
        except ValueError:
            logger.error(f"Invalid payment due date format: {payment_due_str}")
    
    # Set adjudication status based on form data
    if settled_status:
        violation.status = 'SETTLED'
        violation.fine_amount = 0
        remarks = f"SETTLED: {settlement_reason or remarks}"


# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                        
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path



@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now
from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now

# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                        
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path


@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


@login_required
def edit_profile(request):
    # Get or create UserProfile
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'enforcer_id': generate_enforcer_id(),
            'phone_number': '',
            'address': ''
        }
    )

    if request.method == 'POST':
        try:
            user = request.user

            # Update user information
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            # Update profile information
            profile.phone_number = request.POST.get('phone_number')
            profile.address = request.POST.get('address')

            if 'avatar' in request.FILES:
                # Add debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Uploading avatar for user {user.username}")
                logger.info(f"Avatar file name: {request.FILES['avatar'].name}")
                logger.info(f"Using MEDIA_ROOT: {settings.MEDIA_ROOT}")
                
                profile.avatar = request.FILES['avatar']
                logger.info(f"Avatar saved to: {profile.avatar.path if profile.avatar else 'None'}")

            profile.save()

            # Store current page in session
            request.session['current_page'] = 'profile'
            
            # Add success message
            messages.success(request, 'Profile updated successfully!')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Updated Profile',
                details='Profile information was updated',
                category='user'
            )

            return redirect('profile')
        except Exception as e:
            # Log detailed exception info
            import logging, traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Profile update error: {str(e)}")
            logger.error(traceback.format_exc())
            
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    return render(request, 'edit_profile.html', {'user': request.user})


@login_required
def users_list(request):
    user_profiles = UserProfile.objects.all().select_related('user').order_by('user__date_joined')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if search_query:
        user_profiles = user_profiles.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(enforcer_id__icontains=search_query)
        )

    if role_filter:
        user_profiles = user_profiles.filter(role=role_filter)

    if status_filter:
        is_active = status_filter == 'active'
        user_profiles = user_profiles.filter(user__is_active=is_active)

    # Pagination - 10 users per page
    paginator = Paginator(user_profiles, 10)
    page = request.GET.get('page', 1)
    try:
        user_profiles = paginator.page(page)
    except PageNotAnInteger:
        user_profiles = paginator.page(1)
    except EmptyPage:
        user_profiles = paginator.page(paginator.num_pages)
    
    context = {
        'user_profiles': user_profiles,
        'page_obj': user_profiles,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': paginator.count
    }
    return render(request, 'users/users_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        try:
            # Create user
            user = User.objects.create_user(
                username=request.POST['username'],
                password=request.POST['password'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Instead of creating a new profile, update the one created by the signal
            profile = user.userprofile  # This exists because of the signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Created User',
                details=f'Created user account for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User created successfully!')
            # Correctly add query parameter to redirect
            from django.urls import reverse
            return redirect(reverse('users_list') + '?success=created')
        except Exception as e:
            # If there's an error, delete the user to maintain consistency
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error creating user: {str(e)}')
    
    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/create_user.html', {'roles': roles})


@login_required
def user_detail(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    return render(request, 'users/user_detail.html', {'profile': profile})


@login_required
def edit_user(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    
    if request.method == 'POST':
        try:
            user = profile.user
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.save()

            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Updated User',
                details=f'Updated user profile for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User profile updated successfully!')
            return redirect('users_list')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')

    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/edit_user.html', {
        'profile': profile,
        'roles': roles
    })


@login_required
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            profile = get_object_or_404(UserProfile, id=user_id)
            user = profile.user
            
            # Toggle the user's active status
            user.is_active = not user.is_active
            user.save()
            
            # Add success message
            messages.success(request, f'User {user.username} has been {"activated" if user.is_active else "deactivated"} successfully.')
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'message': f'User {user.username} has been {"activated" if user.is_active else "deactivated"}'
            })
        except Exception as e:
            messages.error(request, f'Error changing user status: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def activity_history(request):
    activities = ActivityLog.objects.select_related('user', 'user__userprofile').all()

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        activities = activities.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(user__userprofile__enforcer_id__icontains=search_query)
        ).distinct()

    # Get filter parameters
    category = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_id = request.GET.get('user_id', '')

    # Apply filters
    if category:
        activities = activities.filter(category=category)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            activities = activities.filter(timestamp__gte=date_from)
        except ValueError:
            pass

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to = date_to + timedelta(days=1)
            activities = activities.filter(timestamp__lt=date_to)
        except ValueError:
            pass

    if user_id:
        activities = activities.filter(user_id=user_id)

    # Order by timestamp descending (most recent first)
    activities = activities.order_by('-timestamp')

    # Pagination - 25 activities per page
    paginator = Paginator(activities, 25)
    page = request.GET.get('page', 1)
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)

    # Get all categories and users for filters
    categories = ActivityLog.CATEGORY_CHOICES
    users = User.objects.filter(
        id__in=activities.paginator.object_list.values_list('user_id', flat=True).distinct()
    ).select_related('userprofile')

    context = {
        'activities': activities,
        'categories': categories,
        'users': users,
        'current_category': category,
        'current_user': user_id,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': activities,
        'search_query': search_query,
        'total_count': paginator.count
    }

    return render(request, 'activity_history.html', context)


@login_required
def get_statistics(request, time_range):
    """API endpoint for chart data"""
    today = timezone.now()
    
    if time_range == 'week':
        start_date = today - timedelta(days=7)
        date_format = '%Y-%m-%d'
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
        date_format = '%Y-%m-%d'
    else:  # year
        start_date = today - timedelta(days=365)
        date_format = '%Y-%m'

    # Get violations data
    violations = Violation.objects.filter(
        violation_date__gte=start_date
    ).annotate(
        date=TruncDate('violation_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')

    # Format data for charts
    labels = []
    data = []
    
    # Create a date range for all dates in the period
    date_range = []
    delta = today - start_date
    for i in range(delta.days + 1):
        date_range.append(start_date + timedelta(days=i))
    
    # Create a lookup of existing data
    data_lookup = {item['date']: item['count'] for item in violations}
    
    # Fill in data for all dates
    for date in date_range:
        labels.append(date.strftime(date_format))
        data.append(data_lookup.get(date.date(), 0))

    violations_chart = {
        'labels': labels,
        'datasets': [{
            'label': 'Total Violations',
            'data': data,
            'borderColor': '#2563eb',
            'backgroundColor': 'rgba(37, 99, 235, 0.1)',
            'tension': 0.4,
            'fill': True
        }]
    }

    return JsonResponse({
        'violations': violations_chart
    })


from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now
from django.contrib.auth import logout, authenticate, login
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Value, BooleanField, Case, When, IntegerField, Max, Avg, Min
from django.db.models.functions import Concat, TruncDate, TruncMonth, TruncWeek
from django.utils import timezone
from .models import Violation, Payment, UserProfile, ActivityLog, Violator, Announcement, AnnouncementAcknowledgment, LocationHistory, Operator, Vehicle, OperatorApplication, Driver, ViolationCertificate, DriverVehicleAssignment, ViolationType
from .user_portal.models import UserNotification, UserReport
from django.conf import settings
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, Http404
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import cv2
import numpy as np
from datetime import datetime, timedelta, date
import base64
from django.core.files.base import ContentFile
from django.utils import timezone
from django.contrib.auth.models import User
import random
import string
from .utils import log_activity
import requests
import json
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from idanalyzer import CoreAPI
import os
from django.db import models, transaction
from django.contrib.sessions.backends.db import SessionStore
from .forms import (
    NCAPViolationForm, OperatorForm, ImportOperatorsForm,
    ViolationForm, PaymentForm, ProfileUpdateForm, UserUpdateForm, 
    ViolatorForm, SignatureForm, OperatorForm,
    OperatorImportForm, OperatorApplicationForm, DriverImportForm, DriverForm, VehicleForm, DriverVehicleAssignmentForm
)
from django.core.exceptions import PermissionDenied
import logging
import time
import uuid
import csv, io
import pandas as pd
import xlwt
from django.template.loader import render_to_string
import pytz
import re
import sys
import openpyxl
from decimal import Decimal, InvalidOperation
from django import forms
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.utils.timezone import now

# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                        
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path


@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


@login_required
def edit_profile(request):
    # Get or create UserProfile
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'enforcer_id': generate_enforcer_id(),
            'phone_number': '',
            'address': ''
        }
    )

    if request.method == 'POST':
        try:
            user = request.user

            # Update user information
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            # Update profile information
            profile.phone_number = request.POST.get('phone_number')
            profile.address = request.POST.get('address')

            if 'avatar' in request.FILES:
                # Add debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Uploading avatar for user {user.username}")
                logger.info(f"Avatar file name: {request.FILES['avatar'].name}")
                logger.info(f"Using MEDIA_ROOT: {settings.MEDIA_ROOT}")
                
                profile.avatar = request.FILES['avatar']
                logger.info(f"Avatar saved to: {profile.avatar.path if profile.avatar else 'None'}")

            profile.save()

            # Store current page in session
            request.session['current_page'] = 'profile'
            
            # Add success message
            messages.success(request, 'Profile updated successfully!')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Updated Profile',
                details='Profile information was updated',
                category='user'
            )

            return redirect('profile')
        except Exception as e:
            # Log detailed exception info
            import logging, traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Profile update error: {str(e)}")
            logger.error(traceback.format_exc())
            
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    return render(request, 'edit_profile.html', {'user': request.user})


@login_required
def users_list(request):
    user_profiles = UserProfile.objects.all().select_related('user').order_by('user__date_joined')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if search_query:
        user_profiles = user_profiles.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(enforcer_id__icontains=search_query)
        )

    if role_filter:
        user_profiles = user_profiles.filter(role=role_filter)

    if status_filter:
        is_active = status_filter == 'active'
        user_profiles = user_profiles.filter(user__is_active=is_active)

    # Pagination - 10 users per page
    paginator = Paginator(user_profiles, 10)
    page = request.GET.get('page', 1)
    try:
        user_profiles = paginator.page(page)
    except PageNotAnInteger:
        user_profiles = paginator.page(1)
    except EmptyPage:
        user_profiles = paginator.page(paginator.num_pages)
    
    context = {
        'user_profiles': user_profiles,
        'page_obj': user_profiles,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': paginator.count
    }
    return render(request, 'users/users_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        try:
            # Create user
            user = User.objects.create_user(
                username=request.POST['username'],
                password=request.POST['password'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Instead of creating a new profile, update the one created by the signal
            profile = user.userprofile  # This exists because of the signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Created User',
                details=f'Created user account for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User created successfully!')
            # Correctly add query parameter to redirect
            from django.urls import reverse
            return redirect(reverse('users_list') + '?success=created')
        except Exception as e:
            # If there's an error, delete the user to maintain consistency
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error creating user: {str(e)}')
    
    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/create_user.html', {'roles': roles})


@login_required
def user_detail(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    return render(request, 'users/user_detail.html', {'profile': profile})


@login_required
def edit_user(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    
    if request.method == 'POST':
        try:
            user = profile.user
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.save()

            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Updated User',
                details=f'Updated user profile for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User profile updated successfully!')
            return redirect('users_list')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')

    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/edit_user.html', {
        'profile': profile,
        'roles': roles
    })


@login_required
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            profile = get_object_or_404(UserProfile, id=user_id)
            user = profile.user
            
            # Toggle the user's active status
            user.is_active = not user.is_active
            user.save()
            
            # Add success message
            messages.success(request, f'User {user.username} has been {"activated" if user.is_active else "deactivated"} successfully.')
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'message': f'User {user.username} has been {"activated" if user.is_active else "deactivated"}'
            })
        except Exception as e:
            messages.error(request, f'Error changing user status: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def activity_history(request):
    activities = ActivityLog.objects.select_related('user', 'user__userprofile').all()

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        activities = activities.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(user__userprofile__enforcer_id__icontains=search_query)
        ).distinct()

    # Get filter parameters
    category = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_id = request.GET.get('user_id', '')

    # Apply filters
    if category:
        activities = activities.filter(category=category)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            activities = activities.filter(timestamp__gte=date_from)
        except ValueError:
            pass

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to = date_to + timedelta(days=1)
            activities = activities.filter(timestamp__lt=date_to)
        except ValueError:
            pass

    if user_id:
        activities = activities.filter(user_id=user_id)

    # Order by timestamp descending (most recent first)
    activities = activities.order_by('-timestamp')

    # Pagination - 25 activities per page
    paginator = Paginator(activities, 25)
    page = request.GET.get('page', 1)
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)

    # Get all categories and users for filters
    categories = ActivityLog.CATEGORY_CHOICES
    users = User.objects.filter(
        id__in=activities.paginator.object_list.values_list('user_id', flat=True).distinct()
    ).select_related('userprofile')

    context = {
        'activities': activities,
        'categories': categories,
        'users': users,
        'current_category': category,
        'current_user': user_id,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': activities,
        'search_query': search_query,
        'total_count': paginator.count
    }

    return render(request, 'activity_history.html', context)


@login_required
def get_statistics(request, time_range):
    """API endpoint for chart data"""
    today = timezone.now()
    
    if time_range == 'week':
        start_date = today - timedelta(days=7)
        date_format = '%Y-%m-%d'
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
        date_format = '%Y-%m-%d'
    else:  # year
        start_date = today - timedelta(days=365)
        date_format = '%Y-%m'

    # Get violations data
    violations = Violation.objects.filter(
        violation_date__gte=start_date
    ).annotate(
        date=TruncDate('violation_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')

    # Format data for charts
    labels = []
    data = []
    
    # Create a date range for all dates in the period
    date_range = []
    delta = today - start_date
    for i in range(delta.days + 1):
        date_range.append(start_date + timedelta(days=i))
    
    # Create a lookup of existing data
    data_lookup = {item['date']: item['count'] for item in violations}
    
    # Fill in data for all dates
    for date in date_range:
        labels.append(date.strftime(date_format))
        data.append(data_lookup.get(date.date(), 0))

    violations_chart = {
        'labels': labels,
        'datasets': [{
            'label': 'Total Violations',
            'data': data,
            'borderColor': '#2563eb',
            'backgroundColor': 'rgba(37, 99, 235, 0.1)',
            'tension': 0.4,
            'fill': True
        }]
    }

    return JsonResponse({
        'violations': violations_chart
    })


@login_required
def violation_detail_modal(request, violation_id):
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Check if it's an NCAP violation by checking for any images
        is_ncap = bool(violation.image) or bool(violation.driver_photo) or bool(violation.vehicle_photo) or bool(violation.secondary_photo)
        
        # Get previous violations for the same violator
        previous_violations = Violation.objects.filter(
            violator=violation.violator
        ).exclude(
            id=violation.id
        ).order_by('-violation_date')
        
        # Find violations from same name violators
        same_name_violations = []
        if violation.violator:
            try:
                # Find violations with the same first and last name
                same_name_violations = Violation.objects.filter(
                    violator__first_name__iexact=violation.violator.first_name,
                    violator__last_name__iexact=violation.violator.last_name
                ).exclude(
                    violator=violation.violator  # Exclude the current violator as they're already in previous_violations
                ).order_by('-violation_date')
            except Exception as e:
                print(f"Error finding same name violations: {e}")
        
        # Find violations from the same submission
        submission_violations = []
        if is_ncap:
            try:
                # First, try to find violations with the same submission_id
                if violation.submission_id:
                    submission_violations = list(Violation.objects.filter(
                        submission_id=violation.submission_id
                    ).order_by('id'))
                    
                    print(f"Found {len(submission_violations)} violations with submission_id {violation.submission_id}")
                
                # If no submission_id or no other violations found with it, fall back to time window approach
                if not submission_violations:
                    # Find violations submitted within a 5-minute window of this violation
                    if violation.violation_date:
                        time_window_start = violation.violation_date - timezone.timedelta(minutes=5)
                        time_window_end = violation.violation_date + timezone.timedelta(minutes=5)
                        
                        # Find all violations from the same time window by the same violator
                        submission_violations = list(Violation.objects.filter(
                            violator=violation.violator,
                            violation_date__gte=time_window_start,
                            violation_date__lte=time_window_end
                        ).order_by('id'))
                        
                        print(f"Fallback: Found {len(submission_violations)} violations within 5 min window for violator {violation.violator.id}")
                
                # Debug each violation found
                for idx, v in enumerate(submission_violations):
                    print(f"  Violation #{idx+1}: ID={v.id}, Type={v.violation_type}, Same as current={v.id == int(violation_id)}")
            except Exception as e:
                print(f"Error finding related violations: {e}")
                # In case of error, just use the current violation
                submission_violations = [violation]
        
        # Use the appropriate template
        template_name = 'violations/ncap_violation_detail_modal.html' if is_ncap else 'violations/violation_detail_modal.html'
        
        context = {
            'violation': violation,
            'is_modal': True,
            'previous_violations': previous_violations,
            'same_name_violations': same_name_violations,
            'submission_violations': submission_violations
        }
        
        return render(request, template_name, context)
    except Exception as e:
        print(f"Error loading violation details: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading violation details: {str(e)}</div>', 
            status=500
        )

@login_required
def violations_list(request):
    # Get all violations EXCLUDING NCAP violations (those with any type of image)
    all_violations = Violation.objects.filter(
        Q(image='') | Q(image__isnull=True),
        Q(driver_photo='') | Q(driver_photo__isnull=True),
        Q(vehicle_photo='') | Q(vehicle_photo__isnull=True),
        Q(secondary_photo='') | Q(secondary_photo__isnull=True)
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    # Pagination - 15 items per page
    paginator = Paginator(unique_violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}
    
    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/violations_list.html', context)

@login_required
def update_violation_status(request, violation_id):
    if request.method == 'POST':
        try:
            violation = get_object_or_404(Violation, id=violation_id)
            new_status = request.POST.get('status')
            
            if new_status in dict(Violation.STATUS_CHOICES):
                old_status = violation.status
                violation.status = new_status
                violation.save()
                
                # Log the activity
                log_activity(
                    user=request.user,
                    action='Updated Violation Status',
                    details=f'Changed status of Violation #{violation.id} from {old_status} to {new_status}',
                    category='violation'
                )
                
                messages.success(request, f'Violation status updated successfully to {new_status}')
                return JsonResponse({
                    'status': 'success',
                    'message': f'Violation status updated to {new_status}',
                    'new_status': new_status
                })
            else:
                messages.error(request, 'Invalid status provided')
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid status provided'
                }, status=400)
                
        except Exception as e:
            messages.error(request, f'Error updating violation status: {str(e)}')
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
            
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)

@login_required
def user_detail_modal(request, user_id):
    try:
        # Get the User object first, then its profile
        user = get_object_or_404(User, id=user_id)
        profile = user.userprofile
        
        # Log the user ID for debugging
        logger.debug(f"Fetching user modal for User ID: {user_id}, Username: {user.username}")
        
        context = {
            'profile': profile,
            'user': user,
            'is_modal': True,
            'debug_user_id': user_id  # Add this for debugging
        }
        return render(request, 'users/user_detail.html', context)
    except Exception as e:
        logger.error(f"Error in user_detail_modal for user_id {user_id}: {str(e)}")
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading user details (ID: {user_id}). Error: {str(e)}</div>', 
            status=500
        )

@login_required
def ncap_violations_list(request):
    # Get violations with any images (NCAP violations)
    all_violations = Violation.objects.filter(
        Q(image__isnull=False) | 
        Q(driver_photo__isnull=False) | 
        Q(vehicle_photo__isnull=False) | 
        Q(secondary_photo__isnull=False)
    ).exclude(
        Q(image='') & 
        Q(driver_photo='') & 
        Q(vehicle_photo='') & 
        Q(secondary_photo='')
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)

    # Pagination - 12 items per page
    paginator = Paginator(unique_violations, 12)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/ncap_violations_list.html', context)

@login_required
def adjudication_list(request):
    # Get all violations without PAID status
    violations = Violation.objects.select_related('violator', 'enforcer').exclude(status='PAID').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        violations = violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Pagination - 15 items per page
    paginator = Paginator(violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    return render(request, 'violations/adjudication_list.html', context)

@login_required
def adjudication_form(request, violation_id):
    from django.db.models import Q
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    # Get all other violations from the same violator (excluding current one)
    # Now including all statuses, not just PENDING
    previous_violations = Violation.objects.filter(
        violator=violation.violator
    ).exclude(id=violation_id).order_by('-violation_date')
    
    # Also find violations with same license number but different violator records
    if violation.violator and violation.violator.license_number:
        license_violations = Violation.objects.filter(
            violator__license_number=violation.violator.license_number
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).order_by('-violation_date')
        
        # Combine both querysets
        previous_violations = list(previous_violations) + list(license_violations)
    
    # Also find violations with the same name, even if they're different violator records
    if violation.violator:
        name_violations = Violation.objects.filter(
            Q(violator__first_name__iexact=violation.violator.first_name) &
            Q(violator__last_name__iexact=violation.violator.last_name)
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).exclude(
            id__in=[v.id for v in previous_violations]
        ).order_by('-violation_date')
        
        # Add name-matched violations to the list
        previous_violations = list(previous_violations) + list(name_violations)
        
    # Sort by violation date (most recent first)
    if previous_violations:
        previous_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    context = {
        'violation': violation,
        'previous_violations': previous_violations
    }
    
    return render(request, 'violations/adjudication_form.html', context)

@login_required
def submit_adjudication(request, violation_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    try:
        # Process the main violation
        process_adjudication(request, violation)
        
        # Check if there are batch violations to process
        batch_violation_ids_json = request.POST.get('batch_violation_ids', '[]')
        is_batch = False
        if batch_violation_ids_json and batch_violation_ids_json != '[]':
            batch_violation_ids = json.loads(batch_violation_ids_json)
            is_batch = len(batch_violation_ids) > 0
            
            # Log batch adjudication in the activity log
            if batch_violation_ids:
                ActivityLog.objects.create(
                    user=request.user,
                    action="BATCH_ADJUDICATION",
                    details=f"Batch adjudicated {len(batch_violation_ids)} violations along with violation #{violation_id}",
                    category="violation"
                )
            
            # Process each batch violation
            for batch_id in batch_violation_ids:
                try:
                    batch_violation = Violation.objects.get(id=batch_id)
                    
                    # Get removed violation types for this batch item
                    removed_types_key = f'removed_violation_types_{batch_id}'
                    removed_types_json = request.POST.get(removed_types_key, '[]')
                    
                    # Process the batch violation with its specific removed violations
                    process_adjudication(request, batch_violation, removed_types_json)
                    
                except Violation.DoesNotExist:
                    logger.error(f"Batch violation ID {batch_id} not found during adjudication")
                except Exception as e:
                    logger.error(f"Error processing batch violation {batch_id}: {str(e)}")
        
        # Get settled status for success message
        is_settled = request.POST.get('settled_status', 'false') == 'true'
        
        # Redirect with success parameters
        redirect_url = f"{reverse('adjudication_list')}?success=true"
        if is_settled:
            redirect_url += "&settled=true"
        if is_batch:
            redirect_url += "&batch=true"
            
        return redirect(redirect_url)
    
    except Exception as e:
        logger.error(f"Error in submit_adjudication: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)})

def process_adjudication(request, violation, batch_removed_types_json=None):
    """Helper function to process a single violation adjudication"""
    adjudication_status = request.POST.get('adjudication_status', 'adjudicated')
    requires_approval = request.POST.get('requires_approval', 'true') == 'true'
    settled_status = request.POST.get('settled_status', 'false') == 'true'
    
    # Get form data
    total_penalty = request.POST.get('total_penalty', '0')
    remarks = request.POST.get('remarks', '')
    settlement_reason = request.POST.get('settlement_reason', '')
    
    # Parse removed violations - use batch specific ones if provided
    if batch_removed_types_json:
        removed_violations_json = batch_removed_types_json
    else:
        removed_violations_json = request.POST.get('removed_violations', '[]')
    
    removed_violations = []
    try:
        removed_violations = json.loads(removed_violations_json)
    except json.JSONDecodeError:
        logger.error(f"Error parsing removed violations JSON: {removed_violations_json}")
    
    # Handle payment due date
    payment_due_str = request.POST.get('payment_due', '')
    payment_due_date = None
    if payment_due_str:
        try:
            payment_due_date = datetime.strptime(payment_due_str, '%Y-%m-%d').date()
        except ValueError:
            logger.error(f"Invalid payment due date format: {payment_due_str}")
    
    # Set adjudication status based on form data
    if settled_status:
        violation.status = 'SETTLED'
        violation.fine_amount = 0
        remarks = f"SETTLED: {settlement_reason or remarks}"


# Initialize logger
logger = logging.getLogger(__name__)

def get_client_ip(request):
    """Get the client's IP address from the request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

# Define choices after importing models but before any views
VIOLATION_CHOICES = [
    ('Illegal Parking', 'Illegal Parking (DTS)'),
    ('Entering Prohibited Zones', 'Entering Prohibited Zones'),
    ('Obstruction', 'Obstruction'),
    ('Unlicensed Driver', 'Unlicensed Driver'),
    ('Permitting Hitching', 'Permitting Hitching/Over Loading Passenger(s)'),
    ('Unregistered MV', 'Unregistered MV'),
    ('Refusal to Convey', 'Refusal to convey to proper destination'),
    ('Discourteous Driver', 'Discourteous driver/Conduct'),
    ('Defective Lights', 'No/Defective Lights'),
    ('Expired OR/CR', 'Expired OR/CR'),
    ('No License', 'Failure to Carry DL'),
    ('No Permit', "No MAYOR'S PERMIT/MTOP/POP/PDP"),
    ('Overcharging', 'Overcharging'),
    ('DUI', 'Driving under the influence of liquor/drugs'),
    ('Defective Muffler', 'Operating a vehicle with Defective Muffler'),
    ('Dilapidated', 'Operating a Dilapidated Motorcab'),
    ('Reckless Driving', 'Reckless Driving'),
    ('Others', 'Others')
]

VEHICLE_CLASSIFICATIONS = [
    ('Private', 'Private'),
    ('Public', 'Public'),
    ('Government', 'Government'),
    ('Commercial', 'Commercial')
]

@login_required
def admin_dashboard(request):
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Get violations statistics
    violations = Violation.objects.all()
    total_violations = violations.count()
    weekly_violations = violations.filter(violation_date__gte=week_ago).count()
    monthly_violations = violations.filter(violation_date__gte=month_ago).count()

    # Get status-based counts
    pending_violations = violations.filter(status='PENDING').count()
    paid_violations_count = violations.filter(status='PAID').count()
    settled_violations = violations.filter(status='SETTLED').count()
    overdue_violations = violations.filter(status='OVERDUE').count()

    # Calculate revenue statistics
    paid_violations = violations.filter(status='PAID')
    total_revenue = paid_violations.aggregate(total=Sum('fine_amount'))['total'] or 0
    monthly_revenue = paid_violations.filter(
        receipt_date__gte=month_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    weekly_revenue = paid_violations.filter(
        receipt_date__gte=week_ago
    ).aggregate(total=Sum('fine_amount'))['total'] or 0

    # Get violations trend data (last 30 days)
    dates = []
    violations_data = []
    
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        # Format date as ISO string for proper JavaScript parsing
        dates.append(date.strftime('%Y-%m-%d'))
        count = violations.filter(
            violation_date__date=date.date()
        ).count()
        violations_data.append(count)

    # Convert dates to JSON-safe format
    from django.core.serializers.json import DjangoJSONEncoder
    import json
    
    context = {
        'total_violations': total_violations,
        'total_revenue': total_revenue,
        'pending_violations': pending_violations,
        'overdue_violations': overdue_violations,
        'weekly_violations': weekly_violations,
        'monthly_violations': monthly_violations,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'paid_violations': paid_violations_count,
        'settled_violations': settled_violations,
        'dates': json.dumps(dates, cls=DjangoJSONEncoder),
        'violations_data': json.dumps(violations_data),
        'recent_violations': violations.order_by('-violation_date')[:10]
    }

    return render(request, 'violations/admin_dashboard.html', context)


@login_required
def violation_detail(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    payment = Payment.objects.filter(violation=violation).first()
    
    context = {
        'violation': violation,
        'payment': payment,
        'paymongo_public_key': settings.PAYMONGO_PUBLIC_KEY,
    }
    
    # If it's an HTMX request, return just the content without the base template
    if request.headers.get('HX-Request'):
        return render(request, 'violations/violation_detail.html', context)
    
    return render(request, 'violations/violation_detail.html', context)


@login_required
def process_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)

    try:
        # Create Paymongo checkout session
        url = "https://api.paymongo.com/v1/checkout_sessions"
        
        # Convert amount to centavos (cents)
        amount = int(float(violation.fine_amount) * 100)
        
        payload = {
            "data": {
                "attributes": {
                    "send_email_receipt": True,
                    "show_description": True,
                    "show_line_items": True,
                    "line_items": [
                        {
                            "currency": "PHP",
                            "amount": amount,
                            "name": f"Violation #{violation.id}",
                            "quantity": 1,
                            "description": f"Payment for {violation.violation_type}"
                        }
                    ],
                    "payment_method_types": ["gcash", "card"],
                    "description": f"Payment for Violation #{violation.id}",
                    "reference_number": f"VIO-{violation.id}",
                    "success_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=success",
                    "cancel_url": request.build_absolute_uri(reverse('violation_detail', args=[violation.id])) + "?status=cancelled"
                }
            }
        }

        # Create Basic Auth header
        auth_string = base64.b64encode(f"{settings.PAYMONGO_SECRET_KEY}:".encode()).decode()
        
        headers = {
            "accept": "application/json",
            "Content-Type": "application/json",
            "authorization": f"Basic {auth_string}"
        }

        # Print request details for debugging
        print("Request Headers:", headers)
        print("Request Payload:", json.dumps(payload, indent=2))

        # Create checkout session
        response = requests.post(url, json=payload, headers=headers)
        
        # Print response for debugging
        print("Response Status:", response.status_code)
        print("Response Body:", response.text)
        
        if response.status_code == 200:
            checkout_data = response.json()
            checkout_url = checkout_data["data"]["attributes"]["checkout_url"]
            
            # Update violation status to DISPUTED immediately
            violation.status = 'DISPUTED'
            violation.save()
            
            # Create payment record
            Payment.objects.create(
                violation=violation,
                amount_paid=violation.fine_amount,
                payment_method='online',
                transaction_id=checkout_data["data"]["id"]
            )
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Payment Initiated - Status Updated to Disputed',
                details=f'Payment initiated for Violation #{violation.id}. Status updated to DISPUTED.',
                category='payment'
            )
            
            # Return the checkout URL
            return JsonResponse({
                'checkout_url': checkout_url
            })
        else:
            error_message = response.json() if response.text else 'Unknown error'
            print("Paymongo Error:", error_message)
            return JsonResponse({
                'error': f'Failed to create checkout session: {error_message}'
            }, status=400)

    except Exception as e:
        import traceback
        print("Exception:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def payment_webhook(request):
    if request.method == 'POST':
        try:
            # Parse the webhook data
            payload = json.loads(request.body)
            print("Webhook payload:", json.dumps(payload, indent=2))  # Debug log

            # Check if this is a successful checkout session
            event_type = payload.get('data', {}).get('attributes', {}).get('type')
            
            if event_type == 'checkout_session.completed':
                # Get the checkout session data
                checkout_data = payload['data']['attributes']
                
                # Get the reference number which contains our violation ID
                reference_number = checkout_data.get('reference_number')
                if not reference_number:
                    print("No reference number found in webhook data")
                    return HttpResponse(status=400)
                
                violation_id = reference_number.replace('VIO-', '')
                
                try:
                    # Get the violation
                    violation = Violation.objects.get(id=violation_id)
                    
                    # Create payment record
                    Payment.objects.create(
                        violation=violation,
                        amount_paid=float(checkout_data['line_items'][0]['amount']) / 100,  # Convert from cents
                        payment_method=checkout_data.get('payment_method_used', 'online'),
                        transaction_id=checkout_data['id']
                    )
                    
                    # Update violation status to PAID
                    violation.status = 'PAID'  # Changed from 'DISPUTED' to 'PAID'
                    violation.save()
                    
                    # Log the activity
                    try:
                        user = User.objects.get(username='system')
                    except User.DoesNotExist:
                        user = User.objects.create_user(username='system', password='systemuser123')
                    
                    log_activity(
                        user=user,
                        action='Payment Received - Status Updated to Paid',
                        details=f'Payment received for Violation #{violation.id} via {checkout_data.get("payment_method_used", "online")}. Status updated to PAID.',
                        category='payment'
                    )
                    
                    print(f"Successfully processed payment for violation {violation_id}")
                    return HttpResponse(status=200)
                    
                except Violation.DoesNotExist:
                    print(f"Error: Violation {violation_id} not found")
                    return HttpResponse(status=404)
                except Exception as e:
                    print(f"Error processing payment: {str(e)}")
                    return HttpResponse(status=400)
            
            return HttpResponse(status=200)

        except json.JSONDecodeError as e:
            print(f"Error decoding webhook payload: {str(e)}")
            return HttpResponse(status=400)
        except Exception as e:
            print(f"Webhook processing error: {str(e)}")
            return HttpResponse(status=400)

    return HttpResponse(status=405)  # Method not allowed for non-POST requests


@login_required
def upload_violation(request):
    """View for enforcer to upload a violation (NCAP)"""
    if request.method == 'POST':
        # Check if this request has already been processed
        request_id = request.POST.get('request_id', '')
        if not request_id:
            request_id = str(uuid.uuid4())
        
        session_key = f'processed_violation_{request_id}'
        if request.session.get(session_key):
            # This request has already been processed
            logger.warning(f"Duplicate form submission detected with request_id: {request_id}")
            
            # For AJAX requests, return appropriate response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                response = JsonResponse({
                    'success': False,
                    'message': "This form has already been submitted. Please refresh the page to create a new violation.",
                    'details': "Duplicate submission detected"
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                messages.warning(request, "This form has already been submitted. Please refresh the page to create a new violation.")
                return redirect('violations_list')
        
        # Mark this request as processed
        request.session[session_key] = True
        request.session.save()
        
        try:
            # Extract form data
            data = request.POST
            
            # Validate required fields
            required_fields = ['violation_type', 'location', 'violation_date', 'fine_amount']
            missing_fields = [field for field in required_fields if not data.get(field)]
            
            if missing_fields:
                error_message = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please fill in all required fields.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Validate fine amount is a valid number
            try:
                fine_amount = Decimal(data.get('fine_amount', 0))
                if fine_amount < 0:
                    raise ValueError("Fine amount cannot be negative")
            except (ValueError, InvalidOperation) as e:
                error_message = f"Invalid fine amount: {str(e)}"
                logger.error(f"Validation error: {error_message}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid fine amount.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Get selected IDs if present
            selected_operator_id = data.get('selected_operator_id', '')
            selected_driver_id = data.get('selected_driver_id', '')
            
            # Create a new violator or get an existing one
            try:
                license_number = data.get('license_number', '').strip()
                
                # No longer generate placeholder for empty license numbers
                # If license number is empty, keep it empty
                
                # Extract name components
                full_name = data.get('driver_name', '').strip()
                first_name = ''
                last_name = ''
                
                if full_name:
                    name_parts = full_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else ''
                    last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Try to find existing violator by license number, but only if license number is provided
                if license_number:
                    try:
                        violator = Violator.objects.get(license_number=license_number)
                        
                        # Only update if values are provided
                        if first_name:
                            violator.first_name = first_name
                        if last_name:
                            violator.last_name = last_name
                        if data.get('driver_address'):
                            violator.address = data.get('driver_address')
                        
                        violator.save()
                        logger.info(f"Updated existing violator: {violator.id}")
                        
                    except Violator.DoesNotExist:
                        # Create new violator with the provided license number
                        violator = Violator.objects.create(
                            license_number=license_number,
                            first_name=first_name if first_name else 'Unknown',
                            last_name=last_name if last_name else 'Driver',
                            address=data.get('driver_address', '')
                        )
                        logger.info(f"Created new violator with license: {violator.id}")
                else:
                    # Always create a new violator when no license number is provided
                    violator = Violator.objects.create(
                        license_number='',  # Empty license number
                        first_name=first_name if first_name else 'Unknown',
                        last_name=last_name if last_name else 'Driver',
                        address=data.get('driver_address', '')
                    )
                    logger.info(f"Created new violator without license: {violator.id}")
            except Exception as ve:
                error_message = f"Error processing violator information: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error processing the violator information.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Process date and time
            try:
                violation_date = process_date_time(data.get('violation_date', ''), data.get('violation_time', ''))
            except Exception as date_error:
                error_message = f"Invalid date/time format: {str(date_error)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "Please enter a valid date and time.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Create a new violation
            try:
                violation = Violation.objects.create(
                    violation_type=data.get('violation_type', ''),
                    location=data.get('location', ''),
                    violation_date=violation_date,
                    fine_amount=fine_amount,
                    plate_number=data.get('plate_number', ''),
                    vehicle_type=data.get('vehicle_type', ''),
                    color=data.get('color', ''),  # Correct field name is 'color' not 'vehicle_color'
                    classification=data.get('classification', ''),
                    violator=violator,
                    enforcer=request.user,
                    potpot_number=data.get('potpot_number', ''),
                    operator_address=data.get('operator_address', ''),
                    driver_name=data.get('driver_name', ''),
                    driver_address=data.get('driver_address', ''),
                    pd_number=data.get('pd_number', ''),
                    status='PENDING'
                )
            except Exception as ve:
                error_message = f"Error creating violation record: {str(ve)}"
                logger.error(error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    response = JsonResponse({
                        'success': False,
                        'message': "There was an error creating the violation record.",
                        'details': error_message
                    }, status=400)
                    response['Content-Type'] = 'application/json'
                    return response
                else:
                    messages.error(request, error_message)
                    return render(request, 'violations/upload_violation.html', {
                        'violation_choices': VIOLATION_CHOICES
                    })
            
            # Link to operator if selected
            if selected_operator_id:
                try:
                    operator = Operator.objects.get(pk=selected_operator_id)
                    violation.operator = operator
                    # Don't overwrite fields if they were manually entered in the form
                    if not data.get('operator_name'):
                        violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                    else:
                        violation.operator_name = data.get('operator_name')
                    
                    if not data.get('operator_pd_number'):
                        violation.operator_pd_number = operator.new_pd_number
                    else:
                        violation.operator_pd_number = data.get('operator_pd_number')
                    
                    if not data.get('operator_address'):
                        violation.operator_address = operator.address
                    # potpot_number is already set in the create call above

                    violation.save()
                except Operator.DoesNotExist:
                    pass
            # Use form data if operator info manually entered
            elif data.get('operator_name') or data.get('operator_pd_number'):
                violation.operator_name = data.get('operator_name', '')
                violation.operator_pd_number = data.get('operator_pd_number', '')
                violation.save()
                
            # Link to driver if selected
            if selected_driver_id:
                try:
                    driver = Driver.objects.get(pk=selected_driver_id)
                    # Set driver information directly
                    violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                    violation.pd_number = driver.new_pd_number
                    violation.driver_address = driver.address
                    
                    # Copy other driver fields if available
                    if hasattr(driver, 'novr_number') and driver.novr_number:
                        violation.novr_number = driver.novr_number
                    if hasattr(driver, 'pin_number') and driver.pin_number:
                        violation.pin_number = driver.pin_number
                    if hasattr(driver, 'license_number') and driver.license_number:
                        violation.license_number = driver.license_number
                    
                    violation.save()
                except Driver.DoesNotExist:
                    pass
            # Use form data if driver not selected but info provided
            elif data.get('driver_name'):
                violation.driver_name = data.get('driver_name')
                if data.get('novr_number'): violation.novr_number = data.get('novr_number')
                if data.get('pin_number'): violation.pin_number = data.get('pin_number')
                if data.get('pd_number'): violation.pd_number = data.get('pd_number')
                violation.save()
            
            # Process images
            try:
                process_violation_images(request, violation)
            except Exception as img_error:
                logger.error(f"Error processing images: {str(img_error)}")
                # Continue even if image processing fails
            
            # Create certificate if certificate text is provided
            if data.get('certificate_text'):
                try:
                    ViolationCertificate.objects.create(
                        violation=violation,
                        operations_officer=data.get('operations_officer', ''),
                        ctm_officer=data.get('cttm_officer', ''),
                        certificate_text=data.get('certificate_text', '')
                    )
                except Exception as cert_error:
                    logger.error(f"Error creating certificate: {str(cert_error)}")
                    # Continue even if certificate creation fails
            
            # Log activity
            try:
                log_activity(
                    user=request.user,
                    action='Created Violation',
                    details=f'Violation created for {violation.violator.license_number}',
                    category='violation'
                )
            except Exception as log_error:
                logger.error(f"Error logging activity: {str(log_error)}")
                # Continue even if logging fails
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response for AJAX requests
                response = JsonResponse({
                    'success': True,
                    'message': "Violation has been recorded successfully.",
                    'redirect_url': reverse('violations_list'),
                    'id': violation.id  # Add the violation ID to the response
                })
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - show message and redirect
                messages.success(request, "Violation has been recorded successfully.")
                return redirect('violations_list')
            
        except Exception as e:
            # Log the error
            logger.error(f"Error in upload_violation: {str(e)}")
            error_message = f"An error occurred: {str(e)}"
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                response = JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
                response['Content-Type'] = 'application/json'
                return response
            else:
                # Regular form submission - add error message and continue to render the form
                messages.error(request, error_message)
    
    # Prepare context for GET request
    violation_choices = [
        ('Driving without license', 'Driving without license'),
        ('Reckless driving', 'Reckless driving'),
        ('Driving under influence', 'Driving under influence'),
        ('Speeding', 'Speeding'),
        ('No helmet', 'No helmet'),
        ('No seatbelt', 'No seatbelt'),
        ('Illegal parking', 'Illegal parking'),
        ('Invalid registration', 'Invalid registration'),
        ('Overloading', 'Overloading'),
        ('Driving against traffic', 'Driving against traffic'),
        ('Using mobile phone while driving', 'Using mobile phone while driving'),
        ('Disobeying traffic signs', 'Disobeying traffic signs'),
        ('No plate number', 'No plate number'),
        ('Improper turn', 'Improper turn'),
        ('Defective parts/accessories', 'Defective parts/accessories'),
    ]
    
    return render(request, 'violations/upload_violation.html', {
        'violation_choices': violation_choices
    })


def process_violation_image(image_path):
    """Process the violation image using OpenCV"""
    try:
        # Read image
        img = cv2.imread(image_path)

        # Basic image processing (example)
        img = cv2.resize(img, (800, 600))
        img = cv2.addWeighted(img, 1.2, img, 0, 0)  # Enhance contrast

        # Save processed image
        processed_path = image_path.replace('temp/', 'processed/')
        cv2.imwrite(processed_path, img)

        return processed_path
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return image_path



@login_required
def issue_ticket(request):
    if request.method == 'POST':
        try:
            print("POST data:", request.POST)
            
            # First check if a violator with this license number exists
            license_number = request.POST.get('license_number')
            if license_number:
                license_number = license_number.strip()
            print("License number:", license_number)
            
            # Check if this violation should be associated with a user account
            user_account_id = request.POST.get('user_account_id')
            violator_source = request.POST.get('violator_source')
            
            user_account = None
            if user_account_id and violator_source == 'user':
                try:
                    user_account = User.objects.get(id=user_account_id)
                    print(f"Found user account: {user_account.username} (ID: {user_account_id})")
                except User.DoesNotExist:
                    print(f"User account with ID {user_account_id} not found")
                    user_account = None
            
            # Use filter().first() instead of get() to handle multiple records
            violator = None
            if license_number:
                violator = Violator.objects.filter(license_number=license_number).first()
            
            if violator:
                # Update violator information
                violator.first_name = request.POST.get('first_name')
                violator.last_name = request.POST.get('last_name')
                violator.phone_number = request.POST.get('phone_number')
                violator.address = request.POST.get('address')
                violator.save()
            else:
                # Create new violator if not found
                violator = Violator.objects.create(
                    first_name=request.POST.get('first_name'),
                    last_name=request.POST.get('last_name'),
                    license_number=license_number or '',  # Keep as empty string if no license provided
                    phone_number=request.POST.get('phone_number'),
                    address=request.POST.get('address')
                )

            # Get list of selected violations
            violation_types = request.POST.getlist('violation_type[]')
            print("Selected violations:", violation_types)
            
            if not violation_types:
                return JsonResponse({
                    'success': False,
                    'message': "Please select at least one violation type"
                })

            # Create the violation ticket
            violation = Violation.objects.create(
                violator=violator,
                location=request.POST.get('location'),
                fine_amount=request.POST.get('fine_amount'),
                is_tdz_violation=request.POST.get('is_tdz_violation') == 'on',
                vehicle_type=request.POST.get('vehicle_type'),
                plate_number=request.POST.get('plate_number'),
                color=request.POST.get('color'),
                classification=request.POST.get('classification'),
                registration_number=request.POST.get('registration_number'),
                registration_date=request.POST.get('registration_date') or None,
                vehicle_owner=request.POST.get('vehicle_owner'),
                vehicle_owner_address=request.POST.get('vehicle_owner_address'),
                status='PENDING',
                enforcer=request.user,
                enforcer_signature_date=timezone.now(),
                user_account=user_account  # Link to user account if provided
            )
            
            # Record the association in the activity log
            if user_account:
                ActivityLog.objects.create(
                    user=request.user,
                    action=f"Linked violation #{violation.id} to user account {user_account.username}",
                    details=f"Violation ID: {violation.id} linked to user: {user_account.username}",
                    category="violation"
                )
                
                # Create a notification for the user
                UserNotification.objects.create(
                    user=user_account,
                    message=f"A new violation has been issued to you. Violation ID: {violation.id}",
                    type="VIOLATION",
                    reference_id=violation.id
                )

            # Initialize signature_data as None
            signature_data = None

            # Handle signature if provided
            signature_input = request.POST.get('signature_data')
            if signature_input:
                try:
                    # Check if signature_data is a filename (new approach) or base64 data (old approach)
                    if ';base64,' in signature_input:
                        # Handle old base64 format
                        format_data = signature_input.split(';base64,', 1)
                        if len(format_data) == 2:
                            format_info, imgstr = format_data
                            ext = format_info.split('/')[-1]
                            signature_file = ContentFile(base64.b64decode(imgstr), name=f'signature_{violation.id}.{ext}')
                            violation.violator_signature = signature_file
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_file.name
                    else:
                        # Handle new filename format
                        signature_path = os.path.join('signatures', signature_input)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                violation.violator_signature.save(
                                    f'signature_{violation.id}.png',
                                    ContentFile(f.read()),
                                    save=True
                                )
                            violation.violator_signature_date = timezone.now()
                            violation.save()
                            signature_data = signature_input
                except Exception as e:
                    print(f"Error processing signature: {str(e)}")

            # Add violation types
            violation_types_list = []
            for violation_type in violation_types:
                violation_types_list.append(violation_type)
            
            # Store as JSON
            violation.violation_types = json.dumps(violation_types_list)
            
            # Also store the first type in the single field for compatibility
            if violation_types_list:
                violation.violation_type = violation_types_list[0]
            
            violation.save()

            # Create ticket data dictionary
            ticket_data = {
                'id': violation.id,
                'first_name': violator.first_name,
                'last_name': violator.last_name,
                'license_number': violator.license_number,
                'phone_number': violator.phone_number,
                'address': violator.address,
                'vehicle_type': violation.vehicle_type,
                'plate_number': violation.plate_number,
                'color': violation.color,
                'classification': violation.classification,
                'registration_number': violation.registration_number,
                'registration_date': violation.registration_date,
                'location': violation.location,
                'violations': violation_types,
                'fine_amount': float(violation.fine_amount),
                'is_tdz_violation': violation.is_tdz_violation,
                'violation_date': violation.violation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'signature_data': signature_data,
                'enforcer_name': request.user.get_full_name(),
                'enforcer_id': request.user.userprofile.enforcer_id
            }

            return JsonResponse({
                'success': True,
                'message': 'Ticket issued successfully',
                'ticket_data': ticket_data
            })

        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({
                'success': False,
                'message': str(e)
            })

    # Get all active violation types for the dropdown
    violation_types = ViolationType.objects.filter(is_active=True).order_by('category', 'name')
    
    # Convert to JSON for JavaScript use
    violation_types_json = json.dumps([{
        'name': vt.name, 
        'amount': str(vt.amount),
        'category': vt.category,
        'is_ncap': vt.is_ncap
    } for vt in violation_types])

    return render(request, 'violations/issue_direct_ticket.html', {
        'violation_choices': VIOLATION_CHOICES,
        'vehicle_classifications': VEHICLE_CLASSIFICATIONS,
        'violation_types': violation_types,
        'violation_types_json': violation_types_json
    })


def custom_logout(request):
    logout(request)
    # First create the redirect response
    response = redirect('login')
    # Add cache control headers to prevent back button from showing authenticated pages
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Add success query parameter to trigger SweetAlert
    response['Location'] = response['Location'] + '?logout=success'
    return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Get the next URL from session or use default
            next_url = request.session.get('next', None)
            if next_url:
                del request.session['next']  # Clear the stored URL
                return redirect(next_url)
                
            # Default redirects based on user role with success parameter
            if user.userprofile.role == 'USER':
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("user_portal:user_dashboard")}?login=success')
            else:
                # Add success parameter to show SweetAlert on redirect
                return redirect(f'{reverse("admin_dashboard")}?login=success')
        else:
            messages.error(request, 'Invalid username or password.')
    
    # Add no-cache headers for the login page
    response = render(request, 'registration/login.html')
    response['Cache-Control'] = 'no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def generate_enforcer_id():
    """Generate a unique enforcer ID"""
    prefix = 'ENF'
    while True:
        # Generate random 4 digits
        numbers = ''.join(random.choices(string.digits, k=4))
        enforcer_id = f"{prefix}{numbers}"
        if not UserProfile.objects.filter(enforcer_id=enforcer_id).exists():
            return enforcer_id


def register(request):
    if request.method == 'POST':
        try:
            # Check for existing username
            username = request.POST['username']
            email = request.POST['email']
            license_number = request.POST['license_number']
            
            # Check for duplicate username
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists. Please choose a different username.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate email
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email address already exists. Please use a different email.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Check for duplicate license
            if UserProfile.objects.filter(license_number=license_number).exists():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Validate password match
            if request.POST['password'] != request.POST.get('confirm_password', ''):
                messages.error(request, 'Passwords do not match.')
                return render(request, 'registration/register.html', {'form_data': request.POST})
            
            # Create user
            user = User.objects.create_user(
                username=username,
                password=request.POST['password'],
                email=email,
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Create user profile
            profile = UserProfile.objects.get(user=user)  # Get profile created by signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = 'USER'  # Set role as USER
            profile.license_number = license_number  # Add license number
            profile.save()

            # Handle avatar upload
            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
                profile.save()

            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except IntegrityError as e:
            # Handle database integrity errors (e.g., unique constraints)
            if 'username' in str(e).lower():
                messages.error(request, 'Username already exists. Please choose a different username.')
            elif 'email' in str(e).lower():
                messages.error(request, 'Email address already exists. Please use a different email.')
            elif 'license_number' in str(e).lower():
                messages.error(request, 'License number already registered. Please contact support if this is an error.')
            else:
                messages.error(request, f'Registration failed due to database error: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'registration/register.html', {'form_data': request.POST})

    return render(request, 'registration/register.html')


@login_required
def profile(request):
    # Get the user's violations
    user_violations = Violation.objects.filter(user_account=request.user)
    
    # Count open (unpaid) violations
    open_violations_count = user_violations.exclude(status__in=['PAID', 'SETTLED']).count()
    
    # Calculate total fines due
    total_fines_due = user_violations.exclude(status__in=['PAID', 'SETTLED']).aggregate(
        total=models.Sum('fine_amount', default=0)
    )['total'] or 0
    
    # Get latest notifications
    notifications = UserNotification.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Add context for showing messages
    context = {
        'user_violations': user_violations,
        'open_violations_count': open_violations_count,
        'total_fines_due': total_fines_due,
        'notifications': notifications,
        'show_messages': request.session.get('current_page') == 'profile'
    }
    
    return render(request, 'profile.html', context)


@login_required
def edit_profile(request):
    # Get or create UserProfile
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={
            'enforcer_id': generate_enforcer_id(),
            'phone_number': '',
            'address': ''
        }
    )

    if request.method == 'POST':
        try:
            user = request.user

            # Update user information
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST['email']
            user.save()

            # Update profile information
            profile.phone_number = request.POST.get('phone_number')
            profile.address = request.POST.get('address')

            if 'avatar' in request.FILES:
                # Add debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Uploading avatar for user {user.username}")
                logger.info(f"Avatar file name: {request.FILES['avatar'].name}")
                logger.info(f"Using MEDIA_ROOT: {settings.MEDIA_ROOT}")
                
                profile.avatar = request.FILES['avatar']
                logger.info(f"Avatar saved to: {profile.avatar.path if profile.avatar else 'None'}")

            profile.save()

            # Store current page in session
            request.session['current_page'] = 'profile'
            
            # Add success message
            messages.success(request, 'Profile updated successfully!')
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Updated Profile',
                details='Profile information was updated',
                category='user'
            )

            return redirect('profile')
        except Exception as e:
            # Log detailed exception info
            import logging, traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Profile update error: {str(e)}")
            logger.error(traceback.format_exc())
            
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    return render(request, 'edit_profile.html', {'user': request.user})


@login_required
def users_list(request):
    user_profiles = UserProfile.objects.all().select_related('user').order_by('user__date_joined')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if search_query:
        user_profiles = user_profiles.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(enforcer_id__icontains=search_query)
        )

    if role_filter:
        user_profiles = user_profiles.filter(role=role_filter)

    if status_filter:
        is_active = status_filter == 'active'
        user_profiles = user_profiles.filter(user__is_active=is_active)

    # Pagination - 10 users per page
    paginator = Paginator(user_profiles, 10)
    page = request.GET.get('page', 1)
    try:
        user_profiles = paginator.page(page)
    except PageNotAnInteger:
        user_profiles = paginator.page(1)
    except EmptyPage:
        user_profiles = paginator.page(paginator.num_pages)
    
    context = {
        'user_profiles': user_profiles,
        'page_obj': user_profiles,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'total_count': paginator.count
    }
    return render(request, 'users/users_list.html', context)


@login_required
def create_user(request):
    if request.method == 'POST':
        try:
            # Create user
            user = User.objects.create_user(
                username=request.POST['username'],
                password=request.POST['password'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )

            # Instead of creating a new profile, update the one created by the signal
            profile = user.userprofile  # This exists because of the signal
            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Created User',
                details=f'Created user account for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User created successfully!')
            # Correctly add query parameter to redirect
            from django.urls import reverse
            return redirect(reverse('users_list') + '?success=created')
        except Exception as e:
            # If there's an error, delete the user to maintain consistency
            if 'user' in locals():
                user.delete()
            messages.error(request, f'Error creating user: {str(e)}')
    
    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/create_user.html', {'roles': roles})


@login_required
def user_detail(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    return render(request, 'users/user_detail.html', {'profile': profile})


@login_required
def edit_user(request, user_id):
    profile = get_object_or_404(UserProfile, id=user_id)
    
    if request.method == 'POST':
        try:
            user = profile.user
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.save()

            profile.phone_number = request.POST['phone_number']
            profile.address = request.POST['address']
            profile.role = request.POST['role']

            if 'avatar' in request.FILES:
                profile.avatar = request.FILES['avatar']
            
            profile.save()

            # Log the activity
            log_activity(
                user=request.user,
                action='Updated User',
                details=f'Updated user profile for {user.get_full_name()}',
                category='user'
            )

            # Store current page in session
            request.session['current_page'] = 'users_list'
            messages.success(request, 'User profile updated successfully!')
            return redirect('users_list')
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')

    roles = UserProfile.ROLE_CHOICES
    return render(request, 'users/edit_user.html', {
        'profile': profile,
        'roles': roles
    })


@login_required
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        try:
            profile = get_object_or_404(UserProfile, id=user_id)
            user = profile.user
            
            # Toggle the user's active status
            user.is_active = not user.is_active
            user.save()
            
            # Add success message
            messages.success(request, f'User {user.username} has been {"activated" if user.is_active else "deactivated"} successfully.')
            
            return JsonResponse({
                'success': True,
                'is_active': user.is_active,
                'message': f'User {user.username} has been {"activated" if user.is_active else "deactivated"}'
            })
        except Exception as e:
            messages.error(request, f'Error changing user status: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def activity_history(request):
    activities = ActivityLog.objects.select_related('user', 'user__userprofile').all()

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        activities = activities.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(action__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(user__userprofile__enforcer_id__icontains=search_query)
        ).distinct()

    # Get filter parameters
    category = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_id = request.GET.get('user_id', '')

    # Apply filters
    if category:
        activities = activities.filter(category=category)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            activities = activities.filter(timestamp__gte=date_from)
        except ValueError:
            pass

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to = date_to + timedelta(days=1)
            activities = activities.filter(timestamp__lt=date_to)
        except ValueError:
            pass

    if user_id:
        activities = activities.filter(user_id=user_id)

    # Order by timestamp descending (most recent first)
    activities = activities.order_by('-timestamp')

    # Pagination - 25 activities per page
    paginator = Paginator(activities, 25)
    page = request.GET.get('page', 1)
    try:
        activities = paginator.page(page)
    except PageNotAnInteger:
        activities = paginator.page(1)
    except EmptyPage:
        activities = paginator.page(paginator.num_pages)

    # Get all categories and users for filters
    categories = ActivityLog.CATEGORY_CHOICES
    users = User.objects.filter(
        id__in=activities.paginator.object_list.values_list('user_id', flat=True).distinct()
    ).select_related('userprofile')

    context = {
        'activities': activities,
        'categories': categories,
        'users': users,
        'current_category': category,
        'current_user': user_id,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': activities,
        'search_query': search_query,
        'total_count': paginator.count
    }

    return render(request, 'activity_history.html', context)


@login_required
def get_statistics(request, time_range):
    """API endpoint for chart data"""
    today = timezone.now()
    
    if time_range == 'week':
        start_date = today - timedelta(days=7)
        date_format = '%Y-%m-%d'
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
        date_format = '%Y-%m-%d'
    else:  # year
        start_date = today - timedelta(days=365)
        date_format = '%Y-%m'

    # Get violations data
    violations = Violation.objects.filter(
        violation_date__gte=start_date
    ).annotate(
        date=TruncDate('violation_date')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')

    # Format data for charts
    labels = []
    data = []
    
    # Create a date range for all dates in the period
    date_range = []
    delta = today - start_date
    for i in range(delta.days + 1):
        date_range.append(start_date + timedelta(days=i))
    
    # Create a lookup of existing data
    data_lookup = {item['date']: item['count'] for item in violations}
    
    # Fill in data for all dates
    for date in date_range:
        labels.append(date.strftime(date_format))
        data.append(data_lookup.get(date.date(), 0))

    violations_chart = {
        'labels': labels,
        'datasets': [{
            'label': 'Total Violations',
            'data': data,
            'borderColor': '#2563eb',
            'backgroundColor': 'rgba(37, 99, 235, 0.1)',
            'tension': 0.4,
            'fill': True
        }]
    }

    return JsonResponse({
        'violations': violations_chart
    })


@login_required
def violation_detail_modal(request, violation_id):
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Check if it's an NCAP violation by checking for any images
        is_ncap = bool(violation.image) or bool(violation.driver_photo) or bool(violation.vehicle_photo) or bool(violation.secondary_photo)
        
        # Get previous violations for the same violator
        previous_violations = Violation.objects.filter(
            violator=violation.violator
        ).exclude(
            id=violation.id
        ).order_by('-violation_date')
        
        # Find violations from same name violators
        same_name_violations = []
        if violation.violator:
            try:
                # Find violations with the same first and last name
                same_name_violations = Violation.objects.filter(
                    violator__first_name__iexact=violation.violator.first_name,
                    violator__last_name__iexact=violation.violator.last_name
                ).exclude(
                    violator=violation.violator  # Exclude the current violator as they're already in previous_violations
                ).order_by('-violation_date')
            except Exception as e:
                print(f"Error finding same name violations: {e}")
        
        # Find violations from the same submission
        submission_violations = []
        if is_ncap:
            try:
                # First, try to find violations with the same submission_id
                if violation.submission_id:
                    submission_violations = list(Violation.objects.filter(
                        submission_id=violation.submission_id
                    ).order_by('id'))
                    
                    print(f"Found {len(submission_violations)} violations with submission_id {violation.submission_id}")
                
                # If no submission_id or no other violations found with it, fall back to time window approach
                if not submission_violations:
                    # Find violations submitted within a 5-minute window of this violation
                    if violation.violation_date:
                        time_window_start = violation.violation_date - timezone.timedelta(minutes=5)
                        time_window_end = violation.violation_date + timezone.timedelta(minutes=5)
                        
                        # Find all violations from the same time window by the same violator
                        submission_violations = list(Violation.objects.filter(
                            violator=violation.violator,
                            violation_date__gte=time_window_start,
                            violation_date__lte=time_window_end
                        ).order_by('id'))
                        
                        print(f"Fallback: Found {len(submission_violations)} violations within 5 min window for violator {violation.violator.id}")
                
                # Debug each violation found
                for idx, v in enumerate(submission_violations):
                    print(f"  Violation #{idx+1}: ID={v.id}, Type={v.violation_type}, Same as current={v.id == int(violation_id)}")
            except Exception as e:
                print(f"Error finding related violations: {e}")
                # In case of error, just use the current violation
                submission_violations = [violation]
        
        # Use the appropriate template
        template_name = 'violations/ncap_violation_detail_modal.html' if is_ncap else 'violations/violation_detail_modal.html'
        
        context = {
            'violation': violation,
            'is_modal': True,
            'previous_violations': previous_violations,
            'same_name_violations': same_name_violations,
            'submission_violations': submission_violations
        }
        
        return render(request, template_name, context)
    except Exception as e:
        print(f"Error loading violation details: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading violation details: {str(e)}</div>', 
            status=500
        )

@login_required
def search_violators(request):
    """
    API endpoint to search for violators with the same first/last name for the adjudication page
    Returns matching violations in JSON format
    """
    first_name = request.GET.get('first_name', '').strip()
    last_name = request.GET.get('last_name', '').strip()
    current_violation_id = request.GET.get('current_violation', 0)
    
    if not first_name and not last_name:
        return JsonResponse({
            'success': False,
            'message': 'Please provide at least a first name or last name to search',
            'violations': []
        })
    
    # Build query to search for violations with matching violator names
    query = Q()
    if first_name:
        query |= Q(violator__first_name__icontains=first_name)
    if last_name:
        query |= Q(violator__last_name__icontains=last_name)
    
    # Get all violations with the matching name, excluding the current violation
    violations = Violation.objects.filter(query).exclude(id=current_violation_id).select_related('violator')
    
    # Format the results for JSON response
    results = []
    for violation in violations:
        # Format the violation date for display
        violation_date = violation.violation_date.strftime('%b %d, %Y') if violation.violation_date else 'N/A'
        
        # Get a list of violation types
        violation_types = []
        if violation.violation_type:
            violation_types = [vtype.strip() for vtype in violation.violation_type.split(',')]
        
        # Format the violation data
        formatted_violation = {
            'id': violation.id,
            'violation_date': violation_date,
            'violation_type': violation.violation_type,
            'violation_types': violation_types,
            'status': violation.status,
            'status_display': violation.get_status_display(),
            'fine_amount': str(violation.fine_amount),
            'violator_name': f"{violation.violator.first_name} {violation.violator.last_name}"
        }
        results.append(formatted_violation)
    
    return JsonResponse({
        'success': True,
        'message': f"Found {len(results)} violations",
        'violations': results
    })

@login_required
def violations_list(request):
    # Get all violations EXCLUDING NCAP violations (those with any type of image)
    all_violations = Violation.objects.filter(
        Q(image='') | Q(image__isnull=True),
        Q(driver_photo='') | Q(driver_photo__isnull=True),
        Q(vehicle_photo='') | Q(vehicle_photo__isnull=True),
        Q(secondary_photo='') | Q(secondary_photo__isnull=True)
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    # Pagination - 15 items per page
    paginator = Paginator(unique_violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}
    
    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/violations_list.html', context)

@login_required
def update_violation_status(request, violation_id):
    if request.method == 'POST':
        try:
            violation = get_object_or_404(Violation, id=violation_id)
            new_status = request.POST.get('status')
            
            if new_status in dict(Violation.STATUS_CHOICES):
                old_status = violation.status
                violation.status = new_status
                violation.save()
                
                # Log the activity
                log_activity(
                    user=request.user,
                    action='Updated Violation Status',
                    details=f'Changed status of Violation #{violation.id} from {old_status} to {new_status}',
                    category='violation'
                )
                
                messages.success(request, f'Violation status updated successfully to {new_status}')
                return JsonResponse({
                    'status': 'success',
                    'message': f'Violation status updated to {new_status}',
                    'new_status': new_status
                })
            else:
                messages.error(request, 'Invalid status provided')
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid status provided'
                }, status=400)
                
        except Exception as e:
            messages.error(request, f'Error updating violation status: {str(e)}')
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
            
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)

@login_required
def user_detail_modal(request, user_id):
    try:
        # Get the User object first, then its profile
        user = get_object_or_404(User, id=user_id)
        profile = user.userprofile
        
        # Log the user ID for debugging
        logger.debug(f"Fetching user modal for User ID: {user_id}, Username: {user.username}")
        
        context = {
            'profile': profile,
            'user': user,
            'is_modal': True,
            'debug_user_id': user_id  # Add this for debugging
        }
        return render(request, 'users/user_detail.html', context)
    except Exception as e:
        logger.error(f"Error in user_detail_modal for user_id {user_id}: {str(e)}")
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading user details (ID: {user_id}). Error: {str(e)}</div>', 
            status=500
        )

@login_required
def ncap_violations_list(request):
    # Get violations with any images (NCAP violations)
    all_violations = Violation.objects.filter(
        Q(image__isnull=False) | 
        Q(driver_photo__isnull=False) | 
        Q(vehicle_photo__isnull=False) | 
        Q(secondary_photo__isnull=False)
    ).exclude(
        Q(image='') & 
        Q(driver_photo='') & 
        Q(vehicle_photo='') & 
        Q(secondary_photo='')
    ).select_related('violator', 'enforcer').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_violations = all_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Group violations by violator full name
    violations_by_name = {}
    for violation in all_violations:
        full_name = f"{violation.violator.first_name} {violation.violator.last_name}".strip().lower()
        
        if full_name in violations_by_name:
            # Add to existing violator's violations
            violations_by_name[full_name]['all_violations'].append(violation)
            
            # Keep the most recent violation as the representative
            if violation.violation_date > violations_by_name[full_name]['violation'].violation_date:
                violations_by_name[full_name]['violation'] = violation
        else:
            # First time seeing this name
            violations_by_name[full_name] = {
                'violation': violation,
                'all_violations': [violation]
            }
    
    # Convert back to a list of violations (using the most recent per person)
    unique_violations = [info['violation'] for info in violations_by_name.values()]
    
    # Sort by violation date (most recent first)
    unique_violations.sort(key=lambda v: v.violation_date, reverse=True)

    # Pagination - 12 items per page
    paginator = Paginator(unique_violations, 12)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    # Check if coming from the print form
    printed = request.GET.get('printed', False)
    # if printed:
    #     messages.success(request, "The violation was successfully printed.")

    # Create mapping from violations to their counts
    violation_counts = {info['violation'].id: len(info['all_violations']) for info in violations_by_name.values()}

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': len(unique_violations),
        'violation_counts': violation_counts,
        'violations_by_name': violations_by_name
    }
    return render(request, 'violations/ncap_violations_list.html', context)

@login_required
def adjudication_list(request):
    # Get all violations without PAID status
    violations = Violation.objects.select_related('violator', 'enforcer').exclude(status='PAID').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        violations = violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Pagination - 15 items per page
    paginator = Paginator(violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    return render(request, 'violations/adjudication_list.html', context)

@login_required
def adjudication_form(request, violation_id):
    from django.db.models import Q
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    # Get all other violations from the same violator (excluding current one)
    # Now including all statuses, not just PENDING
    previous_violations = Violation.objects.filter(
        violator=violation.violator
    ).exclude(id=violation_id).order_by('-violation_date')
    
    # Also find violations with same license number but different violator records
    if violation.violator and violation.violator.license_number:
        license_violations = Violation.objects.filter(
            violator__license_number=violation.violator.license_number
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).order_by('-violation_date')
        
        # Combine both querysets
        previous_violations = list(previous_violations) + list(license_violations)
    
    # Also find violations with the same name, even if they're different violator records
    if violation.violator:
        name_violations = Violation.objects.filter(
            Q(violator__first_name__iexact=violation.violator.first_name) &
            Q(violator__last_name__iexact=violation.violator.last_name)
        ).exclude(
            Q(id=violation_id) | Q(violator=violation.violator)
        ).exclude(
            id__in=[v.id for v in previous_violations]
        ).order_by('-violation_date')
        
        # Add name-matched violations to the list
        previous_violations = list(previous_violations) + list(name_violations)
        
    # Sort by violation date (most recent first)
    if previous_violations:
        previous_violations.sort(key=lambda v: v.violation_date, reverse=True)
    
    context = {
        'violation': violation,
        'previous_violations': previous_violations
    }
    
    return render(request, 'violations/adjudication_form.html', context)

@login_required
def submit_adjudication(request, violation_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    violation = get_object_or_404(Violation, id=violation_id)
    
    try:
        # Process the main violation
        process_adjudication(request, violation)
        
        # Check if there are batch violations to process
        batch_violation_ids_json = request.POST.get('batch_violation_ids', '[]')
        is_batch = False
        if batch_violation_ids_json and batch_violation_ids_json != '[]':
            batch_violation_ids = json.loads(batch_violation_ids_json)
            is_batch = len(batch_violation_ids) > 0
            
            # Log batch adjudication in the activity log
            if batch_violation_ids:
                ActivityLog.objects.create(
                    user=request.user,
                    action="BATCH_ADJUDICATION",
                    details=f"Batch adjudicated {len(batch_violation_ids)} violations along with violation #{violation_id}",
                    category="violation"
                )
            
            # Process each batch violation
            for batch_id in batch_violation_ids:
                try:
                    batch_violation = Violation.objects.get(id=batch_id)
                    
                    # Get removed violation types for this batch item
                    removed_types_key = f'removed_violation_types_{batch_id}'
                    removed_types_json = request.POST.get(removed_types_key, '[]')
                    
                    # Process the batch violation with its specific removed violations
                    process_adjudication(request, batch_violation, removed_types_json)
                    
                except Violation.DoesNotExist:
                    logger.error(f"Batch violation ID {batch_id} not found during adjudication")
                except Exception as e:
                    logger.error(f"Error processing batch violation {batch_id}: {str(e)}")
        
        # Get settled status for success message
        is_settled = request.POST.get('settled_status', 'false') == 'true'
        
        # Redirect with success parameters
        redirect_url = f"{reverse('adjudication_list')}?success=true"
        if is_settled:
            redirect_url += "&settled=true"
        if is_batch:
            redirect_url += "&batch=true"
            
        return redirect(redirect_url)
    
    except Exception as e:
        logger.error(f"Error in submit_adjudication: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)})

def process_adjudication(request, violation, batch_removed_types_json=None):
    """Helper function to process a single violation adjudication"""
    adjudication_status = request.POST.get('adjudication_status', 'adjudicated')
    requires_approval = request.POST.get('requires_approval', 'true') == 'true'
    settled_status = request.POST.get('settled_status', 'false') == 'true'
    
    # Get form data
    total_penalty = request.POST.get('total_penalty', '0')
    remarks = request.POST.get('remarks', '')
    settlement_reason = request.POST.get('settlement_reason', '')
    
    # Parse removed violations - use batch specific ones if provided
    if batch_removed_types_json:
        removed_violations_json = batch_removed_types_json
    else:
        removed_violations_json = request.POST.get('removed_violations', '[]')
    
    removed_violations = []
    try:
        removed_violations = json.loads(removed_violations_json)
    except json.JSONDecodeError:
        logger.error(f"Error parsing removed violations JSON: {removed_violations_json}")
    
    # Handle payment due date
    payment_due_str = request.POST.get('payment_due', '')
    payment_due_date = None
    if payment_due_str:
        try:
            payment_due_date = datetime.strptime(payment_due_str, '%Y-%m-%d').date()
        except ValueError:
            logger.error(f"Invalid payment due date format: {payment_due_str}")
    
    # Set adjudication status based on form data
    if settled_status:
        violation.status = 'SETTLED'
        violation.fine_amount = 0
        remarks = f"SETTLED: {settlement_reason or remarks}"
    elif len(removed_violations) == len(violation.get_violation_types()):
        violation.status = 'REMOVED'
        violation.fine_amount = 0
        remarks = f"REMOVED: All violations removed. {remarks}"
    else:
        violation.status = 'ADJUDICATED'
        
        # For batch items, recalculate the fine amount proportionally if violations were removed
        if batch_removed_types_json and removed_violations:
            original_types = violation.get_violation_types()
            remaining_count = len(original_types) - len(removed_violations)
            if remaining_count > 0 and len(original_types) > 0:
                # Adjust fine proportionally based on removed violations
                original_fine = float(violation.fine_amount)
                violation.fine_amount = original_fine * (remaining_count / len(original_types))
            else:
                # All violations removed for this batch item
                violation.fine_amount = 0
                violation.status = 'REMOVED'
        else:
            # For main violation, use the submitted total penalty
            violation.fine_amount = float(total_penalty)
            
            # For the main violation, note if it includes batch penalties
            batch_ids_json = request.POST.get('batch_violation_ids', '[]')
            if batch_ids_json and batch_ids_json != '[]':
                try:
                    batch_ids = json.loads(batch_ids_json)
                    if batch_ids and len(batch_ids) > 0:
                        # Get information about batch violations
                        batch_violations = Violation.objects.filter(id__in=batch_ids)
                        total_batch_amount = sum([v.fine_amount for v in batch_violations])
                        
                        # Create detailed batch information
                        batch_details = "\n\n=== BATCH ADJUDICATION SUMMARY ===\n"
                        batch_details += f"This adjudication includes {len(batch_ids)} additional violation(s) with a total of ₱{total_batch_amount:.2f}.\n"
                        batch_details += "Batch violations included:\n"
                        
                        for v in batch_violations:
                            batch_details += f"- Citation #{v.id}: {v.violation_type} (₱{v.fine_amount:.2f}) from {v.violation_date.strftime('%b %d, %Y')}\n"
                        
                        # Add batch details to remarks
                        remarks += batch_details
                except json.JSONDecodeError:
                    logger.error(f"Error parsing batch violation IDs: {batch_ids_json}")
                except Exception as e:
                    logger.error(f"Error processing batch details: {str(e)}")
        
        # Remove specific violations if any were unchecked
        if removed_violations:
            current_types = violation.get_violation_types()
            remaining_types = [t for t in current_types if t not in removed_violations]
            violation.violation_type = ', '.join(remaining_types)
    
    # Record adjudication details
    violation.adjudicated_by = request.user
    violation.adjudication_date = timezone.now()
    violation.adjudication_remarks = remarks
    violation.payment_due_date = payment_due_date
    violation.requires_approval = requires_approval
    
    # Save the updated violation
    violation.save()
    
    # Log the adjudication
    ActivityLog.objects.create(
        user=request.user,
        action="ADJUDICATION",
        details=f"Adjudicated violation #{violation.id} as {violation.status}",
        category="violation"
    )
    
    return violation

@login_required
def announcements_list(request):
    user_role = request.user.userprofile.role
    current_time = timezone.now()
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    priority_filter = request.GET.get('priority', '')
    
    # For admins and supervisors, show all announcements by default including inactive ones
    if user_role in ['ADMIN', 'SUPERVISOR']:
        announcements = Announcement.objects.all().select_related('created_by').order_by('-created_at')
    else:
        # For regular users, show only active announcements
        announcements = Announcement.objects.filter(
            is_active=True
        ).select_related('created_by').order_by('-created_at')
    
    # Apply filters
    if search_query:
        announcements = announcements.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(created_by__first_name__icontains=search_query) |
            Q(created_by__last_name__icontains=search_query) |
            Q(created_by__userprofile__enforcer_id__icontains=search_query)
        ).distinct()
    
    if category_filter:
        announcements = announcements.filter(category=category_filter)
        
    if priority_filter:
        announcements = announcements.filter(priority=priority_filter)
    
    # Pagination - 8 announcements per page
    paginator = Paginator(announcements, 8)
    page = request.GET.get('page', 1)
    try:
        announcements = paginator.page(page)
    except PageNotAnInteger:
        announcements = paginator.page(1)
    except EmptyPage:
        announcements = paginator.page(paginator.num_pages)

    context = {
        'announcements': announcements,
        'page_obj': announcements,
        'search_query': search_query,
        'category_filter': category_filter,
        'priority_filter': priority_filter,
        'total_count': paginator.count,
        'categories': Announcement.CATEGORY_CHOICES,
        'priorities': Announcement.PRIORITY_CHOICES,
        'admin_view': ''  # No longer needed but keep for backward compatibility
    }
    return render(request, 'announcements/announcements_list.html', context)

@login_required
def create_announcement(request):
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, 'You do not have permission to create announcements.')
        return redirect('announcements_list')

    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        priority = request.POST.get('priority')
        category = request.POST.get('category', 'GENERAL')
        target_audience = request.POST.get('target_audience', 'ALL')
        geographic_area = request.POST.get('geographic_area')
        
        # Boolean fields
        is_active = request.POST.get('is_active') == 'on'
        is_popup = request.POST.get('is_popup') == 'on'
        requires_acknowledgment = request.POST.get('requires_acknowledgment') == 'on'
        send_as_notification = request.POST.get('send_as_notification') == 'on'
        
        # Date fields
        publish_date = request.POST.get('publish_date')
        publish_date = timezone.make_aware(datetime.strptime(publish_date, '%Y-%m-%dT%H:%M')) if publish_date else None
        
        expiration_date = request.POST.get('expiration_date')
        expiration_date = timezone.make_aware(datetime.strptime(expiration_date, '%Y-%m-%dT%H:%M')) if expiration_date else None

        if title and content:
            announcement = Announcement.objects.create(
                title=title,
                content=content,
                priority=priority,
                category=category,
                target_audience=target_audience,
                geographic_area=geographic_area,
                is_active=is_active,
                is_popup=is_popup,
                requires_acknowledgment=requires_acknowledgment,
                created_by=request.user,
                publish_date=publish_date,
                expiration_date=expiration_date
            )
            
            # Create notifications if send_as_notification is checked or if it's a popup
            if send_as_notification or is_popup:
                create_notifications_from_announcement(announcement)
                
            # Store current page in session
            request.session['current_page'] = 'announcements_list'
            messages.success(request, 'Announcement created successfully.')
            return redirect('announcements_list')
        else:
            messages.error(request, 'Please fill in all required fields.')

    return render(request, 'announcements/create_announcement.html')

@login_required
def edit_announcement(request, announcement_id):
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, 'You do not have permission to edit announcements.')
        return redirect('announcements_list')

    announcement = get_object_or_404(Announcement, id=announcement_id)

    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        priority = request.POST.get('priority')
        category = request.POST.get('category', 'GENERAL')
        target_audience = request.POST.get('target_audience', 'ALL')
        geographic_area = request.POST.get('geographic_area')
        
        # Boolean fields
        is_active = request.POST.get('is_active') == 'on'
        is_popup = request.POST.get('is_popup') == 'on'
        requires_acknowledgment = request.POST.get('requires_acknowledgment') == 'on'
        send_as_notification = request.POST.get('send_as_notification') == 'on'
        
        # Date fields
        publish_date = request.POST.get('publish_date')
        publish_date = timezone.make_aware(datetime.strptime(publish_date, '%Y-%m-%dT%H:%M')) if publish_date else None
        
        expiration_date = request.POST.get('expiration_date')
        expiration_date = timezone.make_aware(datetime.strptime(expiration_date, '%Y-%m-%dT%H:%M')) if expiration_date else None

        if title and content:
            announcement.title = title
            announcement.content = content
            announcement.priority = priority
            announcement.category = category
            announcement.target_audience = target_audience
            announcement.geographic_area = geographic_area
            announcement.is_active = is_active
            announcement.is_popup = is_popup
            announcement.requires_acknowledgment = requires_acknowledgment
            announcement.publish_date = publish_date
            announcement.expiration_date = expiration_date
            announcement.save()
            
            # Create notifications if send_as_notification is checked or if it's a popup
            if send_as_notification or is_popup:
                create_notifications_from_announcement(announcement)
                
            # Store current page in session
            request.session['current_page'] = 'announcements_list'
            messages.success(request, 'Announcement updated successfully.')
            return redirect('announcements_list')
        else:
            messages.error(request, 'Please fill in all required fields.')

    context = {
        'announcement': announcement
    }
    return render(request, 'announcements/edit_announcement.html', context)

def create_notifications_from_announcement(announcement):
    """Create user notifications from an announcement"""
    try:
        # Determine which users should receive this notification based on target audience
        if announcement.target_audience == 'ALL':
            users = User.objects.all()
        else:
            # Filter users by role matching the target audience
            users = User.objects.filter(userprofile__role=announcement.target_audience)
        
        # Create a notification for each user
        for user in users:
            # Check if notification referencing this announcement already exists for user
            if not UserNotification.objects.filter(user=user, reference_id=announcement.id, type='SYSTEM').exists():
                notification_type = 'SYSTEM'
                if announcement.priority == 'HIGH':
                    notification_type = 'VIOLATION'  # Use violation type for high priority
                
                UserNotification.objects.create(
                    user=user,
                    type=notification_type,
                    message=f"{announcement.title} - {announcement.content[:100]}" + ('...' if len(announcement.content) > 100 else ''),
                    reference_id=announcement.id
                )
        
        return True
    except Exception as e:
        print(f"Error creating notifications from announcement: {str(e)}")
        return False

@login_required
def delete_announcement(request, announcement_id):
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, 'You do not have permission to delete announcements.')
        return redirect('announcements_list')

    announcement = get_object_or_404(Announcement, id=announcement_id)
    announcement.delete()
    messages.success(request, 'Announcement deleted successfully.')
    return redirect('announcements_list')

@login_required
def mark_announcement_seen(request, announcement_id):
    """Mark a popup announcement as seen in the session to prevent it from showing again"""
    try:
        # Initialize seen_popup_announcements list if it doesn't exist
        if 'seen_popup_announcements' not in request.session:
            request.session['seen_popup_announcements'] = []
        
        # Add this announcement ID to the list if not already there
        seen_announcements = request.session['seen_popup_announcements']
        if announcement_id not in seen_announcements:
            seen_announcements.append(announcement_id)
            request.session['seen_popup_announcements'] = seen_announcements
            request.session.modified = True
            print(f"Added announcement {announcement_id} to seen list for user {request.user.username}")
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        print(f"Error marking announcement as seen: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def get_popup_announcement(request):
    """API endpoint to get the latest popup announcement"""
    try:
        user_role = request.user.userprofile.role
        current_time = timezone.now()
        
        # Debug info
        print(f"Checking popup announcements for user {request.user.username}, role: {user_role}")
        
        # Initialize seen announcements list in session if not present
        if 'seen_popup_announcements' not in request.session:
            request.session['seen_popup_announcements'] = []
        
        seen_announcements = request.session['seen_popup_announcements']
        print(f"User has seen these announcements: {seen_announcements}")
        
        # Get the latest active popup announcement for this user's role
        # that hasn't been seen by this user yet
        announcement = Announcement.objects.filter(
            is_active=True, 
            is_popup=True
        ).filter(
            # Check for target audience
            Q(target_audience='ALL') | Q(target_audience=user_role)
        ).filter(
            # Check for publish and expiration dates
            Q(publish_date__isnull=True) | Q(publish_date__lte=current_time)
        ).filter(
            Q(expiration_date__isnull=True) | Q(expiration_date__gt=current_time)
        ).exclude(
            # Exclude announcements already seen in this session
            id__in=seen_announcements
        ).order_by('-created_at').first()
        
        if announcement:
            # Debug
            print(f"Found popup announcement: {announcement.title} (ID: {announcement.id})")
            
            # Check if user needs to acknowledge and has already done so
            acknowledged = False
            if announcement.requires_acknowledgment:
                acknowledged = AnnouncementAcknowledgment.objects.filter(
                    announcement=announcement,
                    user=request.user
                ).exists()
                
                if acknowledged:
                    print(f"User has already acknowledged announcement {announcement.id}")
                    return JsonResponse({'status': 'no_announcements'})
            
            # Increment view count
            announcement.view_count += 1
            announcement.save(update_fields=['view_count'])
            
            # Prepare response data
            data = {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,  # Use content directly without data-* attributes
                'priority': announcement.priority,
                'created_at': announcement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'created_by': f"{announcement.created_by.first_name} {announcement.created_by.last_name}",
                'category': announcement.get_category_display() if announcement.category else None,
                'requires_acknowledgment': announcement.requires_acknowledgment
            }
            return JsonResponse(data)
        else:
            print("No active popup announcements found")
        
        return JsonResponse({'status': 'no_announcements'})
    except Exception as e:
        print(f"Error in get_popup_announcement: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def reset_popup_announcement(request, announcement_id):
    """Reset a popup announcement so it shows again for all users"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, 'You do not have permission to reset popup announcements.')
        return redirect('announcements_list')
        
    announcement = get_object_or_404(Announcement, id=announcement_id)
    
    # Make sure it's a popup announcement
    if not announcement.is_popup:
        messages.error(request, 'This is not a popup announcement.')
        return redirect('announcements_list')
    
    # Update the announcement to force a refresh (updated_at will change)
    announcement.save()
    
    # Clear session data for this announcement for all active sessions
    from django.contrib.sessions.models import Session
    active_sessions = Session.objects.filter(expire_date__gt=timezone.now())
    
    session_count = 0
    for session in active_sessions:
        try:
            session_data = session.get_decoded()
            if 'seen_popup_announcements' in session_data:
                seen_announcements = session_data['seen_popup_announcements']
                if announcement.id in seen_announcements:
                    seen_announcements.remove(announcement.id)
                    session_data['seen_popup_announcements'] = seen_announcements
                    session.session_data = SessionStore().encode(session_data)
                    session.save()
                    session_count += 1
        except:
            # Skip any problematic sessions
            continue
            
    messages.success(request, f'Popup announcement reset for {session_count} active sessions.')
    return redirect('announcements_list')

@login_required
def enforcer_map(request):
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR', 'ENFORCER']:
        messages.error(request, 'You do not have permission to view the enforcer map.')
        return redirect('admin_dashboard')
    
    return render(request, 'enforcer_map.html')

@login_required
@csrf_exempt
def update_location(request):
    # Check if this is a browser request rather than an XHR/fetch request
    is_browser_request = request.headers.get('Accept', '').find('text/html') != -1 and not request.headers.get('X-Requested-With')
    
    # If it's a browser request, redirect to appropriate page instead of showing JSON error
    if is_browser_request:
        messages.error(request, "This endpoint should not be accessed directly in the browser.")
        return redirect('login')
    
    # For GET requests or non-POST requests, return a clear message
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error', 
            'message': 'This endpoint requires POST method'
        }, status=405)
    
    # If user is not an enforcer, return a clear message instead of a generic error
    if request.user.userprofile.role != 'ENFORCER':
        return JsonResponse({
            'status': 'error', 
            'message': 'Location tracking is only available for enforcers'
        }, status=403)

    try:
        data = json.loads(request.body)
        profile = request.user.userprofile
        
        # Update location efficiently
        profile.latitude = data['latitude']
        profile.longitude = data['longitude']
        profile.last_location_update = timezone.now()
        profile.is_active_duty = data.get('is_active', True)
        
        # Save additional data if provided
        if 'accuracy' in data:
            profile.accuracy = data['accuracy']
        if 'heading' in data:
            profile.heading = data['heading']
        if 'speed' in data:
            profile.speed = data['speed']
        if 'battery_level' in data:
            profile.battery_level = data['battery_level']
        if 'device_info' in data:
            profile.device_info = data['device_info']
        
        # Use update_fields to optimize the query
        profile.save()
        
        # Also save to location history for path tracking
        LocationHistory.objects.create(
            user_profile=profile,
            latitude=profile.latitude,
            longitude=profile.longitude,
            accuracy=profile.accuracy if hasattr(profile, 'accuracy') else None,
            heading=profile.heading if hasattr(profile, 'heading') else None,
            speed=profile.speed if hasattr(profile, 'speed') else None,
            is_active_duty=profile.is_active_duty,
            battery_level=profile.battery_level if hasattr(profile, 'battery_level') else None,
            device_info=profile.device_info if hasattr(profile, 'device_info') else None
        )
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        logger.error(f"Error updating location: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def get_enforcer_locations(request):
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR', 'ENFORCER']:
        return JsonResponse({
            'status': 'error',
            'message': 'Permission denied'
        }, status=403)

    # Get all enforcers with their locations, optimized query
    current_time = timezone.now()
    inactive_threshold = current_time - timezone.timedelta(minutes=5)

    # Base query to get all enforcers
    base_query = UserProfile.objects.filter(
        role='ENFORCER',
        latitude__isnull=False,
        longitude__isnull=False
    ).select_related('user').only(
        'id', 'user__first_name', 'user__last_name',
        'enforcer_id', 'role', 'latitude', 'longitude',
        'last_location_update', 'avatar', 'is_active_duty'
    )
    
    # Process all enforcers and mark them as active/inactive based on last update time
    locations = []
    for profile in base_query:
        is_active = profile.is_active_duty and profile.last_location_update and profile.last_location_update > inactive_threshold
        
        locations.append({
            'id': profile.id,
            'name': f"{profile.user.first_name} {profile.user.last_name}",
            'enforcer_id': profile.enforcer_id,
            'role': profile.role,
            'latitude': float(profile.latitude),
            'longitude': float(profile.longitude),
            'last_update': profile.last_location_update.isoformat() if profile.last_location_update else None,
            'avatar_url': profile.avatar.url if profile.avatar else None,
            'is_active': is_active,
            'inactive_time': (current_time - profile.last_location_update).total_seconds() // 60 if profile.last_location_update else None
        })

    return JsonResponse({
        'status': 'success',
        'locations': locations
    })

@login_required
def get_enforcer_path(request):
    """Get historical location data for an enforcer to display their movement path"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR', 'ENFORCER']:
        return JsonResponse({
            'status': 'error',
            'message': 'Permission denied'
        }, status=403)
    
    enforcer_id = request.GET.get('enforcer_id')
    hours = int(request.GET.get('hours', 24))  # Default to 24 hours
    
    if not enforcer_id:
        return JsonResponse({
            'status': 'error',
            'message': 'Enforcer ID is required'
        }, status=400)
    
    try:
        # Get the enforcer profile
        enforcer_profile = UserProfile.objects.get(id=enforcer_id)
        
        # Get location history for the specified time range
        time_threshold = timezone.now() - timezone.timedelta(hours=hours)
        
        path_query = LocationHistory.objects.filter(
            user_profile=enforcer_profile,
            timestamp__gt=time_threshold
        ).order_by('timestamp')
        
        # Format the path data
        path = [{
            'latitude': float(point.latitude),
            'longitude': float(point.longitude),
            'timestamp': point.timestamp.isoformat(),
            'is_active': point.is_active_duty,
            'accuracy': point.accuracy,
            'speed': point.speed,
            'heading': point.heading
        } for point in path_query]
        
        return JsonResponse({
            'status': 'success',
            'enforcer_id': enforcer_id,
            'enforcer_name': f"{enforcer_profile.user.first_name} {enforcer_profile.user.last_name}",
            'hours': hours,
            'path': path
        })
        
    except UserProfile.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Enforcer not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error fetching enforcer path: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }, status=500)

# Add this function to check if user is authorized
def is_payment_recorder(user):
    return user.is_authenticated and user.userprofile.role in ['ADMIN', 'SUPERVISOR', 'ADJUDICATOR']

# Add this view for recording payments
@user_passes_test(is_payment_recorder)
def record_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    
    if request.method == 'POST':
        try:
            receipt_number = request.POST.get('receipt_number')
            receipt_date = request.POST.get('receipt_date')
            remarks = request.POST.get('remarks')
            
            violation.status = 'PAID'
            violation.receipt_number = receipt_number
            violation.receipt_date = receipt_date
            violation.payment_date = timezone.now()
            violation.payment_remarks = remarks
            violation.processed_by = request.user
            violation.payment_amount = violation.fine_amount
            violation.save()
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Recorded Payment',
                details=f'Recorded payment for Violation #{violation.id}',
                category='violation'
            )
            
            messages.success(request, 'Payment recorded successfully.')
            return redirect('payment_processing')
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
    
    context = {
        'violation': violation,
        'today': timezone.now()
    }
    return render(request, 'violations/payment_receipt.html', context)

# Add this view for payment records
@login_required
def payment_records(request):
    # Get all paid violations
    paid_violations = Violation.objects.filter(
        status='PAID'
    ).select_related('violator', 'enforcer').order_by('-receipt_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        paid_violations = paid_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(plate_number__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(receipt_number__icontains=search_query)
        ).distinct()
    
    # Pagination
    paginator = Paginator(paid_violations, 10)
    page = request.GET.get('page')
    
    try:
        paid_violations = paginator.page(page)
    except PageNotAnInteger:
        paid_violations = paginator.page(1)
    except EmptyPage:
        paid_violations = paginator.page(paginator.num_pages)
    
    context = {
        'paid_violations': paid_violations,
        'page_obj': paid_violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    
    return render(request, 'payments/payment_records.html', context)

@login_required
def print_receipt(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    
    if violation.status != 'PAID':
        messages.error(request, 'Cannot print receipt for unpaid violation.')
        return redirect('payment_records')
    
    context = {
        'violation': violation,
        'today': timezone.now()
    }
    return render(request, 'payments/print_receipt.html', context)

@login_required
def payment_detail_modal(request, violation_id):
    """View for showing just payment details and basic violation information in a modal."""
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Check if the violation is paid
        if violation.status != 'PAID':
            return HttpResponse(
                '<div class="alert alert-warning">This violation has not been paid yet.</div>'
            )
        
        # Check if it's an NCAP violation
        is_ncap = bool(violation.image or violation.driver_photo or 
                       violation.vehicle_photo or violation.secondary_photo)
        
        context = {
            'violation': violation,
            'is_ncap': is_ncap
        }
        
        return render(request, 'payments/payment_detail_modal.html', context)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f'<div class="alert alert-danger">Error loading payment details: {str(e)}</div>', 
            status=500
        )

def is_supervisor(user):
    return user.userprofile.role == 'SUPERVISOR' or user.userprofile.role == 'ADMIN'

def is_treasurer(user):
    return user.userprofile.role == 'TREASURER' or user.userprofile.role == 'ADMIN'

@login_required
@user_passes_test(is_supervisor)
def pending_approvals(request):
    # Get violations that are adjudicated but not yet approved/rejected
    violations = Violation.objects.filter(
        status='ADJUDICATED'
    ).select_related('violator', 'enforcer', 'adjudicated_by').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        violations = violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    return render(request, 'violations/pending_approvals.html', context)

@login_required
@user_passes_test(is_supervisor)
def approve_adjudication(request, violation_id):
    if request.method == 'POST':
        violation = get_object_or_404(Violation, id=violation_id)
        try:
            # Check if the fine amount is zero, mark as SETTLED instead of APPROVED
            if violation.fine_amount == 0:
                violation.status = 'SETTLED'
                
                # Log the activity for zero amount (settled) cases
                log_action = 'Approved & Settled'
                log_details = f'Approved and settled Violation #{violation.id} (zero amount)'
            else:
                violation.status = 'APPROVED'
                
                # Log the activity for regular approvals
                log_action = 'Approved Adjudication'
                log_details = f'Approved adjudication for Violation #{violation.id}'
            
            violation.approved_by = request.user
            violation.approval_date = timezone.now()
            violation.save()
            
            # Log the activity
            log_activity(
                user=request.user,
                action=log_action,
                details=log_details,
                category='violation'
            )
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
@user_passes_test(is_supervisor)
def reject_adjudication(request, violation_id):
    if request.method == 'POST':
        try:
            violation = get_object_or_404(Violation, id=violation_id)
            
            # Parse request body as JSON
            data = json.loads(request.body)
            
            # Get rejection reason
            reason = data.get('reason', '')
            
            # Update violation
            violation.status = 'REJECTED'
            violation.rejection_reason = reason
            violation.approved_by = None
            violation.approval_date = None
            violation.save()
            
            # Return success
            return JsonResponse({
                'success': True
            })
        except Exception as e:
            # Return error
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    # Return method not allowed for non-POST requests
    return JsonResponse({
        'success': False,
        'message': 'Method not allowed'
    }, status=405)

@login_required
@user_passes_test(is_treasurer)
def payment_processing(request):
    # Get violations that are approved but not yet paid
    # Exclude violations with zero fine amounts as they're auto-settled
    violations = Violation.objects.filter(
        status='APPROVED'
    ).exclude(
        fine_amount=0
    ).select_related('violator', 'enforcer', 'adjudicated_by').order_by('-violation_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        violations = violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    return render(request, 'violations/payment_processing.html', context)

@login_required
@user_passes_test(is_treasurer)
def record_payment(request, violation_id):
    violation = get_object_or_404(Violation, id=violation_id)
    
    if request.method == 'POST':
        try:
            # Handle JSON request
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                receipt_number = data.get('receipt_number')
                receipt_date = data.get('receipt_date')
                payment_amount = data.get('payment_amount')
                remarks = data.get('remarks')
            else:
                # Handle form data
                receipt_number = request.POST.get('receipt_number')
                receipt_date = request.POST.get('receipt_date')
                payment_amount = request.POST.get('payment_amount')
                remarks = request.POST.get('remarks')
            
            if not all([receipt_number, receipt_date, payment_amount]):
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            violation.status = 'PAID'
            violation.receipt_number = receipt_number
            violation.receipt_date = receipt_date
            violation.payment_date = timezone.now()
            violation.payment_remarks = remarks
            violation.processed_by = request.user
            violation.payment_amount = payment_amount
            violation.save()
            
            # Log the activity
            log_activity(
                user=request.user,
                action='Recorded Payment',
                details=f'Recorded payment for Violation #{violation.id}',
                category='violation'
            )
            
            # Return JSON response for AJAX requests
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': 'Payment recorded successfully'
                })
            
            # For form submissions, redirect
            messages.success(request, 'Payment recorded successfully.')
            return redirect('payment_processing')
            
        except Exception as e:
            # Return JSON error response for AJAX requests
            if request.content_type == 'application/json':
                return JsonResponse({
                    'success': False,
                    'message': f'Error recording payment: {str(e)}'
                }, status=400)
            
            # For form submissions, show error message
            messages.error(request, f'Error recording payment: {str(e)}')
    
    # For GET requests, render the template
    context = {
        'violation': violation,
        'today': timezone.now()
    }
    return render(request, 'violations/payment_receipt.html', context)

# Landing Page Views
def landing_home(request):
    return render(request, 'landing/home.html')

def about(request):
    return render(request, 'landing/about.html')

def services(request):
    return render(request, 'landing/services.html')

def contact(request):
    return render(request, 'contact.html')

@csrf_exempt
def scan_document(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        # Get the image file from request
        image_file = request.FILES.get('image')
        scan_type = request.POST.get('scan_type')
        
        print(f"Received scan request - Type: {scan_type}")
        
        if not image_file:
            return JsonResponse({'success': False, 'message': 'No image provided'})
            
        # Initialize ID Analyzer API
        api_key = os.getenv('ID_ANALYZER_API_KEY')
        if not api_key:
            print("Error: ID Analyzer API key not configured")
            return JsonResponse({'success': False, 'message': 'ID Analyzer API key not configured'})
            
        # Save temporary file
        temp_path = os.path.join('temp', f'doc_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg')
        os.makedirs('temp', exist_ok=True)
        
        print(f"Saving temporary file to: {temp_path}")
        
        with open(temp_path, 'wb+') as destination:
            for chunk in image_file.chunks():
                destination.write(chunk)

        try:
            print("Initializing ID Analyzer API...")
            # Initialize Core API with your api key and region
            coreapi = CoreAPI(api_key, "US")

            # Enable document authentication using quick module
            coreapi.enable_authentication(True, 'quick')
            
            # Enable face detection
            coreapi.face_detection = True
            
            print("Scanning document...")
            # Perform scan
            result = coreapi.scan(document_primary=temp_path)
            print("Scan result:", result)
            
            # Process results based on scan type
            if scan_type == 'license':
                # Extract license information
                if result.get('result'):
                    data_result = result['result']
                    # Properly format the address from address1 field
                    address = data_result.get('address1', '')
                    # Format the full name if individual parts are not available
                    first_name = data_result.get('firstName', '')
                    last_name = data_result.get('lastName', '')
                    if not first_name and not last_name and data_result.get('fullName'):
                        names = data_result['fullName'].split(' ', 1)
                        first_name = names[0]
                        last_name = names[1] if len(names) > 1 else ''

                    response_data = {
                        'success': True,
                        'license_number': data_result.get('documentNumber', ''),
                        'first_name': first_name,
                        'last_name': last_name,
                        'address': address,
                        'phone_number': data_result.get('phone', '')
                    }
                    print("License data extracted:", response_data)
                else:
                    print("No data extracted from license")
                    raise Exception('No data extracted from license')
                    
            else:  # vehicle registration
                # Extract vehicle information
                if result.get('result'):
                    data_result = result['result']
                    # Properly format the address
                    address = data_result.get('address1', data_result.get('address', ''))
                    response_data = {
                        'success': True,
                        'plate_number': data_result.get('vehicle', {}).get('plateNumber', ''),
                        'vehicle_type': data_result.get('vehicle', {}).get('type', ''),
                        'color': data_result.get('vehicle', {}).get('color', ''),
                        'registration_number': data_result.get('documentNumber', ''),
                        'owner_name': data_result.get('fullName', ''),
                        'owner_address': address
                    }
                    print("Vehicle data extracted:", response_data)
                else:
                    print("No data extracted from vehicle registration")
                    raise Exception('No data extracted from vehicle registration')

            # Add authentication results if available
            if result.get('authentication'):
                auth_result = result['authentication']
                response_data['authentication_score'] = auth_result.get('score', 0)
                print("Authentication score:", response_data['authentication_score'])
                
            return JsonResponse(response_data)
            
        except Exception as e:
            print(f"ID Analyzer API error: {str(e)}")
            # Handle API-specific errors
            return JsonResponse({
                'success': False,
                'message': f'ID Analyzer API error: {str(e)}'
            })
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
                print("Temporary file removed")
                
    except Exception as e:
        print(f"Error processing document: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error processing document: {str(e)}'
        })

def track_violation(request):
    """View for tracking violations by ticket number"""
    if request.method == 'POST':
        ticket_number = request.POST.get('ticket_number')
        try:
            violation = Violation.objects.get(id=ticket_number)
            return render(request, 'track_violation.html', {
                'violation': violation,
                'found': True
            })
        except Violation.DoesNotExist:
            return render(request, 'track_violation.html', {
                'error': 'No violation found with this ticket number.',
                'found': False
            })
    return render(request, 'track_violation.html', {'found': None})

@login_required
def save_signature(request):
    if request.method == 'POST' and request.FILES.get('signature'):
        try:
            signature_file = request.FILES['signature']
            # Save to signatures directory
            filename = signature_file.name
            filepath = os.path.join('signatures', filename)
            
            # Ensure signatures directory exists
            os.makedirs('signatures', exist_ok=True)
            
            # Save the file
            with open(filepath, 'wb+') as destination:
                for chunk in signature_file.chunks():
                    destination.write(chunk)
            
            return JsonResponse({
                'success': True,
                'filename': filename,
                'message': 'Signature saved successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error saving signature: {str(e)}'
            }, status=500)
    return JsonResponse({
        'success': False,
        'message': 'Invalid request'
    }, status=400)

@login_required
def get_signature(request, filename):
    try:
        filepath = os.path.join('signatures', filename)
        if os.path.exists(filepath):
            with open(filepath, 'rb') as f:
                return HttpResponse(f.read(), content_type='image/png')
        return HttpResponse('Signature not found', status=404)
    except Exception as e:
        return HttpResponse(f'Error retrieving signature: {str(e)}', status=500)

@login_required
def acknowledge_announcement(request, announcement_id):
    """API endpoint to acknowledge an announcement"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        announcement = get_object_or_404(Announcement, id=announcement_id)
        
        # Record acknowledgment
        AnnouncementAcknowledgment.objects.get_or_create(
            announcement=announcement,
            user=request.user
        )
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def resend_announcement_notification(request, announcement_id):
    """Resend an announcement as a notification to users"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, 'You do not have permission to resend announcement notifications.')
        return redirect('announcements_list')
    
    announcement = get_object_or_404(Announcement, id=announcement_id)
    
    # Send the announcement as a notification
    if create_notifications_from_announcement(announcement):
        messages.success(request, 'Announcement has been sent as notification successfully.')
    else:
        messages.error(request, 'Error sending notifications. Please try again.')
    
    return redirect('announcements_list')

@login_required
def mark_notification_read(request, notification_id):
    """API endpoint to mark a notification as read"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        notification = get_object_or_404(
            UserNotification,
            id=notification_id,
            user=request.user
        )
        notification.is_read = True
        notification.save()
        
        # Count unread notifications for the user
        unread_count = UserNotification.objects.filter(
            user=request.user, 
            is_read=False
        ).count()
        
        return JsonResponse({
            'status': 'success',
            'unread_count': unread_count
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def get_announcement_details(request, announcement_id):
    """API endpoint to get announcement details"""
    try:
        announcement = get_object_or_404(Announcement, id=announcement_id)
        
        # Check if user is allowed to see this announcement based on target audience
        if announcement.target_audience != 'ALL' and announcement.target_audience != request.user.userprofile.role:
            return JsonResponse({'status': 'error', 'message': 'You do not have permission to view this announcement'}, status=403)
        
        # Increment view count
        announcement.view_count += 1
        announcement.save()
        
        # Check if user has acknowledged this announcement
        acknowledged = AnnouncementAcknowledgment.objects.filter(
            announcement=announcement,
            user=request.user
        ).exists()
        
        # Format data for response
        data = {
            'id': announcement.id,
            'title': announcement.title,
            'content': announcement.content,
            'rich_content': announcement.rich_content,
            'priority': announcement.priority,
            'category': announcement.category,
            'created_at': announcement.created_at.isoformat(),
            'created_by': announcement.created_by.get_full_name() or announcement.created_by.username,
            'requires_acknowledgment': announcement.requires_acknowledgment,
            'is_acknowledged': acknowledged,
            'view_count': announcement.view_count
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def save_ncap_violation(request):
    """View to save a new NCAP violation"""
    if request.method == 'POST':
        try:
            form = NCAPViolationForm(request.POST, request.FILES)
            
            # Get selected IDs if present
            selected_operator_id = request.POST.get('selected_operator_id', '')
            selected_driver_id = request.POST.get('selected_driver_id', '')
            
            # Get the unique submission ID (if provided)
            submission_id = request.POST.get('request_id', '')
            if not submission_id:
                submission_id = uuid.uuid4().hex  # Generate a new ID if not provided
            
            # Get violation penalties JSON data
            violation_penalties_json = request.POST.get('violation_penalties', '[]')
            try:
                violation_penalties = json.loads(violation_penalties_json)
                print(f"Loaded violation penalties: {violation_penalties}")
            except json.JSONDecodeError as e:
                violation_penalties = []
                print(f"Error decoding violation penalties: {e}")
                
            # Fallback to single violation if no specific violations found
            if not violation_penalties or len(violation_penalties) == 0:
                single_type = form.cleaned_data.get('violation_type', 'Unspecified Violation')
                print(f"Using single violation type: {single_type}")
                violation_penalties = [{'text': single_type, 'rate': form.cleaned_data.get('fine_amount', 0)}]
            
            print(f"Processing submission with ID: {submission_id}, found {len(violation_penalties)} violations")
            
            # Store all created violations for processing images
            created_violations = []
            
            if form.is_valid():
                for i, violation_data in enumerate(violation_penalties):
                    violation_type = violation_data.get('text', 'Unspecified Violation')
                    
                    # Create a violation instance but don't save yet
                    violation = form.save(commit=False)
                    
                    # Override the violation type with the specific one
                    # Ensure violation_type is not too long even though we've changed to TextField
                    violation.violation_type = violation_type
                    
                    # Set the fine amount from the violation data if available
                    if 'rate' in violation_data:
                        violation.fine_amount = violation_data.get('rate', violation.fine_amount)
                    
                    # Set the submission ID to track this batch
                    violation.submission_id = submission_id
                    
                    # Set the enforcer to the current user
                    violation.enforcer = request.user
                    
                    # Get or create the violator (only once for all violations)
                    if i == 0:
                        # Check if license number is provided
                        license_number = form.cleaned_data.get('license_number', '')
                        
                        if license_number:
                            # If license number is provided, try to find by license number
                            violator, created = Violator.objects.get_or_create(
                                license_number=license_number,
                                defaults={
                                    'first_name': form.cleaned_data['first_name'],
                                    'last_name': form.cleaned_data['last_name']
                                }
                            )
                            
                            # Update violator information if it already exists
                            if not created:
                                violator.first_name = form.cleaned_data['first_name']
                                violator.last_name = form.cleaned_data['last_name']
                                violator.save()
                        else:
                            # If no license number, always create a new violator
                            violator = Violator.objects.create(
                                license_number='',
                                first_name=form.cleaned_data['first_name'],
                                last_name=form.cleaned_data['last_name']
                            )
                    
                    violation.violator = violator
                    
                    # Link to operator if selected (same for all violations)
                    if selected_operator_id:
                        try:
                            operator = Operator.objects.get(pk=selected_operator_id)
                            violation.operator = operator
                            # Don't overwrite fields if they were manually entered in the form
                            if not request.POST.get('operator_name'):
                                violation.operator_name = f"{operator.first_name} {operator.middle_initial + ' ' if operator.middle_initial else ''}{operator.last_name}"
                            else:
                                violation.operator_name = request.POST.get('operator_name')
                            
                            if not request.POST.get('operator_pd_number'):
                                violation.operator_pd_number = operator.new_pd_number
                            else:
                                violation.operator_pd_number = request.POST.get('operator_pd_number')
                            
                            if not request.POST.get('operator_address'):
                                violation.operator_address = operator.address
                            # potpot_number is already set in the create call above

                            violation.save()
                        except Operator.DoesNotExist:
                            pass
                    # Use form data if operator info manually entered
                    elif request.POST.get('operator_name') or request.POST.get('operator_pd_number'):
                        violation.operator_name = request.POST.get('operator_name', '')
                        violation.operator_pd_number = request.POST.get('operator_pd_number', '')
                        violation.save()
                    
                    # Link to driver if selected (same for all violations)
                    if selected_driver_id:
                        try:
                            driver = Driver.objects.get(pk=selected_driver_id)
                            # Set driver information directly
                            violation.driver_name = f"{driver.first_name} {driver.middle_initial + ' ' if driver.middle_initial else ''}{driver.last_name}"
                            violation.pd_number = driver.new_pd_number
                            violation.driver_address = driver.address
                            
                            # Copy other driver fields if available
                            if hasattr(driver, 'novr_number') and driver.novr_number:
                                violation.novr_number = driver.novr_number
                            if hasattr(driver, 'pin_number') and driver.pin_number:
                                violation.pin_number = driver.pin_number
                            if hasattr(driver, 'license_number') and driver.license_number:
                                violation.license_number = driver.license_number
                            
                            violation.save()
                        except Driver.DoesNotExist:
                            pass
                    # Use form data if driver not selected but info provided
                    elif request.POST.get('driver_name'):
                        violation.driver_name = request.POST.get('driver_name')
                        if request.POST.get('novr_number'): violation.novr_number = request.POST.get('novr_number')
                        if request.POST.get('pin_number'): violation.pin_number = request.POST.get('pin_number')
                        if request.POST.get('pd_number'): violation.pd_number = request.POST.get('pd_number')
                        violation.save()
                    
                    # Save the violation
                    violation.save()
                    created_violations.append(violation)
                    
                    print(f"Created violation #{violation.id} of type '{violation_type}' with submission ID: {submission_id}")
                
                # Only process images for the first violation to avoid duplication
                if created_violations:
                    try:
                        # Process images only for the first violation
                        process_violation_images(request, created_violations[0])
                    except Exception as img_error:
                        logger.error(f"Error processing images: {str(img_error)}")
                    
                    # Create certificate if certificate text is provided - only for the first violation
                    if request.POST.get('certificate_text'):
                        try:
                            ViolationCertificate.objects.create(
                                violation=created_violations[0],
                                operations_officer=request.POST.get('operations_officer', ''),
                                ctm_officer=request.POST.get('cttm_officer', ''),
                                certificate_text=request.POST.get('certificate_text', '')
                            )
                        except Exception as cert_error:
                            logger.error(f"Error creating certificate: {str(cert_error)}")
                
                # Check if this is an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Return JSON response for AJAX requests
                    return JsonResponse({
                        'success': True,
                        'message': f"{len(created_violations)} NCAP violation(s) recorded successfully!",
                        'redirect_url': reverse('ncap_violations_list'),
                        'id': created_violations[0].id if created_violations else None
                    })
                else:
                    # Regular form submission
                    messages.success(request, f"{len(created_violations)} NCAP violation(s) recorded successfully!")
                    return redirect('ncap_violations_list')
            else:
                # Form validation failed
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Return JSON with validation errors for AJAX
                    return JsonResponse({
                        'success': False,
                        'message': "Please correct the errors below.",
                        'errors': form.errors.as_json()
                    }, status=400)
                else:
                    messages.error(request, "Please correct the errors below.")
        except Exception as e:
            # Log the error
            logger.error(f"Error in save_ncap_violation: {str(e)}")
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return a JSON response with error details for AJAX requests
                return JsonResponse({
                    'success': False,
                    'message': "An error occurred while processing your request.",
                    'details': str(e)
                }, status=400)
            else:
                # Regular form submission - add error message
                messages.error(request, f"An error occurred: {str(e)}")
    
    # If GET request or form invalid
    form = NCAPViolationForm()
    
    return render(request, 'violations/create_ncap_violation.html', {
        'form': form,
        'title': 'Create NCAP Violation',
        'button_text': 'Record Violation'
    })

@login_required
def create_ncap_violation(request):
    """
    View function to display the form for creating a new NCAP violation
    """
    if request.method == 'POST':
        # The POST handling is done in save_ncap_violation view
        # This is just a fallback in case form is submitted directly to this URL
        return save_ncap_violation(request)
    
    # Create a new empty form
    form = NCAPViolationForm()
    
    # Render the template with the form
    context = {
        'form': form,
        'title': 'Create NCAP Violation',
        'button_text': 'Record Violation'
    }
    
    return render(request, 'violations/create_ncap_violation.html', context)

@login_required
def operator_list(request):
    """View to list all operators with search and pagination"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, "You don't have permission to access the operators management page.")
        return redirect('home')
    
    # Search functionality
    query = request.GET.get('q', '')
    if query:
        operators = Operator.objects.filter(
            Q(last_name__icontains=query) |
            Q(first_name__icontains=query) |
            Q(new_pd_number__icontains=query) |
            Q(old_pd_number__icontains=query) |
            Q(address__icontains=query)
        )
    else:
        operators = Operator.objects.all()
    
    # Clean up middle_initial values - convert 'nan' to None
    for operator in operators:
        if operator.middle_initial == 'nan':
            operator.middle_initial = None
        
    # Pagination
    paginator = Paginator(operators, 10)  # Show 10 operators per page
    page_number = request.GET.get('page')
    operators_page = paginator.get_page(page_number)
    
    context = {
        'operators': operators_page,
        'query': query,
        'title': 'Operators Management',
    }
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action="Viewed operators list",
        category="general"
    )
    
    return render(request, 'operators/operator_list.html', context)


@login_required
def operator_detail(request, pk):
    """View to display detailed information about an operator"""
    if not request.user.is_authenticated:
        messages.error(request, "You must be logged in to view operator details.")
        return redirect('login')
    
    # Get the operator or return 404
    operator = get_object_or_404(Operator, pk=pk)
    
    # Get vehicles associated with this operator
    vehicles = Vehicle.objects.filter(operator=operator).order_by('-created_at')
    
    context = {
        'operator': operator,
        'vehicles': vehicles,
        'title': f'Operator Details: {operator.full_name()}',
    }
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Viewed operator details for: {operator.full_name()}",
        category="general"
    )
    
    return render(request, 'operators/operator_detail.html', context)


@login_required
def operator_create(request):
    """View to create a new operator"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, "You don't have permission to create operators.")
        return redirect('operator_list')
    
    if request.method == 'POST':
        form = OperatorForm(request.POST)
        if form.is_valid():
            operator = form.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Created operator: {operator.last_name}, {operator.first_name}",
                category="user"
            )
            
            messages.success(request, f"Operator {operator.last_name}, {operator.first_name} created successfully.")
            return redirect('operator_list')
    else:
        form = OperatorForm()
    
    context = {
        'form': form,
        'title': 'Create Operator',
    }
    
    return render(request, 'operators/operator_form.html', context)


@login_required
def operator_update(request, pk):
    """View to update an existing operator"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, "You don't have permission to update operators.")
        return redirect('operator_list')
    
    operator = get_object_or_404(Operator, pk=pk)
    
    if request.method == 'POST':
        form = OperatorForm(request.POST, instance=operator)
        if form.is_valid():
            operator = form.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Updated operator: {operator.last_name}, {operator.first_name}",
                category="user"
            )
            
            messages.success(request, f"Operator {operator.last_name}, {operator.first_name} updated successfully.")
            return redirect('operator_list')
    else:
        form = OperatorForm(instance=operator)
    
    context = {
        'form': form,
        'operator': operator,
        'title': 'Update Operator',
    }
    
    return render(request, 'operators/operator_form.html', context)


@login_required
def operator_delete(request, pk):
    """View to disable an operator"""
    if not request.user.userprofile.role in ['ADMIN']:
        messages.error(request, "You don't have permission to disable operators.")
        return redirect('operator_list')
    
    operator = get_object_or_404(Operator, pk=pk)
    
    if request.method == 'POST':
        operator_name = f"{operator.last_name}, {operator.first_name}"
        
        try:
            # Check if there are any active vehicle assignments
            active_assignments = Vehicle.objects.filter(operator=operator, active=True).exists()
            if active_assignments:
                return JsonResponse({
                    'success': False, 
                    'error': 'Cannot disable operator with active vehicle assignments. Please end all vehicle assignments first.'
                })
            
            # Disable the operator instead of deleting
            operator.active = False
            operator.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Disabled operator: {operator_name}",
                category="user"
            )
            
            messages.success(request, f"Operator {operator_name} has been disabled successfully.")
        except Exception as e:
            messages.error(request, f"Error disabling operator: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        
        return JsonResponse({'success': True})
    
    # If it's not a POST request, return a 405 Method Not Allowed response
    return JsonResponse({
        'success': False, 
        'error': 'Method not allowed'
    }, status=405)


@login_required
def operator_vehicles(request, operator_id):
    """View to display all vehicles associated with an operator"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR', 'CLERK', 'TREASURER']:
        messages.error(request, "You don't have permission to view operator vehicles.")
        return redirect('operator_list')
    
    operator = get_object_or_404(Operator, pk=operator_id)
    vehicles = Vehicle.objects.filter(operator=operator)
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Viewed vehicles for operator: {operator.last_name}, {operator.first_name}",
        category="user"
    )
    
    context = {
        'operator': operator, 
        'vehicles': vehicles,
        'title': f"{operator.first_name} {operator.last_name}'s Vehicles",
    }
    
    return render(request, 'operators/operator_vehicles.html', context)


@login_required
def operator_import(request):
    """View to import operators from an Excel or CSV file"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, "You don't have permission to import operators.")
        return redirect('operator_list')
    
    if request.method == 'POST':
        form = OperatorImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            preview_only = request.POST.get('preview_only') == 'true'
            
            # Get file extension to determine file type
            extension = file.name.split('.')[-1].lower()
            
            try:
                # Import necessary modules
                import re
                import sys
                import io
                import pandas as pd
                from django.http import JsonResponse
                
                # Check if this is an AJAX request
                is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                
                # Process the file based on its type
                if extension == 'csv':
                    # Process CSV file
                    data = []
                    decoded_file = file.read().decode('utf-8')
                    io_string = io.StringIO(decoded_file)
                    reader = csv.DictReader(io_string)
                    for row in reader:
                        data.append(row)
                elif extension in ['xlsx', 'xls']:
                    # Process Excel file
                    try:
                        # First try reading with headers
                        df = pd.read_excel(file)
                    except Exception as e:
                        try:
                            # Try reading without headers
                            df = pd.read_excel(file, header=None)
                            # Assign default column names
                            df.columns = [f"Column_{i}" for i in range(df.shape[1])]
                        except Exception as e2:
                            error_msg = f"Error reading Excel file: {str(e)}"
                            if is_ajax:
                                return JsonResponse({"error": error_msg}, status=400)
                            messages.error(request, error_msg)
                            return redirect('operator_import')
                    
                    # Debug info
                    print(f"DataFrame shape: {df.shape}", file=sys.stderr)
                    print(f"Column names: {df.columns.tolist()}", file=sys.stderr)
                    print(f"First 3 rows:\n{df.head(3)}", file=sys.stderr)
                    
                    # Check if first row might contain column headers
                    if len(df) > 0:
                        first_row = df.iloc[0].tolist()
                        print(f"First row (potential headers): {first_row}", file=sys.stderr)
                    
                        # Check if the first row looks like headers (strings instead of numbers)
                        headers_detected = all([not isinstance(x, (int, float)) or pd.isna(x) for x in first_row if not pd.isna(x)])
                        if headers_detected and any(['NAME' in str(x).upper() for x in first_row if not pd.isna(x)]):
                            # First row looks like headers, use it as column names
                            df.columns = [str(x).strip() if not pd.isna(x) else f"Column_{i}" for i, x in enumerate(first_row)]
                            df = df.iloc[1:].reset_index(drop=True)
                            print(f"Using first row as headers. New columns: {df.columns.tolist()}", file=sys.stderr)
                    
                    # Define patterns for better column matching
                    column_patterns = {
                        'last_name': ['LAST NAME', 'LASTNAME', 'LAST', 'SURNAME', 'FAMILY NAME', 'LN', 'FAMILY'],
                        'first_name': ['FIRST NAME', 'FIRSTNAME', 'FIRST', 'GIVEN NAME', 'FN', 'GIVEN', 'NAME'],
                        'middle_initial': ['MIDDLE INITIAL', 'MI', 'M.I.', 'MIDDLE', 'MIDDLE NAME', 'M.I', 'INIT', 'INITIAL'],
                        'address': ['ADDRESS', 'ADDR', 'LOCATION', 'RESIDENCE', 'HOME ADDRESS', 'LOC', 'RES', 'HOME'],
                        'old_pd_number': ['OLD PD NO', 'OLD P.D. NO.', 'OLD P.D. NO', 'OLD PD', 'OLD PO NO', 'OLD P.O. NO.', 'OLD P.O. NO', 'OLD PO', 'PREVIOUS PD', 'OLD ID', 'PREVIOUS ID', 'OLD P.D NO', 'OLD', 'OLD NUMBER', 'PREV', 'PREVIOUS'],
                        'new_pd_number': ['NEW PD NO', 'NEW P.D. NO.', 'NEW P.D. NO', 'NEW PD', 'NEW PO NO', 'NEW P.O. NO.', 'NEW P.O. NO', 'NEW PO', 'CURRENT PD', 'NEW ID', 'CURRENT ID', 'PD NUMBER', 'ID NUMBER', 'NEW', 'PD NO', 'P.D. NO', 'PD', 'ID', 'NUMBER', 'NO']
                    }
                    
                    # Clean column names for better matching
                    clean_columns = {}
                    for col in df.columns:
                        # Convert to uppercase, remove special characters, and trim whitespace
                        clean_col = re.sub(r'[^A-Z0-9 ]', '', str(col).upper()).strip()
                        clean_columns[col] = clean_col
                    
                    print(f"Cleaned column names: {clean_columns}", file=sys.stderr)
                    
                    # Match each column to the most similar pattern
                    column_mapping = {}
                    
                    # First try exact matches
                    for field, patterns in column_patterns.items():
                        for col, clean_col in clean_columns.items():
                            for pattern in patterns:
                                if clean_col == pattern:
                                    column_mapping[field] = col
                                    break
                    
                    # If we're missing any fields, try partial matches
                    for field in column_patterns:
                        if field not in column_mapping:
                            for col, clean_col in clean_columns.items():
                                # Skip columns already mapped
                                if col in column_mapping.values():
                                    continue
                            
                                for pattern in column_patterns[field]:
                                    if pattern in clean_col or clean_col in pattern:
                                        column_mapping[field] = col
                                        break
                    
                    # Last resort: for fields still missing, use positional assumptions
                    required_positions = {
                        'last_name': 0,     # First column
                        'first_name': 1,    # Second column
                        'middle_initial': 2, # Third column
                        'address': 3,      # Fourth column
                        'old_pd_number': 4, # Fifth column
                        'new_pd_number': 5  # Sixth column
                    }
                    
                    for field in column_patterns:
                        if field not in column_mapping and field in required_positions:
                            pos = required_positions[field]
                            if pos < len(df.columns):
                                column_mapping[field] = df.columns[pos]
                    
                    print(f"Detected column mapping: {column_mapping}", file=sys.stderr)
                    
                    # Check if all required columns are present
                    required_columns = ['last_name', 'first_name', 'new_pd_number']
                    missing_columns = [col for col in required_columns if col not in column_mapping]
                    
                    if missing_columns:
                        # Print more debug info to help diagnose the issue
                        error_msg = f"Could not find columns for: {', '.join(missing_columns)}. Please check your file has columns for these fields."
                        available_columns = f"Available columns in your file: {', '.join(df.columns.tolist())}"
                        
                        if is_ajax:
                            return JsonResponse({
                                "error": error_msg,
                                "details": available_columns
                            }, status=400)
                        
                        messages.error(request, error_msg)
                        messages.info(request, available_columns)
                        return redirect('operator_import')
                    
                    # Create a new DataFrame with mapped columns
                    named_df = pd.DataFrame()
                    for field, col in column_mapping.items():
                        named_df[field] = df[col]
                    
                    # Debug output
                    print(f"Named columns: {named_df.columns.tolist()}", file=sys.stderr)
                    print(f"First 3 rows after mapping:\n{named_df.head(3)}", file=sys.stderr)
                    
                    # Process data and validate
                    data = []
                    for _, row in named_df.iterrows():
                        # Convert numeric values to strings and handle NaN values
                        last_name = str(row.get('last_name', '')).strip() if pd.notna(row.get('last_name')) else ''
                        first_name = str(row.get('first_name', '')).strip() if pd.notna(row.get('first_name')) else ''
                        middle_initial = str(row.get('middle_initial', '')).strip() if pd.notna(row.get('middle_initial')) else ''
                        address = str(row.get('address', '')).strip() if pd.notna(row.get('address')) else ''
                        old_pd_number = str(row.get('old_pd_number', '')).strip() if pd.notna(row.get('old_pd_number')) else ''
                        new_pd_number = str(row.get('new_pd_number', '')).strip() if pd.notna(row.get('new_pd_number')) else ''
                        
                        # Skip empty or invalid rows
                        if not last_name or not first_name or not new_pd_number:
                            continue
                            
                        # Create record
                        operator_data = {
                            'last_name': last_name,
                            'first_name': first_name,
                            'middle_initial': middle_initial,
                            'address': address,
                            'old_pd_number': old_pd_number,
                            'new_pd_number': new_pd_number,
                            'status': 'new',  # Default status
                            'note': ''
                        }
                        
                        # Check if the new_pd_number already exists
                        existing = Operator.objects.filter(new_pd_number=new_pd_number).first()
                        if existing:
                            operator_data['status'] = 'update'
                            operator_data['note'] = 'Will update existing operator data'
                        
                        data.append(operator_data)
                else:
                    messages.error(request, f"Unsupported file format: {extension}. Please upload a CSV or Excel file.")
                    return redirect('operator_import')
                
                # Debug info for the processed data
                print(f"Processed {len(data)} records for import")
                
                # If no data was processed, show error
                if not data:
                    error_msg = "No valid data found in the file. Please check the format and try again."
                    if is_ajax:
                        return JsonResponse({
                            "success": True,
                            "total": 0,
                            "new": 0,
                            "updated": 0,
                            "errors": 0,
                            "message": error_msg
                        })
                    messages.warning(request, error_msg)
                    return redirect('operator_import')
                
                # Summary of found data
                new_count = sum(1 for item in data if item['status'] == 'new')
                update_count = sum(1 for item in data if item['status'] == 'update')
                
                # If this is just a preview, return the data for display
                if preview_only:
                    context = {
                        'data': data,
                        'title': 'Preview Imported Operators',
                        'form': form,
                        'new_count': new_count,
                        'update_count': update_count,
                        'error_count': 0
                    }
                
                    # Store data in session for later processing
                    request.session['import_data'] = data
                
                    if is_ajax:
                        return JsonResponse({
                            "success": True,
                            "preview": data[:50],  # Send only first 50 records to avoid large response
                            "total": len(data),
                            "new": new_count,
                            "updated": update_count,
                            "errors": 0
                        })
                
                    return render(request, 'operators/operator_import_preview.html', context)
                
                # If this is final import, process the data
                success_count = 0
                error_count = 0
                error_messages = []
                
                try:
                    for row in data:
                        try:
                            last_name = row.get('last_name')
                            first_name = row.get('first_name')
                            middle_initial = row.get('middle_initial')
                            address = row.get('address')
                            old_pd_number = row.get('old_pd_number')
                            new_pd_number = row.get('new_pd_number')
                            
                            # Skip empty rows
                            if not last_name or not first_name or not new_pd_number:
                                continue
                            
                            # Check if the operator already exists
                            existing = Operator.objects.filter(new_pd_number=new_pd_number).first()
                            if existing:
                                # Update existing operator
                                existing.last_name = last_name
                                existing.first_name = first_name
                                existing.middle_initial = middle_initial
                                existing.address = address
                                existing.old_pd_number = old_pd_number
                                existing.save()
                            else:
                                # Create new operator
                                Operator.objects.create(
                                    last_name=last_name,
                                    first_name=first_name,
                                    middle_initial=middle_initial,
                                    address=address,
                                    old_pd_number=old_pd_number,
                                    new_pd_number=new_pd_number
                                )
                            
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            error_messages.append(f"Error with row {success_count + error_count}: {str(e)}")
                            continue
                    
                    # Log activity
                    ActivityLog.objects.create(
                        user=request.user,
                        action=f"Imported operators: {success_count} successful, {error_count} failed",
                        category="user"
                    )
                    
                    # Clear the session data
                    if 'import_data' in request.session:
                        del request.session['import_data']
                    
                    if is_ajax:
                        return JsonResponse({
                            "success": True,
                            "total": success_count + error_count,
                            "new": new_count,
                            "updated": update_count,
                            "errors": error_count,
                            "message": f"Import completed: {success_count} operators imported successfully, {error_count} failed."
                        })
                    
                    if error_count > 0:
                        # Log the first few errors to help with debugging
                        error_details = "; ".join(error_messages[:5])
                        if len(error_messages) > 5:
                            error_details += f"; and {len(error_messages) - 5} more errors"
                        messages.warning(request, f"Import completed with issues: {success_count} operators imported successfully, {error_count} failed. Errors: {error_details}")
                    else:
                        messages.success(request, f"Import completed: {success_count} operators imported successfully.")
                    
                    return redirect('operator_list')
                except Exception as e:
                    error_msg = f"Error processing file: {str(e)}"
                    import traceback
                    print(traceback.format_exc())
                    
                    if is_ajax:
                        return JsonResponse({"error": error_msg}, status=500)
                        
                    messages.error(request, error_msg)
                    return redirect('operator_import')
            except Exception as e:
                error_msg = f"Error processing file: {str(e)}"
                import traceback
                print(traceback.format_exc())
                
                if is_ajax:
                    return JsonResponse({"error": error_msg}, status=500)
                    
                messages.error(request, error_msg)
                return redirect('operator_import')
    else:
        form = OperatorImportForm()
    
    context = {
        'form': form,
        'title': 'Import Operators',
    }
    
    return render(request, 'operators/operator_import.html', context)


@login_required
def operator_import_confirm(request):
    """View to save imported operators after preview"""
    if not request.user.userprofile.role in ['ADMIN', 'SUPERVISOR']:
        messages.error(request, "You don't have permission to import operators.")
        return redirect('operator_list')
    
    if request.method == 'POST':
        import_data = request.session.get('import_data', [])
        if not import_data:
            messages.error(request, "No import data found. Please try again.")
            return redirect('operator_import')
        
        # Process the data
        success_count = 0
        error_count = 0
        error_messages = []
        
        for row in import_data:
            try:
                # With our cleaned data, we can directly access the lowercase keys
                last_name = row.get('last_name')
                first_name = row.get('first_name')
                middle_initial = row.get('middle_initial')
                address = row.get('address')
                old_pd_number = row.get('old_pd_number')
                new_pd_number = row.get('new_pd_number')
                
                # Skip empty rows
                if not last_name or not first_name or not new_pd_number:
                    continue
                
                # Check if the operator already exists
                existing = Operator.objects.filter(new_pd_number=new_pd_number).first()
                if existing:
                    # Update existing operator
                    existing.last_name = last_name
                    existing.first_name = first_name
                    existing.middle_initial = middle_initial
                    existing.address = address
                    existing.old_pd_number = old_pd_number
                    existing.save()
                else:
                    # Create new operator
                    Operator.objects.create(
                        last_name=last_name,
                        first_name=first_name,
                        middle_initial=middle_initial,
                        address=address,
                        old_pd_number=old_pd_number,
                        new_pd_number=new_pd_number
                    )
                
                success_count += 1
            
            except Exception as e:
                error_count += 1
                error_messages.append(f"Error with row {success_count + error_count}: {str(e)}")
                continue
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action=f"Imported operators: {success_count} successful, {error_count} failed",
            category="user"
        )
        
        # Clear the session data
        del request.session['import_data']
        
        if error_count > 0:
            # Log the first few errors to help with debugging
            error_details = "; ".join(error_messages[:5])
            if len(error_messages) > 5:
                error_details += f"; and {len(error_messages) - 5} more errors"
            messages.warning(request, f"Import completed with issues: {success_count} operators imported successfully, {error_count} failed. Errors: {error_details}")
        else:
            messages.success(request, f"Import completed: {success_count} operators imported successfully.")
        
        return redirect('operator_list')
    
    return redirect('operator_import')


@login_required
def operator_export_excel(request):
    """View to export all operators data to Excel file"""
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Operators Data"
    
    # Add headers
    headers = ['Last Name', 'First Name', 'Middle Initial', 'Address', 'Old PD Number', 'New PD Number']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    # Get all operators from database
    operators = Operator.objects.all().order_by('last_name', 'first_name')
    
    # Add data rows
    for row_num, operator in enumerate(operators, 2):  # Start from row 2 (after header)
        row = [
            operator.last_name,
            operator.first_name,
            operator.middle_initial if operator.middle_initial else '',
            operator.address,
            operator.old_pd_number if operator.old_pd_number else '',
            operator.new_pd_number
        ]
        
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    column_widths = [20, 20, 15, 40, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=operators_data.xlsx'
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Exported operators data to Excel",
        category="export"
    )
    
    # Save the workbook to the response
    workbook.save(response)
    return response

@login_required
def api_search_operators(request):
    """API endpoint to search for operators by PD number"""
    query = request.GET.get('q', '')
    
    if not query:
        return JsonResponse([], safe=False)
        
    # Search for operators by new_pd_number, old_pd_number, po_number, or name
    operators = Operator.objects.filter(
        Q(new_pd_number__icontains=query) | 
        Q(old_pd_number__icontains=query) | 
        Q(po_number__icontains=query) |
        Q(last_name__icontains=query) |
        Q(first_name__icontains=query)
    )[:10]  # Limit to 10 results
    
    # Format the results
    results = []
    for operator in operators:
        # Get the vehicle information for this operator
        vehicles = Vehicle.objects.filter(operator=operator)
        vehicle_data = []
        
        for vehicle in vehicles:
            vehicle_data.append({
                'id': vehicle.id,
                'new_pd_number': vehicle.new_pd_number,
                'old_pd_number': vehicle.old_pd_number,
                'vehicle_type': vehicle.vehicle_type,
                'plate_number': vehicle.plate_number,
                'color': vehicle.color,
                'classification': vehicle.vehicle_type  # Use vehicle_type as classification
            })
        
        operator_data = {
            'id': operator.id,
            'first_name': operator.first_name,
            'last_name': operator.last_name,
            'middle_initial': operator.middle_initial,
            'address': operator.address,
            'old_pd_number': operator.old_pd_number,
            'new_pd_number': operator.new_pd_number,
            'po_number': operator.po_number,
            'vehicles': vehicle_data
        }
        
        results.append(operator_data)
    
    return JsonResponse(results, safe=False)

@login_required
def api_get_operator(request):
    """API endpoint to get operator details by ID"""
    operator_id = request.GET.get('id')
    
    if not operator_id:
        return JsonResponse({'error': 'No operator ID provided'}, status=400)
    
    try:
        operator = Operator.objects.get(id=operator_id)
        
        # Get the vehicle information for this operator
        vehicles = Vehicle.objects.filter(operator=operator)
        vehicle_data = []
        
        for vehicle in vehicles:
            vehicle_data.append({
                'id': vehicle.id,
                'new_pd_number': vehicle.new_pd_number,
                'old_pd_number': vehicle.old_pd_number,
                'vehicle_type': vehicle.vehicle_type,
                'plate_number': vehicle.plate_number,
                'color': vehicle.color,
                'classification': vehicle.vehicle_type  # Use vehicle_type as classification
            })
        
        operator_data = {
            'id': operator.id,
            'first_name': operator.first_name,
            'last_name': operator.last_name,
            'middle_initial': operator.middle_initial,
            'address': operator.address,
            'old_pd_number': operator.old_pd_number,
            'new_pd_number': operator.new_pd_number,
            'po_number': operator.po_number,
            'vehicles': vehicle_data
        }
        
        return JsonResponse(operator_data)
    
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'Operator not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_search_drivers(request):
    """API endpoint to search for drivers by PD number"""
    query = request.GET.get('q', '')
    
    if not query:
        return JsonResponse([], safe=False)
        
    # Search for drivers by new_pd_number, old_pd_number, or name
    drivers = Driver.objects.filter(
        Q(new_pd_number__icontains=query) | 
        Q(old_pd_number__icontains=query) | 
        Q(last_name__icontains=query) |
        Q(first_name__icontains=query)
    )[:10]  # Limit to 10 results
    
    # Format the results
    results = []
    for driver in drivers:
        driver_data = {
            'id': driver.id,
            'first_name': driver.first_name,
            'last_name': driver.last_name,
            'middle_initial': driver.middle_initial,
            'address': driver.address,
            'old_pd_number': driver.old_pd_number,
            'new_pd_number': driver.new_pd_number,
            'license_number': driver.license_number,
            'po_number': getattr(driver, 'po_number', None),  # Include if available
            'contact_number': getattr(driver, 'contact_number', None)
        }
        
        # Add operator info if available
        if driver.operator:
            # Get the vehicle information for this operator
            vehicles = Vehicle.objects.filter(operator=driver.operator)
            vehicle_data = []
            
            for vehicle in vehicles:
                vehicle_data.append({
                    'id': vehicle.id,
                    'new_pd_number': vehicle.new_pd_number,
                    'old_pd_number': vehicle.old_pd_number,
                    'vehicle_type': vehicle.vehicle_type,
                    'plate_number': vehicle.plate_number,
                    'color': vehicle.color,
                    'classification': vehicle.vehicle_type  # Use vehicle_type as classification
                })
            
            driver_data['operator'] = {
                'id': driver.operator.id,
                'first_name': driver.operator.first_name,
                'last_name': driver.operator.last_name,
                'middle_initial': driver.operator.middle_initial,
                'new_pd_number': driver.operator.new_pd_number,
                'old_pd_number': driver.operator.old_pd_number,
                'po_number': driver.operator.po_number,
                'address': driver.operator.address,
                'vehicles': vehicle_data
            }
        
        results.append(driver_data)
    
    return JsonResponse(results, safe=False)

@login_required
def api_get_driver(request):
    """API endpoint to get driver details by ID"""
    driver_id = request.GET.get('id')
    
    if not driver_id:
        return JsonResponse({'error': 'No driver ID provided'}, status=400)
    
    try:
        driver = Driver.objects.get(id=driver_id)
        
        # Get the vehicle information for related driver assignments
        current_vehicles = []
        for assignment in DriverVehicleAssignment.objects.filter(driver=driver, end_date__isnull=True):
            current_vehicles.append({
                'id': assignment.vehicle.id,
                'new_pd_number': assignment.vehicle.new_pd_number,
                'old_pd_number': assignment.vehicle.old_pd_number,
                'vehicle_type': assignment.vehicle.vehicle_type,
                'plate_number': assignment.vehicle.plate_number,
                'color': assignment.vehicle.color,
                'classification': assignment.vehicle.vehicle_type  # Use vehicle_type as classification
            })
        
        driver_data = {
            'id': driver.id,
            'first_name': driver.first_name,
            'last_name': driver.last_name,
            'middle_initial': driver.middle_initial,
            'address': driver.address,
            'old_pd_number': driver.old_pd_number,
            'new_pd_number': driver.new_pd_number,
            'license_number': driver.license_number,
            'po_number': getattr(driver, 'po_number', None),
            'contact_number': getattr(driver, 'contact_number', None),
            'vehicles': current_vehicles
        }
        
        # Add operator info if available
        if driver.operator:
            # Get the vehicle information for this operator
            vehicles = Vehicle.objects.filter(operator=driver.operator)
            vehicle_data = []
            
            for vehicle in vehicles:
                vehicle_data.append({
                    'id': vehicle.id,
                    'new_pd_number': vehicle.new_pd_number,
                    'old_pd_number': vehicle.old_pd_number,
                    'vehicle_type': vehicle.vehicle_type,
                    'plate_number': vehicle.plate_number,
                    'color': vehicle.color,
                    'classification': vehicle.vehicle_type
                })
            
            driver_data['operator'] = {
                'id': driver.operator.id,
                'first_name': driver.operator.first_name,
                'last_name': driver.operator.last_name,
                'middle_initial': driver.operator.middle_initial,
                'new_pd_number': driver.operator.new_pd_number,
                'old_pd_number': driver.operator.old_pd_number,
                'po_number': driver.operator.po_number,
                'address': driver.operator.address,
                'vehicles': vehicle_data
            }
        
        return JsonResponse(driver_data)
    
    except Driver.DoesNotExist:
        return JsonResponse({'error': 'Driver not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting driver: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def process_violation_images(request, violation):
    """Process uploaded images for a violation"""
    try:
        # Get image files from request
        driver_photo = request.FILES.get('driver_photo')
        vehicle_photo = request.FILES.get('vehicle_photo')
        secondary_photo = request.FILES.get('secondary_photo')
        primary_image = request.FILES.get('image')
        
        # Maximum file size (10MB)
        MAX_FILE_SIZE = 10 * 1024 * 1024
        
        # Allowed file types
        ALLOWED_TYPES = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        
        # Process each image if it exists
        if driver_photo:
            try:
                # Validate file size
                if driver_photo.size > MAX_FILE_SIZE:
                    logger.warning(f"Driver photo too large: {driver_photo.size} bytes")
                    # We'll continue instead of raising an exception
                elif driver_photo.content_type not in ALLOWED_TYPES:
                    logger.warning(f"Invalid file type for driver photo: {driver_photo.content_type}")
                    # We'll continue instead of raising an exception
                else:
                    violation.driver_photo = driver_photo
                    logger.info(f"Driver photo processed for violation {violation.id}")
            except Exception as e:
                logger.error(f"Error processing driver photo: {str(e)}")
            
        if vehicle_photo:
            try:
                # Validate file size
                if vehicle_photo.size > MAX_FILE_SIZE:
                    logger.warning(f"Vehicle photo too large: {vehicle_photo.size} bytes")
                    # We'll continue instead of raising an exception
                elif vehicle_photo.content_type not in ALLOWED_TYPES:
                    logger.warning(f"Invalid file type for vehicle photo: {vehicle_photo.content_type}")
                    # We'll continue instead of raising an exception
                else:
                    violation.vehicle_photo = vehicle_photo
                    logger.info(f"Vehicle photo processed for violation {violation.id}")
            except Exception as e:
                logger.error(f"Error processing vehicle photo: {str(e)}")
            
        if secondary_photo:
            try:
                # Validate file size
                if secondary_photo.size > MAX_FILE_SIZE:
                    logger.warning(f"Secondary photo too large: {secondary_photo.size} bytes")
                    # We'll continue instead of raising an exception
                elif secondary_photo.content_type not in ALLOWED_TYPES:
                    logger.warning(f"Invalid file type for secondary photo: {secondary_photo.content_type}")
                    # We'll continue instead of raising an exception
                else:
                    violation.secondary_photo = secondary_photo
                    logger.info(f"Secondary photo processed for violation {violation.id}")
            except Exception as e:
                logger.error(f"Error processing secondary photo: {str(e)}")
                
        if primary_image:
            try:
                # Validate file size
                if primary_image.size > MAX_FILE_SIZE:
                    logger.warning(f"Primary image too large: {primary_image.size} bytes")
                    # We'll continue instead of raising an exception
                elif primary_image.content_type not in ALLOWED_TYPES:
                    logger.warning(f"Invalid file type for primary image: {primary_image.content_type}")
                    # We'll continue instead of raising an exception
                else:
                    violation.image = primary_image
                    logger.info(f"Primary image processed for violation {violation.id}")
            except Exception as e:
                logger.error(f"Error processing primary image: {str(e)}")
            
        # Save the violation with the attached images
        violation.save()
        logger.info(f"All images processed and saved for violation {violation.id}")
        
    except Exception as e:
        logger.error(f"Error processing violation images: {str(e)}")
        # Don't raise the exception, just log it and continue
        pass

def process_date_time(date_str, time_str):
    """Process date and time strings into a datetime object"""
    try:
        if date_str:
            date_obj = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            
            if time_str:
                # Parse time if provided
                time_obj = timezone.datetime.strptime(time_str, '%H:%M').time()
                return timezone.datetime.combine(date_obj, time_obj)
            else:
                # Default to midnight if no time is provided
                return timezone.datetime.combine(date_obj, timezone.datetime.min.time())
        
        # Default to now if no date is provided
        return timezone.now()
    except Exception as e:
        logger.error(f"Error processing date/time: {str(e)}")
        return timezone.now()

@login_required
def vehicle_list(request):
    """View to list all vehicles for an operator"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to view vehicles.")
        return redirect('operator_dashboard')
    
    # Get all vehicles for this operator
    vehicles = Vehicle.objects.filter(operator=operator).order_by('-created_at')
    
    return render(request, 'operators/vehicle_list.html', {
        'vehicles': vehicles,
        'operator': operator
    })


@login_required
def driver_list(request):
    """View to list all drivers for an operator"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to view drivers.")
        return redirect('operator_dashboard')
    
    # Get all drivers for this operator
    drivers = Driver.objects.filter(operator=operator).order_by('last_name', 'first_name')
    
    return render(request, 'operators/driver_list.html', {
        'drivers': drivers,
        'operator': operator
    })

@login_required
def operator_vehicle_list(request):
    """View to list all vehicles for an operator"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to view vehicles.")
        return redirect('operator_dashboard')
    
    # Get all vehicles for this operator
    vehicles = Vehicle.objects.filter(operator=operator).order_by('-created_at')
    
    return render(request, 'operators/vehicle_list.html', {
        'vehicles': vehicles,
        'operator': operator
    })


@login_required
def operator_register_vehicle(request):
    """View for operators to register a new vehicle"""
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to add vehicles.")
        return redirect('operator_dashboard')
    
    # Get user's personal vehicles
    personal_vehicles = []
    try:
        from traffic_violation_system.user_portal.models import VehicleRegistration
        personal_vehicles = VehicleRegistration.objects.filter(user=request.user)
    except:
        pass  # Fail silently if VehicleRegistration model doesn't exist
    
    if request.method == 'POST':
        form = VehicleForm(request.POST, operator=operator)
        if form.is_valid():
            vehicle = form.save(commit=False)
            vehicle.operator = operator
            vehicle.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Registered new vehicle with PD number {vehicle.new_pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Vehicle successfully registered with PD number {vehicle.new_pd_number}")
            return redirect('operator_dashboard')  # Redirect to dashboard instead of vehicle list
    else:
        # Check if we're converting a personal vehicle
        personal_vehicle_id = request.GET.get('personal_vehicle_id')
        if personal_vehicle_id:
            try:
                from traffic_violation_system.user_portal.models import VehicleRegistration
                personal_vehicle = VehicleRegistration.objects.get(id=personal_vehicle_id, user=request.user)
                # Pre-populate the form with data from personal vehicle
                initial_data = {
                    'vehicle_type': 'Other',  # Default to 'Other' since the choices may be different
                    'plate_number': personal_vehicle.plate_number,
                    'color': personal_vehicle.color,
                    'year_model': personal_vehicle.year_model,
                    'capacity': 4,  # Default capacity
                    'notes': f"Converted from personal vehicle. Make: {personal_vehicle.make}, Model: {personal_vehicle.model}"
                }
                form = VehicleForm(operator=operator, initial=initial_data)
            except (VehicleRegistration.DoesNotExist, ImportError):
                form = VehicleForm(operator=operator)
        else:
            form = VehicleForm(operator=operator)
    
    return render(request, 'operators/vehicle_form.html', {
        'form': form,
        'is_create': True,
        'title': 'Register New Vehicle',
        'submit_text': 'Register Vehicle',
        'personal_vehicles': personal_vehicles
    })


@login_required
def operator_edit_vehicle(request, vehicle_id):
    """View for operators to edit a vehicle"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to edit vehicles.")
        return redirect('operator_dashboard')
    
    # Get the vehicle or return 404
    vehicle = get_object_or_404(Vehicle, id=vehicle_id, operator=operator)
    
    if request.method == 'POST':
        form = VehicleForm(request.POST, instance=vehicle, operator=operator)
        if form.is_valid():
            form.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Updated vehicle with PD number {vehicle.new_pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Vehicle details updated successfully")
            return redirect('vehicle_list')
    else:
        form = VehicleForm(instance=vehicle, operator=operator)
    
    return render(request, 'operators/vehicle_form.html', {
        'form': form,
        'is_create': False,
        'title': 'Edit Vehicle',
        'submit_text': 'Update Vehicle',
        'vehicle': vehicle
    })


@login_required
def operator_delete_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    
    # Check if the user is the operator of this vehicle or an admin
    if not request.user.is_staff and request.user.userprofile.operator != vehicle.operator:
        messages.error(request, "You don't have permission to delete this vehicle.")
        return redirect('operator_dashboard')
    
    if request.method == 'POST':
        vehicle.delete()
        messages.success(request, "Vehicle deleted successfully.")
        return redirect('operator_dashboard')
    
    return render(request, 'operators/delete_vehicle.html', {'vehicle': vehicle})


@login_required
def operator_driver_list(request):
    """View to list all drivers for an operator"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to view drivers.")
        return redirect('operator_dashboard')
    
    # Get all drivers for this operator
    drivers = Driver.objects.filter(operator=operator).order_by('last_name', 'first_name')
    
    return render(request, 'operators/driver_list.html', {
        'drivers': drivers,
        'operator': operator
    })


@login_required
def operator_register_driver(request):
    """View for operators to register a new driver"""
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to add drivers.")
        return redirect('operator_dashboard')
    
    if request.method == 'POST':
        form = DriverForm(request.POST, operator=operator)
        if form.is_valid():
            driver = form.save(commit=False)
            driver.operator = operator
            driver.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Registered new driver {driver.get_full_name()} with PD number {driver.new_pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Driver {driver.get_full_name()} successfully registered with PD number {driver.new_pd_number}")
            return redirect('driver_detail', pk=driver.id)
    else:
        form = DriverForm(operator=operator)
    
    return render(request, 'operators/driver_form.html', {
        'form': form,
        'is_create': True,
        'title': 'Register New Driver',
        'submit_text': 'Register Driver',
        'cancel_url': 'operator_dashboard'  # Add cancel URL for the form
    })


@login_required
def operator_edit_driver(request, driver_id):
    """View for operators to edit a driver"""
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to edit drivers.")
        return redirect('operator_dashboard')
    
    driver = get_object_or_404(Driver, id=driver_id, operator=operator)
    
    if request.method == 'POST':
        form = DriverForm(request.POST, instance=driver, operator=operator)
        if form.is_valid():
            form.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Updated driver {driver.get_full_name()} with PD number {driver.new_pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Driver {driver.get_full_name()} updated successfully")
            return redirect('operator_dashboard')  # Redirect to dashboard
    else:
        form = DriverForm(instance=driver, operator=operator)
    
    return render(request, 'operators/driver_form.html', {
        'form': form,
        'driver': driver,
        'is_create': False,
        'title': f'Edit Driver: {driver.get_full_name()}',
        'submit_text': 'Save Changes',
        'cancel_url': 'operator_dashboard'  # Add cancel URL for the form
    })


@login_required
def operator_delete_driver(request, driver_id):
    """View for operators to disable a driver"""
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to disable drivers.")
        return redirect('operator_dashboard')
    
    driver = get_object_or_404(Driver, id=driver_id, operator=operator)
    
    if request.method == 'POST':
        driver_name = driver.get_full_name()
        pd_number = driver.new_pd_number
        
        try:
            # Check if there are any active vehicle assignments
            active_assignments = DriverVehicleAssignment.objects.filter(driver=driver, is_active=True).exists()
            if active_assignments:
                return JsonResponse({
                    'success': False, 
                    'error': 'Cannot disable driver with active vehicle assignments. Please end all vehicle assignments first.'
                })
            
            # Disable the driver instead of deleting
            driver.active = False
            driver.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Disabled driver {driver_name} with PD number {pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Driver {driver_name} has been disabled successfully.")
        except Exception as e:
            messages.error(request, f"Error disabling driver: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        
        return JsonResponse({'success': True})
    
    # If it's not a POST request, return a 405 Method Not Allowed response
    return JsonResponse({
        'success': False, 
        'error': 'Method not allowed'
    }, status=405)


@login_required
def operator_assign_driver_to_vehicle(request):
    """View for operators to assign a driver to a vehicle"""
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to assign drivers to vehicles.")
        return redirect('operator_dashboard')
    
    # Pre-select driver or vehicle if provided in query parameters
    initial_data = {}
    if request.GET.get('driver'):
        try:
            driver_id = int(request.GET.get('driver'))
            driver = get_object_or_404(Driver, id=driver_id, operator=operator)
            initial_data['driver'] = driver.id
        except (ValueError, TypeError):
            pass
    
    if request.GET.get('vehicle'):
        try:
            vehicle_id = int(request.GET.get('vehicle'))
            vehicle = get_object_or_404(Vehicle, id=vehicle_id, operator=operator)
            initial_data['vehicle'] = vehicle.id
        except (ValueError, TypeError):
            pass
    
    if request.method == 'POST':
        form = DriverVehicleAssignmentForm(request.POST, operator=operator)
        if form.is_valid():
            driver = form.cleaned_data['driver']
            vehicle = form.cleaned_data['vehicle']
            
            # Check if there's already an active assignment for this vehicle
            existing_assignment = DriverVehicleAssignment.objects.filter(
                vehicle=vehicle,
                is_active=True
            ).first()
            
            if existing_assignment:
                # End the existing assignment
                existing_assignment.end_assignment()
                messages.info(request, f"Previous assignment of {existing_assignment.driver.get_full_name()} to this vehicle has been ended.")
            
            # Create the new assignment
            assignment = form.save(commit=False)
            assignment.created_by = request.user
            assignment.save()
            
            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Assigned driver {driver.get_full_name()} to vehicle {vehicle.new_pd_number}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Driver {driver.get_full_name()} has been assigned to vehicle {vehicle.new_pd_number}")
            return redirect('operator_dashboard')  # Redirect to dashboard instead of assignment list
    else:
        form = DriverVehicleAssignmentForm(operator=operator, initial=initial_data)
    
    return render(request, 'operators/assignment_form.html', {
        'form': form,
        'title': 'Assign Driver to Vehicle',
        'submit_text': 'Assign Driver'
    })


@login_required
def operator_end_driver_assignment(request, assignment_id):
    """View for operators to end a driver-vehicle assignment"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to manage driver assignments.")
        return redirect('operator_dashboard')
    
    # Get the assignment or return 404
    assignment = get_object_or_404(
        DriverVehicleAssignment, 
        id=assignment_id,
        vehicle__operator=operator,
        active=True
    )
    
    if request.method == 'POST':
        # End the assignment
        driver_name = assignment.driver.get_full_name()
        vehicle_pd = assignment.vehicle.new_pd_number
        
        assignment.end_assignment()
        
        # Log the activity
        ActivityLog.objects.create(
            user=request.user,
            action=f"Ended assignment of driver {driver_name} from vehicle {vehicle_pd}",
            ip_address=get_client_ip(request)
        )
        
        messages.success(request, f"Driver {driver_name} has been unassigned from vehicle {vehicle_pd}")
        return redirect('assignment_list')
    
    return render(request, 'operators/end_assignment.html', {
        'assignment': assignment
    })


@login_required
def operator_assignment_list(request):
    """View to list all driver-vehicle assignments for an operator"""
    # Check if user is an operator
    try:
        operator = Operator.objects.get(user=request.user)
    except Operator.DoesNotExist:
        messages.error(request, "You must be a registered operator to view assignments.")
        return redirect('operator_dashboard')
    
    # Get active and inactive assignments
    active_assignments = DriverVehicleAssignment.objects.filter(
        vehicle__operator=operator,
        is_active=True
    ).select_related('driver', 'vehicle').order_by('-start_date')
    
    past_assignments = DriverVehicleAssignment.objects.filter(
        vehicle__operator=operator,
        is_is_active=False
    ).select_related('driver', 'vehicle').order_by('-end_date')[:10]  # Limit to last 10
    
    return render(request, 'operators/assignment_list.html', {
        'active_assignments': active_assignments,
        'past_assignments': past_assignments,
        'operator': operator
    })

# Add a simple redirect for driver_list
@login_required
def driver_list(request):
    """Simple redirect to operator dashboard"""
    return redirect('operator_dashboard')

# Update the driver_list redirect to check for admin access
@login_required
def driver_list(request):
    """Enhanced driver list view that checks for admin access"""
    # If user is admin/staff, use the admin driver list view
    if request.user.is_staff:
        return redirect('admin_driver_list')
    # Otherwise redirect to operator dashboard for non-admin users
    return redirect('operator_dashboard')

@login_required
def admin_driver_list(request):
    """View to display a list of all drivers with search and pagination"""
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to view the list of drivers.")
        return redirect('user_portal:user_dashboard')
    
    # Get search query
    search_query = request.GET.get('q', '')
    
    # Get all drivers ordered by last name, first name
    drivers = Driver.objects.all().order_by('last_name', 'first_name')
    
    # Apply search filter if provided
    if search_query:
        drivers = drivers.filter(
            Q(last_name__icontains=search_query) | 
            Q(first_name__icontains=search_query) |
            Q(new_pd_number__icontains=search_query) |
            Q(old_pd_number__icontains=search_query)
        )
    
    # Get current date for comparison with expiration dates
    current_date = timezone.now().date()
    date_30_days_later = current_date + timedelta(days=30)
    
    # Add expiration status for each driver
    for driver in drivers:
        driver.expiration_status = driver.get_expiration_status()
    
    # Pagination - 20 drivers per page
    paginator = Paginator(drivers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'drivers/driver_list.html', {
        'drivers': page_obj,
        'query': search_query,
        'now': current_date,
        'date_30_days_later': date_30_days_later
    })

@login_required
def driver_detail(request, pk):
    """View to display detailed information about a driver"""
    if not request.user.is_staff and not hasattr(request.user, 'operator'):
        messages.error(request, "You don't have permission to view driver details.")
        return redirect('user_portal:user_dashboard')
    
    try:
        driver = Driver.objects.get(pk=pk)
        
        # If user is not staff, check if they are the operator of this driver
        if not request.user.is_staff:
            try:
                operator = Operator.objects.get(user=request.user)
                if driver.operator != operator:
                    messages.error(request, "You don't have permission to view this driver.")
                    return redirect('operator_dashboard')
            except Operator.DoesNotExist:
                messages.error(request, "You don't have permission to view driver details.")
                return redirect('user_portal:user_dashboard')
        
        # Get current date for comparison with expiration date
        current_date = timezone.now().date()
        expiration_status = driver.get_expiration_status()
        
        return render(request, 'drivers/driver_detail.html', {
            'driver': driver,
            'now': current_date,
            'expiration_status': expiration_status
        })
        
    except Driver.DoesNotExist:
        messages.error(request, "Driver not found.")
        return redirect('admin_driver_list')

@user_passes_test(lambda u: u.is_staff)
def admin_report_list(request):
    """View all user reports in admin panel with filtering and search options."""
    # Get search parameters
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    report_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Start with all reports
    reports = UserReport.objects.all().select_related('user', 'assigned_to')
    
    # Apply filters
    if q:
        reports = reports.filter(
            Q(subject__icontains=q) | 
            Q(description__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )
    
    if status:
        reports = reports.filter(status=status)
    
    if report_type:
        reports = reports.filter(type=report_type)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            reports = reports.filter(created_at__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            reports = reports.filter(created_at__lte=date_to)
        except ValueError:
            pass
    
    # Order by most recent
    reports = reports.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(reports, 15)  # Show 15 reports per page
    page = request.GET.get('page', 1)
    
    try:
        reports = paginator.page(page)
    except PageNotAnInteger:
        reports = paginator.page(1)
    except EmptyPage:
        reports = paginator.page(paginator.num_pages)
    
    context = {
        'reports': reports,
        'status_choices': UserReport.STATUS_CHOICES,
        'report_types': UserReport.REPORT_TYPES,
        'current_filters': {
            'q': q,
            'status': status,
            'type': report_type,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/reports/report_list_partial.html', context)
    
    return render(request, 'admin/reports/report_list.html', context)

@user_passes_test(lambda u: u.is_staff)
def admin_report_view(request, report_id):
    """View a single report in a modal or detailed page."""
    report = get_object_or_404(UserReport, id=report_id)
    
    context = {
        'report': report,
    }
    
    # If this is an AJAX request, return the modal content
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/reports/report_modal.html', context)
    
    # Otherwise, return the full page
    return render(request, 'admin/reports/report_detail.html', context)

@user_passes_test(lambda u: u.is_staff)
def admin_report_dispute(request, report_id):
    """Dispute a report with an optional reason."""
    report = get_object_or_404(UserReport, id=report_id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        
        # Update report status to disputed
        report.status = 'CLOSED'
        report.resolution_notes = reason
        report.resolved_at = timezone.now()
        report.assigned_to = request.user
        report.save()
        
        # Create notification for the user
        UserNotification.objects.create(
            user=report.user,
            type='STATUS',
            message=f'Your report "{report.subject}" has been disputed. Reason: {reason}'
        )
        
        # Return a JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Report has been disputed successfully.',
                'status': 'CLOSED',
                'resolved_at': report.resolved_at.strftime('%Y-%m-%d %H:%M'),
                'assigned_to': report.assigned_to.get_full_name() or report.assigned_to.username
            })
        
        messages.success(request, 'Report has been disputed successfully.')
        return redirect('admin_report_list')
    
    # If this is a GET request, return the dispute form
    context = {
        'report': report,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/reports/report_dispute_modal.html', context)
    
    return render(request, 'admin/reports/report_dispute.html', context)

@user_passes_test(lambda u: u.is_staff)
def admin_report_update_status(request, report_id):
    """Update the status of a report."""
    report = get_object_or_404(UserReport, id=report_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        # Validate the status
        valid_statuses = [status for status, _ in UserReport.STATUS_CHOICES]
        if status not in valid_statuses:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid status value.',
                }, status=400)
            
            messages.error(request, 'Invalid status value.')
            return redirect('admin_report_list')
        
        # Update the report
        report.status = status
        
        # If resolving or closing, add resolution details
        if status in ['RESOLVED', 'CLOSED']:
            report.resolution_notes = notes
            report.resolved_at = timezone.now()
            report.assigned_to = request.user
        else:
            # For other statuses, just update the assigned admin
            report.assigned_to = request.user
        
        report.save()
        
        # Create notification for the user
        status_display = dict(UserReport.STATUS_CHOICES).get(status, status)
        UserNotification.objects.create(
            user=report.user,
            type='STATUS',
            message=f'Your report "{report.subject}" status has been updated to {status_display}.'
        )
        
        # Return a JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Report status updated to {status_display}.',
                'status': status,
                'status_display': status_display,
                'resolved_at': report.resolved_at.strftime('%Y-%m-%d %H:%M') if report.resolved_at else None,
                'assigned_to': report.assigned_to.get_full_name() or report.assigned_to.username
            })
        
        messages.success(request, f'Report status updated to {status_display}.')
        return redirect('admin_report_list')
    
    # If this is a GET request, return the status update form
    context = {
        'report': report,
        'status_choices': UserReport.STATUS_CHOICES,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/reports/report_status_modal.html', context)
    
    return render(request, 'admin/reports/report_status.html', context)

@user_passes_test(lambda u: u.is_staff)
def admin_report_export(request):
    """Export reports to Excel."""
    import xlwt
    from django.http import HttpResponse
    
    # Get filtered reports based on query parameters
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    report_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Start with all reports
    reports = UserReport.objects.all().select_related('user', 'assigned_to')
    
    # Apply filters
    if q:
        reports = reports.filter(
            Q(subject__icontains=q) | 
            Q(description__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )
    
    if status:
        reports = reports.filter(status=status)
    
    if report_type:
        reports = reports.filter(type=report_type)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            reports = reports.filter(created_at__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            reports = reports.filter(created_at__lte=date_to)
        except ValueError:
            pass
    
    # Create a workbook and add a worksheet
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Reports')
    
    # Define styles
    header_style = xlwt.easyxf('font: bold on; pattern: pattern solid, fore_colour gray25;')
    date_style = xlwt.easyxf(num_format_str='dd/mm/yyyy hh:mm')
    
    # Write header row
    header_row = [
        'ID', 'User', 'Type', 'Subject', 'Description', 'Location', 
        'Incident Date', 'Status', 'Created At', 'Updated At', 
        'Assigned To', 'Resolution Notes', 'Resolved At'
    ]
    
    for col_num, column_title in enumerate(header_row):
        worksheet.write(0, col_num, column_title, header_style)
    
    # Write data rows
    for row_num, report in enumerate(reports, 1):
        # Format dates and handle nulls
        incident_date = report.incident_date.strftime('%Y-%m-%d %H:%M') if report.incident_date else ''
        created_at = report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else ''
        updated_at = report.updated_at.strftime('%Y-%m-%d %H:%M') if report.updated_at else ''
        resolved_at = report.resolved_at.strftime('%Y-%m-%d %H:%M') if report.resolved_at else ''
        
        # Get report type and status display names
        report_type_display = dict(UserReport.REPORT_TYPES).get(report.type, report.type)
        status_display = dict(UserReport.STATUS_CHOICES).get(report.status, report.status)
        
        # Format user and assigned_to names
        user_name = f"{report.user.get_full_name()} ({report.user.username})" if report.user else ''
        assigned_to = f"{report.assigned_to.get_full_name()} ({report.assigned_to.username})" if report.assigned_to else ''
        
        # Write the data row
        row = [
            report.id,
            user_name,
            report_type_display,
            report.subject,
            report.description,
            report.location,
            incident_date,
            status_display,
            created_at,
            updated_at,
            assigned_to,
            report.resolution_notes,
            resolved_at
        ]
        
        for col_num, cell_value in enumerate(row):
            worksheet.write(row_num, col_num, cell_value)
    
    # Set column widths
    column_widths = [10, 30, 20, 40, 50, 30, 20, 15, 20, 20, 30, 50, 20]
    for i, width in enumerate(column_widths):
        worksheet.col(i).width = 256 * width  # 256 is the width of one character
    
    # Create the HttpResponse with Excel content type
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reports.xls'
    
    # Save the workbook to the response
    workbook.save(response)
    return response

@login_required
def print_violation_form(request, violation_id):
    """View for printing a violation form/ticket"""
    violation = get_object_or_404(Violation, id=violation_id)
    
    # Check if this is an NCAP violation (with images)
    is_ncap = bool(violation.image) or bool(violation.driver_photo) or bool(violation.vehicle_photo) or bool(violation.secondary_photo)
    
    # Choose the appropriate template based on violation type
    template_name = 'violations/ncap_print_form.html' if is_ncap else 'violations/ticket_template.html'
    
    # Add the 'now' context variable for the timestamp in the template
    from django.utils import timezone
    
    context = {
        'violation': violation,
        'now': timezone.now()
    }
    
    # Comment out success message
    # messages.success(request, "The violation was successfully printed.")
    
    return render(request, template_name, context)

@user_passes_test(lambda u: u.is_superuser)
def debug_media(request):
    """Diagnostic view to check media files and paths (only accessible to superusers)"""
    import os
    from django.conf import settings
    
    media_root = settings.MEDIA_ROOT
    media_url = settings.MEDIA_URL
    storage_class = settings.DEFAULT_FILE_STORAGE
    
    # Check media directories
    media_dirs = {
        'media_root': {
            'path': media_root,
            'exists': os.path.exists(media_root),
            'is_dir': os.path.isdir(media_root) if os.path.exists(media_root) else False,
            'permissions': oct(os.stat(media_root).st_mode & 0o777) if os.path.exists(media_root) else None,
        }
    }
    
    # Check subdirectories
    for subdir in ['avatars', 'signatures', 'driver_photos', 'vehicle_photos', 'secondary_photos', 'qr_codes']:
        subdir_path = os.path.join(media_root, subdir)
        media_dirs[subdir] = {
            'path': subdir_path,
            'exists': os.path.exists(subdir_path),
            'is_dir': os.path.isdir(subdir_path) if os.path.exists(subdir_path) else False,
            'permissions': oct(os.stat(subdir_path).st_mode & 0o777) if os.path.exists(subdir_path) else None,
        }
    
    # Check for profile images
    from traffic_violation_system.models import UserProfile
    profiles = UserProfile.objects.all()
    profile_images = []
    
    for profile in profiles:
        if profile.avatar:
            try:
                image_info = {
                    'user': profile.user.username,
                    'field_value': str(profile.avatar),
                    'url': profile.avatar.url if hasattr(profile.avatar, 'url') else None,
                    'path': profile.avatar.path if hasattr(profile.avatar, 'path') else None,
                    'exists': os.path.exists(profile.avatar.path) if hasattr(profile.avatar, 'path') else False,
                    'size': os.path.getsize(profile.avatar.path) if hasattr(profile.avatar, 'path') and os.path.exists(profile.avatar.path) else None,
                    'obj': profile.avatar  # Add the actual avatar object for our template tag
                }
                profile_images.append(image_info)
            except Exception as e:
                profile_images.append({
                    'user': profile.user.username,
                    'field_value': str(profile.avatar),
                    'error': str(e)
                })
    
    context = {
        'media_url': media_url,
        'storage_class': storage_class,
        'media_dirs': media_dirs,
        'profile_images': profile_images,
        'profile_count': profiles.count(),
        'profiles_with_avatar': sum(1 for p in profiles if p.avatar),
    }
    
    return render(request, 'debug/media_debug.html', context)

@login_required
def operator_apply(request):
    if request.user.userprofile.is_operator:
        messages.info(request, 'You are already registered as an operator.')
        return redirect('operator_dashboard')
    
    # Check if user already has a pending application
    has_pending_application = OperatorApplication.objects.filter(
        user=request.user,
        status='PENDING'
    ).exists()
    
    if has_pending_application and not request.path == '/operator/application/status/':
        messages.info(request, 'You already have a pending operator application. Please check the status before submitting another application.')
        return redirect('operator_application_status')
            
    if request.method == 'POST':
        form = OperatorApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.user = request.user
            
            # Store the number of units in extra_data as JSON
            import json
            number_of_units = request.POST.get('number_of_units', '1')
            vehicle_type = request.POST.get('vehicle_type', 'Potpot')
            
            try:
                number_of_units = int(number_of_units)
                if number_of_units < 1:
                    number_of_units = 1
                elif number_of_units > 100:
                    number_of_units = 100
            except (ValueError, TypeError):
                number_of_units = 1
                
            application.extra_data = json.dumps({
                'number_of_units': number_of_units,
                'vehicle_type': vehicle_type
            })
            application.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Submitted operator application",
                category="user"
            )
            
            # Use URL parameter instead of Django messages
            return redirect(f"{reverse('operator_application_status')}?application=success")
    else:
        form = OperatorApplicationForm()
            
    return render(request, 'operators/operator_apply.html', {
        'form': form,
        'has_pending_application': has_pending_application
    })

@login_required
def operator_application_status(request):
    applications = OperatorApplication.objects.filter(user=request.user).order_by('-submitted_at')
    
    return render(request, 'operators/operator_application_status.html', {
        'applications': applications,
        'title': 'Operator Application Status'
    })

@login_required
def operator_dashboard(request):
    if not request.user.userprofile.is_operator:
        messages.error(request, 'You are not registered as an operator.')
        return redirect('user_portal:user_dashboard')
            
    # Get operator-specific information
    try:
        operator = Operator.objects.get(user=request.user)
        vehicles = Vehicle.objects.filter(operator=operator)
        
        # Get drivers associated with this operator - use direct relationship
        drivers = Driver.objects.filter(operator=operator).distinct()
        
        # Get violations for this operator's vehicles
        violations = Violation.objects.filter(
            operator=operator
        ).order_by('-violation_date')
        
        violations_count = violations.count()
        
        # Get vehicles without active driver assignments
        unassigned_vehicles = Vehicle.objects.filter(
            operator=operator,
            active=True
        ).exclude(
            id__in=DriverVehicleAssignment.objects.filter(
                is_active=True
            ).values('vehicle_id')
        ).order_by('new_pd_number')
        
        unassigned_count = unassigned_vehicles.count()
        
    except Operator.DoesNotExist:
        # Fix the inconsistency: The user has is_operator=True but no operator profile
        # This can happen if an operator was deleted in the admin without updating the user profile
        user_profile = request.user.userprofile
        user_profile.is_operator = False
        user_profile.operator_since = None
        user_profile.save()
        
        # Log the issue
        ActivityLog.objects.create(
            user=request.user,
            action=f"Inconsistent operator status detected and fixed",
            details="User had is_operator=True but no Operator record existed. Status reset to non-operator.",
            category="user"
        )
        
        messages.warning(request, 'Your operator profile was not found. Your status has been reset. If you believe this is an error, please contact support.')
        return redirect('user_portal:user_dashboard')
    
    return render(request, 'operators/operator_dashboard.html', {
        'vehicles': vehicles,
        'drivers': drivers,
        'violations': violations,
        'violations_count': violations_count,
        'unassigned_vehicles': unassigned_vehicles,
        'unassigned_count': unassigned_count
    })

@login_required
def operator_applications_manage(request):
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('user_portal:user_dashboard')
            
    # Get all applications ordered by submission date
    all_applications = OperatorApplication.objects.all().order_by('-submitted_at')
    
    # Separate pending and processed applications
    pending_applications = all_applications.filter(status='PENDING')
    processed_applications = all_applications.exclude(status='PENDING')
    
    return render(request, 'operators/operator_applications_manage.html', {
        'pending_applications': pending_applications,
        'processed_applications': processed_applications,
        'title': 'Operator Applications'
    })

@login_required
def operator_application_review(request, application_id):
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('user_portal:user_dashboard')
    
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    try:
        application = get_object_or_404(OperatorApplication, id=application_id)
        
        if request.method == 'POST':
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            if status not in ['APPROVED', 'REJECTED']:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid status. Must be APPROVED or REJECTED.'
                    }, status=400)
                else:
                    messages.error(request, 'Invalid status. Must be APPROVED or REJECTED.')
                    return redirect('operator_applications_manage')
            
            # Use transaction to ensure all database operations succeed or fail together
            with transaction.atomic():
                # Update application status
                application.status = status
                application.notes = notes
                application.processed_at = timezone.now()
                application.processed_by = request.user
                application.save()
                
                # If approved, create or update an Operator record for this user
                if status == 'APPROVED':
                                        # Get user profile for address and other details
                    user_profile = application.user.userprofile
                    
                    # Generate a unique PD number
                    highest_pd = 0
                    try:
                        # Get all PD numbers, filtering out non-numeric ones
                        pd_numbers = Operator.objects.filter(
                            new_pd_number__iregex=r'^\d+$'
                        ).values_list('new_pd_number', flat=True)
                        
                        # Convert to integers for proper sorting
                        numeric_pds = [int(pd) for pd in pd_numbers if pd and pd.isdigit()]
                        
                        if numeric_pds:
                            highest_pd = max(numeric_pds)
                    except Exception as e:
                        # If there's any error, log it but continue with default numbering
                        print(f"Error determining highest PD number: {str(e)}")
                    
                    # Set the new PD number as next in sequence (XXX format)
                    new_pd_number = str(highest_pd + 1).zfill(3)
                    old_pd_number = ""  # Leave old PD number blank as requested
                    
                    # Create or update operator record with safe attribute access
                    operator, created = Operator.objects.get_or_create(
                        user=application.user,
                        defaults={
                            'first_name': application.user.first_name,
                            'last_name': application.user.last_name,
                            'address': user_profile.address,
                            'new_pd_number': new_pd_number,
                            'old_pd_number': old_pd_number,
                        }
                    )
                    
                    # Generate a new PO number if created
                    if created:
                        def generate_po_number():
                            # Generate a random 3-digit number
                            import random
                            while True:
                                po_number = f"PO-{random.randint(100, 999)}"
                                
                                # Check if it already exists
                                if not Operator.objects.filter(po_number=po_number).exists():
                                    return po_number
                        
                        operator.po_number = generate_po_number()
                        operator.save()
                        
                        # Update the user profile to mark them as an operator
                        user_profile.is_operator = True
                        user_profile.operator_since = timezone.now()
                        user_profile.save()
                        
                        
                        # Check if number_of_units was specified in the application
                        import json
                        units_created = 0
                        vehicles_created = []
                        
                        try:
                            if application.extra_data:
                                # Parse the extra_data JSON
                                extra_data = json.loads(application.extra_data)
                                number_of_units = extra_data.get('number_of_units', 1)
                                vehicle_type = extra_data.get('vehicle_type', 'Potpot')
                                
                                # Ensure it's an integer and within limits
                                number_of_units = int(number_of_units)
                                if number_of_units > 100:
                                    number_of_units = 100  # Limit to 100 units
                                
                                if number_of_units > 0:
                                    # Find the highest existing potpot number in the system
                                    highest_potpot = 0
                                    try:
                                        # Get all potpot numbers, filtering out non-numeric ones
                                        potpot_numbers = Vehicle.objects.filter(
                                            potpot_number__iregex=r'^\d+$'
                                        ).values_list('potpot_number', flat=True)
                                        
                                        # Convert to integers for proper sorting
                                        numeric_potpots = [int(pn) for pn in potpot_numbers if pn and pn.isdigit()]
                                        
                                        if numeric_potpots:
                                            highest_potpot = max(numeric_potpots)
                                    except Exception as e:
                                        # If there's any error, log it but continue with default numbering
                                        print(f"Error determining highest potpot number: {str(e)}")
                                    
                                    # Start numbering from the highest existing + 1
                                    start_number = highest_potpot + 1
                                    
                                    # Create the specified number of vehicle records
                                    for i in range(start_number, start_number + number_of_units):
                                        # Generate potpot number in XXX format (001, 002, etc.)
                                        potpot_number = str(i).zfill(3)
                                        
                                        # Create a vehicle record for this potpot
                                        vehicle = Vehicle(
                                            operator=operator,
                                            potpot_number=potpot_number,
                                            vehicle_type=vehicle_type,
                                            active=True
                                        )
                                        
                                        # Use the operator's new_pd_number as a base for the vehicle's pd number
                                        # Format: [operator_pd]-[potpot_number]
                                        if operator.new_pd_number:
                                            vehicle.new_pd_number = f"{operator.new_pd_number}-{potpot_number}"
                                        
                                        # Save the vehicle
                                        vehicle.save()
                                        vehicles_created.append(potpot_number)
                                        units_created += 1
                        except Exception as e:
                            # Log error but continue with the approval
                            print(f"Error creating potpot units: {str(e)}")
                        
                        # Create notification for the user
                        from .user_portal.models import UserNotification
                        UserNotification.objects.create(
                            user=application.user,
                            type='SYSTEM',
                            message=f'Your operator application has been approved. You now have operator privileges with {units_created} potpot vehicles created.'
                        )
                        
                        # Log the approval
                        ActivityLog.objects.create(
                            user=request.user,
                            action=f"Approved operator application for {application.user.username}",
                            details=f"Assigned PO Number: {operator.po_number}, Created {units_created} potpot vehicles",
                            category="operator"
                        )
                    else:
                        # Log the approval for existing operator
                        ActivityLog.objects.create(
                            user=request.user,
                            action=f"Approved operator application for existing operator {application.user.username}",
                            category="operator"
                        )
                
                # If rejected, just create notification and log
                else:
                    # Create notification for the user
                    from .user_portal.models import UserNotification
                    UserNotification.objects.create(
                        user=application.user,
                        type='SYSTEM',
                        message=f'Your operator application was not approved. Reason: {notes if notes else "No specific reason provided."}'
                    )
                    
                    # Log the rejection
                    ActivityLog.objects.create(
                        user=request.user,
                        action=f"Rejected operator application for {application.user.username}",
                        details=f"Reason: {notes if notes else 'No reason provided'}",
                        category="operator"
                    )
            
            # Return successful response outside transaction
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Application {status.lower()}.',
                    'status': status
                })
            else:
                messages.success(request, f'Application has been {status.lower()}.')
                return redirect('operator_applications_manage')
        
        # Handle GET request
        if is_ajax:
            try:
                # Return JSON data for AJAX GET request
                user_profile = application.user.userprofile
                
                # Prepare documents data with safe URL handling
                documents = {
                    'business_permit': application.business_permit.url if application.business_permit else None,
                    'police_clearance': application.police_clearance.url if application.police_clearance and application.police_clearance.name else None,
                    'barangay_certificate': application.barangay_certificate.url if application.barangay_certificate and application.barangay_certificate.name else None,
                    'cedula': application.cedula.url if application.cedula and application.cedula.name else None,
                    'cenro_tickets': application.cenro_tickets.url if application.cenro_tickets and application.cenro_tickets.name else None,
                    'mayors_permit': application.mayors_permit.url if application.mayors_permit and application.mayors_permit.name else None,
                    'other_documents': application.other_documents.url if application.other_documents and application.other_documents.name else None,
                }
                
                # Prepare user data
                user_data = {
                    'username': application.user.username,
                    'full_name': application.user.get_full_name(),
                    'email': application.user.email,
                    'phone_number': user_profile.phone_number,
                    'address': user_profile.address,
                }
                
                # Application data
                application_data = {
                    'id': application.id,
                    'company_name': getattr(application, 'company_name', ''),
                    'submitted_at': application.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'status': application.status,
                }
                
                # Get vehicle type and number of units from extra_data
                if application.extra_data:
                    try:
                        import json
                        extra_data = json.loads(application.extra_data)
                        application_data['vehicle_type'] = extra_data.get('vehicle_type', 'Potpot')
                        application_data['number_of_units'] = extra_data.get('number_of_units', 1)
                    except json.JSONDecodeError:
                        # If we can't parse the JSON, use defaults
                        application_data['vehicle_type'] = 'Potpot'
                        application_data['number_of_units'] = 1
                else:
                    # If no extra_data, use defaults
                    application_data['vehicle_type'] = 'Potpot'
                    application_data['number_of_units'] = 1
                
                return JsonResponse({
                    'success': True,
                    'user': user_data,
                    'application': application_data,
                    'documents': documents,
                })
            except Exception as e:
                # Log any errors that occur when preparing JSON data
                logger.error(f"Error preparing JSON data in operator_application_review GET: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'message': f'Error preparing application data: {str(e)}'
                }, status=500)
        else:
            # For regular GET requests, render the template
            return render(request, 'operators/operator_application_review.html', {
                'application': application
            })
        
    except Exception as e:
        # Log the error for debugging
        logger.error(f"Error in operator_application_review: {str(e)}")
        
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)
        else:
            messages.error(request, f'Error processing application: {str(e)}')
            return redirect('operator_applications_manage')

@login_required
def driver_export_excel(request):
    """View to export all drivers data to Excel file"""
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Drivers Data"
    
    # Add headers
    headers = ['Last Name', 'First Name', 'Middle Initial', 'Address', 'Old PD Number', 'New PD Number', 'License Number', 'Contact Number', 'Operator']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    # Get all drivers from database
    drivers = Driver.objects.all().select_related('operator').order_by('last_name', 'first_name')
    
    # Add data rows
    for row_num, driver in enumerate(drivers, 2):  # Start from row 2 (after header)
        row = [
            driver.last_name,
            driver.first_name,
            driver.middle_initial if driver.middle_initial else '',
            driver.address,
            driver.old_pd_number if driver.old_pd_number else '',
            driver.new_pd_number,
            driver.license_number if driver.license_number else '',
            driver.contact_number if driver.contact_number else '',
            f"{driver.operator.last_name}, {driver.operator.first_name}" if driver.operator else ''
        ]
        
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    column_widths = [20, 20, 15, 40, 15, 15, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=drivers_data.xlsx'
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Exported drivers data to Excel",
        category="export"
    )
    
    # Save the workbook to the response
    workbook.save(response)
    return response

@login_required
def driver_create(request):
    """View to create a new driver"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to create drivers.')
        return redirect('admin_driver_list')
        
    if request.method == 'POST':
        form = DriverForm(request.POST)
        # Exclude new_pd_number from validation since we'll auto-generate it
        if 'new_pd_number' in form.errors:
            del form.errors['new_pd_number']
            
        if form.is_valid() or ('new_pd_number' in form.errors and len(form.errors) == 1):
            driver = form.save(commit=False)
            
            # Generate a unique PD number in XXX format (3 digits)
            import random
            highest_pd = 0
            try:
                # Find the highest existing PD number in the system
                pd_numbers = Driver.objects.filter(
                    new_pd_number__regex=r'^\d+$'
                ).values_list('new_pd_number', flat=True)
                
                # Convert to integers for proper sorting
                numeric_pds = [int(pd) for pd in pd_numbers if pd and pd.isdigit()]
                
                if numeric_pds:
                    highest_pd = max(numeric_pds)
            except Exception as e:
                # If there's any error, log it but continue with default numbering
                print(f"Error determining highest PD number: {str(e)}")
            
            # Set the new PD number as next in sequence (XXX format)
            new_pd_number = str(highest_pd + 1).zfill(3)
            driver.new_pd_number = new_pd_number
            
            # Set expiration date to 1 year from now
            from django.utils import timezone
            driver.expiration_date = timezone.now().date() + timezone.timedelta(days=365)
                
            driver.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Created new driver: {driver.first_name} {driver.last_name} with PD number: {new_pd_number}",
                category="driver"
            )
            
            messages.success(request, f'Driver {driver.first_name} {driver.last_name} has been created successfully with PD number: {new_pd_number}.')
            return redirect('driver_detail', pk=driver.id)
    else:
        form = DriverForm()
        
    return render(request, 'drivers/driver_form.html', {
        'form': form,
        'title': 'Create New Driver',
        'is_create': True,
        'submit_text': 'Create Driver'
    })

@login_required
def driver_update(request, pk):
    """View to update an existing driver"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to update drivers.')
        return redirect('admin_driver_list')
        
    driver = get_object_or_404(Driver, pk=pk)
    
    if request.method == 'POST':
        form = DriverForm(request.POST, instance=driver)
        if form.is_valid():
            driver = form.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Updated driver: {driver.first_name} {driver.last_name}",
                category="driver"
            )
            
            messages.success(request, f'Driver {driver.first_name} {driver.last_name} has been updated successfully.')
            return redirect('admin_driver_list')
    else:
        form = DriverForm(instance=driver)
        
    return render(request, 'drivers/driver_form.html', {
        'form': form,
        'driver': driver,
        'title': 'Update Driver',
        'action': 'Update',
        'button_text': 'Save Changes'
    })

@login_required
def driver_disable(request, pk):
    """View to disable a driver"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to disable drivers.')
        return redirect('admin_driver_list')
        
    driver = get_object_or_404(Driver, pk=pk)
    
    if request.method == 'POST':
        driver_name = f"{driver.first_name} {driver.last_name}"
        
        try:
            # Check if there are any active vehicle assignments
            active_assignments = DriverVehicleAssignment.objects.filter(driver=driver, is_active=True).exists()
            if active_assignments:
                return JsonResponse({
                    'success': False, 
                    'error': 'Cannot disable driver with active vehicle assignments. Please end all vehicle assignments first.'
                })
            
            # Disable the driver instead of deleting
            driver.active = False
            driver.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Disabled driver: {driver_name}",
                category="driver"
            )
            
            messages.success(request, f'Driver {driver_name} has been disabled successfully.')
        except Exception as e:
            messages.error(request, f'Error disabling driver: {str(e)}')
            return JsonResponse({'success': False, 'error': str(e)})
        
        return JsonResponse({'success': True})
    
    # If it's not a POST request, return a 405 Method Not Allowed response
    return JsonResponse({
        'success': False, 
        'error': 'Method not allowed'
    }, status=405)

@login_required
def driver_enable(request, pk):
    """View to enable a previously disabled driver"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to enable drivers.')
        return redirect('admin_driver_list')
        
    driver = get_object_or_404(Driver, pk=pk)
    
    if request.method == 'POST':
        driver_name = f"{driver.first_name} {driver.last_name}"
        
        try:
            # Enable the driver
            driver.active = True
            driver.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action=f"Enabled driver: {driver_name}",
                category="driver"
            )
            
            messages.success(request, f'Driver {driver_name} has been enabled successfully.')
        except Exception as e:
            messages.error(request, f'Error enabling driver: {str(e)}')
            return JsonResponse({'success': False, 'error': str(e)})
        
        return JsonResponse({'success': True})
    
    # If it's not a POST request, return a 405 Method Not Allowed response
    return JsonResponse({
        'success': False, 
        'error': 'Method not allowed'
    }, status=405)

@login_required
def driver_import(request):
    """View to handle driver data import"""

@login_required
def driver_import(request):
    """View to handle driver data import"""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to import drivers.')
        return redirect('admin_driver_list')
    
    if request.method == 'POST':
        form = DriverImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Save the file temporarily and prepare for processing
            import_file = request.FILES['file']
            
            # Add debugging information
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"File name: {import_file.name}, Size: {import_file.size}")
            
            # Process the file using pandas
            import pandas as pd
            import tempfile
            import os
            
            # Save uploaded file to a temp file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            
            with open(temp_file.name, 'wb+') as destination:
                for chunk in import_file.chunks():
                    destination.write(chunk)
            
            try:
                # Read the Excel file
                df = pd.read_excel(temp_file.name)
                
                # Add debugging information
                logger.error(f"DataFrame columns: {df.columns.tolist()}")
                logger.error(f"DataFrame rows: {len(df)}")
                
                # Skip the preview & confirmation, directly process the file
                if len(df.columns) >= 6:
                    # Just assume the columns are in order: last name, first name, middle initial, address, old pd, new pd
                    col_last_name = df.columns[0]
                    col_first_name = df.columns[1]
                    col_middle_initial = df.columns[2] if len(df.columns) > 2 else None
                    col_address = df.columns[3] if len(df.columns) > 3 else None
                    col_old_pd = df.columns[4] if len(df.columns) > 4 else None
                    col_new_pd = df.columns[5] if len(df.columns) > 5 else None
                    
                    logger.error(f"Using columns by position: {col_last_name}, {col_first_name}, {col_middle_initial}, {col_address}, {col_old_pd}, {col_new_pd}")
                    
                    # Track import results
                    success_count = 0
                    update_count = 0
                    error_count = 0
                    error_messages = []
                    
                    # Process each row
                    for index, row in df.iterrows():
                        # Skip header rows if present
                        if index == 0 and ('POTPO' in str(row.get(col_last_name, '')) or 'LASTNAME' in str(row.get(col_last_name, ''))):
                            continue
                            
                        # Extract data
                        last_name = str(row.get(col_last_name, ''))
                        first_name = str(row.get(col_first_name, ''))
                        middle_initial = str(row.get(col_middle_initial, ''))
                        address = str(row.get(col_address, ''))
                        old_pd_number = str(row.get(col_old_pd, ''))
                        new_pd_number = str(row.get(col_new_pd, ''))
                        
                        # Basic validation
                        if not last_name or not first_name or not address or not new_pd_number:
                            error_count += 1
                            error_messages.append(f"Row {index+1}: Missing required fields")
                            continue
                        
                        try:
                            # Check if this is an update or new record
                            existing_driver = Driver.objects.filter(new_pd_number=new_pd_number).first()
                            
                            if existing_driver:
                                # Update existing driver
                                existing_driver.last_name = last_name
                                existing_driver.first_name = first_name
                                existing_driver.middle_initial = middle_initial
                                existing_driver.address = address
                                existing_driver.old_pd_number = old_pd_number
                                existing_driver.save()
                                update_count += 1
                            else:
                                # Create new driver
                                Driver.objects.create(
                                    last_name=last_name,
                                    first_name=first_name,
                                    middle_initial=middle_initial,
                                    address=address,
                                    old_pd_number=old_pd_number,
                                    new_pd_number=new_pd_number
                                )
                                success_count += 1
                        except Exception as e:
                            error_count += 1
                            error_messages.append(f"Row {index+1}: {str(e)}")
                            logger.error(f"Error importing driver {last_name}, {first_name}: {str(e)}")
                    
                    # Clean up the temp file
                    os.unlink(temp_file.name)
                    
                    # Log activity
                    ActivityLog.objects.create(
                        user=request.user,
                        action=f"Imported drivers: {success_count} new, {update_count} updated, {error_count} failed",
                        category="driver"
                    )
                    
                    # Show success message
                    if error_count > 0:
                        messages.warning(request, f'Imported {success_count} new drivers and updated {update_count} existing drivers. {error_count} records had errors.')
                    else:
                        messages.success(request, f'Successfully imported {success_count} new drivers and updated {update_count} existing drivers.')
                    
                    return redirect('admin_driver_list')
                else:
                    os.unlink(temp_file.name)  # Delete temp file
                    messages.error(request, "File doesn't have enough columns. Expected at least 6 columns.")
                    return redirect('driver_import')
                
            except Exception as e:
                os.unlink(temp_file.name)  # Delete temp file
                logger.error(f"Error processing file: {str(e)}")
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect('driver_import')
    else:
        form = DriverImportForm()
    
    return render(request, 'drivers/driver_import.html', {
        'form': form,
        'title': 'Import Drivers'
    })

@login_required
def operator_template_excel(request):
    """View to export empty operators template for data import"""
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Operators Template"
    
    # Add headers
    headers = ['Last Name', 'First Name', 'Middle Initial', 'Address', 'Old PD Number', 'New PD Number']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    # Add example data row
    example_data = ['Doe', 'John', 'A', '123 Main St, Anytown', 'PD-12345', 'PD-67890']
    for col_num, value in enumerate(example_data, 1):
        sheet.cell(row=2, column=col_num).value = value
        sheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(italic=True)
        sheet.cell(row=2, column=col_num).fill = openpyxl.styles.PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Add a note row
    note = "This is an example row. Please delete and replace with your actual data."
    sheet.cell(row=3, column=1).value = note
    sheet.cell(row=3, column=1).font = openpyxl.styles.Font(italic=True, color="888888")
    sheet.merge_cells('A3:F3')
    
    # Add 5 empty rows for the user to fill in
    for row_num in range(4, 9):
        for col_num in range(1, 7):
            sheet.cell(row=row_num, column=col_num).value = ""
    
    # Adjust column widths
    column_widths = [20, 20, 15, 40, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=operators_template.xlsx'
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Downloaded operators import template",
        category="user"
    )
    
    # Save the workbook to the response
    workbook.save(response)
    return response

@login_required
def driver_template_excel(request):
    """View to export empty drivers template for data import"""
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Drivers Template"
    
    # Add headers
    headers = ['Last Name', 'First Name', 'Middle Initial', 'Address', 'Old PD Number', 'New PD Number', 'License Number', 'Contact Number']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    # Add example data row
    example_data = ['Doe', 'John', 'A', '123 Main St, Anytown', 'PD-12345', 'PD-67890', 'DL-123456', '09123456789']
    for col_num, value in enumerate(example_data, 1):
        sheet.cell(row=2, column=col_num).value = value
        sheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(italic=True)
        sheet.cell(row=2, column=col_num).fill = openpyxl.styles.PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Add a note row
    note = "This is an example row. Please delete and replace with your actual data."
    sheet.cell(row=3, column=1).value = note
    sheet.cell(row=3, column=1).font = openpyxl.styles.Font(italic=True, color="888888")
    sheet.merge_cells('A3:H3')
    
    # Add 5 empty rows for the user to fill in
    for row_num in range(4, 9):
        for col_num in range(1, 9):
            sheet.cell(row=row_num, column=col_num).value = ""
    
    # Adjust column widths
    column_widths = [20, 20, 15, 40, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=drivers_template.xlsx'
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action=f"Downloaded drivers import template",
        category="user"
    )
    
    # Save the workbook to the response
    workbook.save(response)
    return response

# Add the missing function here
def check_potpot_number_unique(request):
    """
    AJAX endpoint to check if a potpot number is unique.
    Returns JSON with is_unique: true/false
    """
    number = request.GET.get('number', '')
    
    # Check if the number is valid (3 digits)
    if not number or not number.isdigit() or len(number) != 3:
        return JsonResponse({'is_unique': False, 'error': 'Invalid number format'})
    
    # Check if this potpot number exists in the database
    is_unique = not Vehicle.objects.filter(potpot_number=number).exists()
    
    return JsonResponse({'is_unique': is_unique})

@login_required
def get_adjudicator_notes(request, violation_id):
    """View to fetch adjudicator notes for a violation."""
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Check if the violation has been adjudicated
        if not violation.adjudicated_by or not violation.adjudication_date:
            return JsonResponse({
                'success': True,
                'notes': None
            })
        
        # Format adjudication date
        adjudication_date = violation.adjudication_date.strftime('%B %d, %Y %I:%M %p')
        
        # Get adjudicator name
        adjudicator_name = violation.adjudicated_by.get_full_name() or violation.adjudicated_by.username
        
        # Return adjudicator notes
        return JsonResponse({
            'success': True,
            'notes': violation.adjudication_remarks,
            'adjudicator_name': adjudicator_name,
            'adjudication_date': adjudication_date
        })
    except Exception as e:
        # Return error
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
def adjudication_dashboard(request):
    """
    Dashboard view showing adjudication statistics and rejected adjudications
    """
    # Count total violations
    total_violations = Violation.objects.count()
    
    # Count total adjudications
    total_adjudications = Violation.objects.filter(
        adjudicated_by__isnull=False,
        adjudication_date__isnull=False
    ).count()
    
    # Count approved adjudications
    approved_adjudications = Violation.objects.filter(
        status='APPROVED'
    ).count()
    
    # Count rejected adjudications
    rejected_adjudications = Violation.objects.filter(
        status='REJECTED'
    ).count()
    
    # Pending adjudications (awaiting approval)
    pending_adjudications = Violation.objects.filter(
        status='ADJUDICATED'
    ).count()
    
    # Recent adjudications (last 10)
    recent_adjudications = Violation.objects.filter(
        adjudicated_by__isnull=False
    ).order_by('-adjudication_date')[:10]
    
    # Get a summary of adjudications by adjudicator
    adjudicator_summary = Violation.objects.filter(
        adjudicated_by__isnull=False
    ).values('adjudicated_by__username', 'adjudicated_by__first_name', 'adjudicated_by__last_name'
    ).annotate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='APPROVED')),
        rejected=Count('id', filter=Q(status='REJECTED')),
        pending=Count('id', filter=Q(status='ADJUDICATED'))
    ).order_by('-total')[:5]  # Top 5 adjudicators
    
    context = {
        'total_violations': total_violations,
        'total_adjudications': total_adjudications,
        'approved_adjudications': approved_adjudications,
        'rejected_adjudications': rejected_adjudications,
        'pending_adjudications': pending_adjudications,
        'recent_adjudications': recent_adjudications,
        'adjudicator_summary': adjudicator_summary,
    }
    
    return render(request, 'violations/adjudication_dashboard.html', context)

@login_required
def rejected_adjudications(request):
    """
    View to display all rejected adjudications with options to re-adjudicate
    """
    # Get all rejected adjudications
    rejected_violations = Violation.objects.filter(
        status='REJECTED'
    ).select_related('violator', 'adjudicated_by', 'approved_by').order_by('-adjudication_date')
    
    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        rejected_violations = rejected_violations.filter(
            Q(violator__first_name__icontains=search_query) |
            Q(violator__last_name__icontains=search_query) |
            Q(violator__license_number__icontains=search_query) |
            Q(violation_type__icontains=search_query) |
            Q(id__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(rejected_violations, 15)
    page = request.GET.get('page', 1)
    try:
        violations = paginator.page(page)
    except PageNotAnInteger:
        violations = paginator.page(1)
    except EmptyPage:
        violations = paginator.page(paginator.num_pages)

    context = {
        'violations': violations,
        'page_obj': violations,
        'search_query': search_query,
        'total_count': paginator.count
    }
    
    return render(request, 'violations/rejected_adjudications.html', context)

@login_required
def violation_payment_info(request, violation_id):
    """
    Return violation details as JSON for payment processing.
    Works for both NCAP and regular violations.
    """
    try:
        violation = get_object_or_404(Violation, id=violation_id)
        
        # Determine if this is an NCAP violation (has images)
        is_ncap = bool(violation.image or violation.driver_photo or 
                      violation.vehicle_photo or violation.secondary_photo)
        
        # Build the response data dictionary with all necessary information
        response_data = {
            'id': violation.id,
            'violation_date': violation.violation_date.isoformat(),
            'formatted_date': violation.violation_date.strftime("%b %d, %Y"),
            'violation_type': violation.violation_type,
            'violation_types': violation.get_violation_types(),
            'status': violation.status,
            'status_display': violation.get_status_display(),
            'status_class': f"bg-{get_status_class(violation.status)}",
            'fine_amount': str(violation.fine_amount),
            'is_ncap': is_ncap,
            'is_tdz_violation': violation.is_tdz_violation,
            'plate_number': violation.plate_number,
            'vehicle_type': violation.vehicle_type,
            'location': violation.location,
        }
        
        # Add due date if available
        if violation.payment_due_date:
            response_data['payment_due_date'] = violation.payment_due_date.isoformat()
            response_data['formatted_due_date'] = violation.payment_due_date.strftime("%b %d, %Y")
        else:
            response_data['payment_due_date'] = None
            response_data['formatted_due_date'] = "Not set"
        
        # Add violator information if available
        if violation.violator:
            response_data['violator_name'] = f"{violation.violator.first_name} {violation.violator.last_name}".strip()
            response_data['license_number'] = violation.violator.license_number
            response_data['phone_number'] = violation.violator.phone_number
            response_data['address'] = violation.violator.address
        
        # Add driver info for NCAP violations
        if is_ncap:
            if violation.driver_name:
                response_data['driver_name'] = violation.driver_name
            if violation.novr_number:
                response_data['novr_number'] = violation.novr_number
            if violation.pin_number:
                response_data['pin_number'] = violation.pin_number
            if violation.operator_name:
                response_data['operator_name'] = violation.operator_name
            if violation.potpot_number:
                response_data['potpot_number'] = violation.potpot_number
        
        return JsonResponse(response_data)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': str(e)
        }, status=500)

def get_status_class(status):
    """
    Convert a violation status to a Bootstrap color class
    """
    status_classes = {
        'PENDING': 'warning',
        'ADJUDICATED': 'info',
        'APPROVED': 'success',
        'REJECTED': 'danger',
        'PAID': 'primary',
        'SETTLED': 'primary',
        'OVERDUE': 'danger',
    }
    return status_classes.get(status, 'secondary')

def index(request):
    """Home page view"""
    return render(request, 'landing/home.html')

@login_required
def export_payment_records_excel(request):
    """Export payment records to Excel format"""
    try:
        # Query parameters (same as payment_records view)
        search_query = request.GET.get('search', '')
        
        # Get paid violations
        violations = Violation.objects.filter(status='PAID').order_by('-receipt_date')
        
        # Apply search filter if provided
        if search_query:
            violations = violations.filter(
                Q(violator__first_name__icontains=search_query) |
                Q(violator__last_name__icontains=search_query) |
                Q(violator__license_number__icontains=search_query) |
                Q(receipt_number__icontains=search_query) |
                Q(violation_type__name__icontains=search_query)
            )
        
        # Create workbook and active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Records"
        
        # Add logo
        logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo.png')
        if os.path.exists(logo_path):
            logo = XLImage(logo_path)
            logo.width = 80
            logo.height = 80
            ws.add_image(logo, 'A1')
        
        # Add ibayaw logo
        ibayaw_logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo_ibayaw.png')
        if os.path.exists(ibayaw_logo_path):
            ibayaw_logo = XLImage(ibayaw_logo_path)
            ibayaw_logo.width = 80
            ibayaw_logo.height = 80
            ws.add_image(ibayaw_logo, 'H1')
        
        # Add headers
        ws.merge_cells('B1:G1')
        ws.cell(row=1, column=2).value = "REPUBLIC OF THE PHILIPPINES"
        ws.cell(row=1, column=2).font = Font(size=12, bold=True)
        ws.cell(row=1, column=2).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('B2:G2')
        ws.cell(row=2, column=2).value = "CITY TRANSPORT AND TRAFFIC MANAGEMENT OFFICE"
        ws.cell(row=2, column=2).font = Font(size=14, bold=True)
        ws.cell(row=2, column=2).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('B3:G3')
        ws.cell(row=3, column=2).value = "Traffic Violation Management System"
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('B4:G4')
        ws.cell(row=4, column=2).value = "City Government of Bayawan City"
        ws.cell(row=4, column=2).alignment = Alignment(horizontal='center')
        
        # Add report title and date
        ws.merge_cells('A6:H6')
        ws.cell(row=6, column=1).value = "PAYMENT RECORDS REPORT"
        ws.cell(row=6, column=1).font = Font(size=14, bold=True)
        ws.cell(row=6, column=1).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A7:H7')
        ws.cell(row=7, column=1).value = f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws.cell(row=7, column=1).alignment = Alignment(horizontal='center')
        
        # Add column headers
        headers = ['Receipt #', 'Date', 'Violation #', 'Violator', 'License Number', 'Violation Type', 'Amount', 'Status']
        header_row = 9
        
        # Style for headers
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        row_num = header_row + 1
        for violation in violations:
            ws.cell(row=row_num, column=1).value = violation.receipt_number
            ws.cell(row=row_num, column=2).value = violation.receipt_date.strftime('%b %d, %Y') if violation.receipt_date else ''
            ws.cell(row=row_num, column=3).value = violation.id
            ws.cell(row=row_num, column=4).value = violation.violator.get_full_name()
            ws.cell(row=row_num, column=5).value = violation.violator.license_number if hasattr(violation.violator, 'license_number') else 'N/A'
            ws.cell(row=row_num, column=6).value = violation.violation_type.name if hasattr(violation.violation_type, 'name') else str(violation.violation_type)
            ws.cell(row=row_num, column=7).value = violation.fine_amount
            ws.cell(row=row_num, column=7).number_format = '₱#,##0.00'
            ws.cell(row=row_num, column=8).value = 'Paid'
            
            # Add border to cells
            for col_num in range(1, 9):
                ws.cell(row=row_num, column=col_num).border = border
            
            row_num += 1
        
        # Adjust column widths
        column_widths = [15, 15, 12, 25, 20, 25, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=payment_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error generating Excel export: {str(e)}", status=500)

@login_required
def export_payment_records_pdf(request):
    """Export payment records to PDF format"""
    try:
        # Query parameters (same as payment_records view)
        search_query = request.GET.get('search', '')
        
        # Get paid violations
        violations = Violation.objects.filter(status='PAID').order_by('-receipt_date')
        
        # Apply search filter if provided
        if search_query:
            violations = violations.filter(
                Q(violator__first_name__icontains=search_query) |
                Q(violator__last_name__icontains=search_query) |
                Q(violator__license_number__icontains=search_query) |
                Q(receipt_number__icontains=search_query) |
                Q(violation_type__name__icontains=search_query)
            )
        
        # Get absolute paths for logo images
        logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo.png')
        ibayaw_logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'logo_ibayaw.png')
        
        # Convert to data URLs if files exist
        logo_data = None
        ibayaw_logo_data = None
        
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as img_file:
                logo_data = base64.b64encode(img_file.read()).decode('utf-8')
        
        if os.path.exists(ibayaw_logo_path):
            with open(ibayaw_logo_path, 'rb') as img_file:
                ibayaw_logo_data = base64.b64encode(img_file.read()).decode('utf-8')
        
        # Prepare context
        context = {
            'violations': violations,
            'total_count': violations.count(),
            'generated_date': now(),
            'search_query': search_query,
            'STATIC_URL': settings.STATIC_URL,
            'logo_data': logo_data,
            'ibayaw_logo_data': ibayaw_logo_data,
            'MEDIA_URL': settings.MEDIA_URL,
            'BASE_DIR': settings.BASE_DIR,
        }
        
        # Render template to string
        template = get_template('payments/payment_records_pdf.html')
        html = template.render(context)
        
        # Create PDF with proper encoding
        result = io.BytesIO()
        pdf = pisa.pisaDocument(
            io.BytesIO(html.encode("UTF-8")), 
            result,
            encoding='UTF-8',
            link_callback=fetch_resources
        )
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = f'payment_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        return HttpResponse("Error generating PDF", status=500)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error generating PDF export: {str(e)}", status=500)
        
def fetch_resources(uri, rel):
    """
    Callback to allow pisa/reportlab to retrieve Images,Stylesheets, etc.
    """
    # Convert URI to absolute path
    if uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
    elif uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    else:
        path = uri  # For absolute paths
    
    # Make sure path exists
    if not os.path.isfile(path):
        path = os.path.join(settings.BASE_DIR, uri)
    
    return path